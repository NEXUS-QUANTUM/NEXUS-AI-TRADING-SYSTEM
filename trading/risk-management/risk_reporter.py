"""
NEXUS AI TRADING SYSTEM - Risk Reporter Module
Copyright © 2026 NEXUS QUANTUM LTD
CEO: Dr X... - Majority Shareholder

Module: trading/risk-management/risk_reporter.py
Description: Comprehensive risk reporting with full API integration
"""

import asyncio
import json
import logging
import os
import tempfile
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, field, asdict
from enum import Enum
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns

from pydantic import BaseModel, Field, validator
from fastapi import HTTPException, status, Response
from fastapi.responses import FileResponse, StreamingResponse

# NEXUS Internal Imports
from shared.configs.risk_config import RiskConfig
from shared.constants.trading_constants import (
    RISK_LEVELS,
    TIME_FRAMES,
    REPORT_TYPES
)
from shared.types.risk import (
    RiskMetrics,
    PortfolioRiskMetrics,
    RiskReport,
    RiskReportConfig
)
from shared.utilities.retry import retry_async
from shared.utilities.logger import get_logger
from shared.utilities.cache_utils import cached

# Database imports
from backend.database.models import (
    Position,
    Order,
    Trade,
    PortfolioSnapshot,
    RiskAlert,
    PerformanceMetric
)
from backend.database.repositories.position_repository import PositionRepository
from backend.database.repositories.trade_repository import TradeRepository
from backend.database.repositories.portfolio_repository import PortfolioRepository
from backend.database.repositories.risk_repository import RiskRepository

# Risk management imports
from trading.risk_management.portfolio_risk import PortfolioRiskManager
from trading.risk_management.risk_limits import RiskLimitsManager
from trading.risk_management.risk_monitor import RiskMonitor

logger = get_logger(__name__)


# =============================================================================
# ENUMS & CONSTANTS
# =============================================================================

class ReportFormat(str, Enum):
    """Supported report formats"""
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class ReportType(str, Enum):
    """Types of risk reports"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    CUSTOM = "custom"
    REAL_TIME = "real_time"
    SUMMARY = "summary"
    DETAILED = "detailed"
    EXECUTIVE = "executive"
    COMPLIANCE = "compliance"
    PERFORMANCE = "performance"
    POSITION = "position"
    LIMIT = "limit"
    STRESS_TEST = "stress_test"
    VaR = "var"
    DRAWDOWN = "drawdown"
    CORRELATION = "correlation"


class ReportFrequency(str, Enum):
    """Frequency of report generation"""
    MANUAL = "manual"
    SCHEDULED = "scheduled"
    ON_DEMAND = "on_demand"
    EVENT_DRIVEN = "event_driven"
    CONTINUOUS = "continuous"


class ChartType(str, Enum):
    """Types of charts"""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    AREA = "area"
    CANDLESTICK = "candlestick"
    HEATMAP = "heatmap"
    HISTOGRAM = "histogram"
    BOXPLOT = "boxplot"
    SCATTER = "scatter"
    DRAWDOWN = "drawdown"
    EQUITY = "equity"
    RISK = "risk"


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class ReportRequest(BaseModel):
    """Request model for generating a report"""
    report_type: ReportType
    portfolio_id: str
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    format: ReportFormat = ReportFormat.PDF
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_detailed: bool = True
    include_recommendations: bool = True
    chart_types: List[ChartType] = []
    metrics: List[str] = []
    sections: List[str] = []
    language: str = "en"
    timezone: str = "UTC"
    email_report: bool = False
    email_recipients: List[str] = []


class ReportResponse(BaseModel):
    """Response model for report generation"""
    report_id: str
    report_type: ReportType
    portfolio_id: str
    generated_at: datetime
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    format: ReportFormat
    file_size: int
    file_url: Optional[str] = None
    summary: Dict[str, Any]
    metrics: Dict[str, Any]
    alerts_count: int
    breaches_count: int
    recommendations_count: int
    charts_generated: int


class ReportSchedule(BaseModel):
    """Configuration for scheduled reports"""
    report_type: ReportType
    portfolio_id: str
    frequency: ReportFrequency
    format: ReportFormat = ReportFormat.PDF
    time: str = "00:00"  # HH:MM format
    day_of_week: Optional[int] = None  # 0=Monday, 6=Sunday
    day_of_month: Optional[int] = None  # 1-31
    recipients: List[str] = []
    enabled: bool = True
    last_generated: Optional[datetime] = None
    next_generation: Optional[datetime] = None


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ReportData:
    """Data container for reports"""
    portfolio_id: str
    report_type: ReportType
    start_date: datetime
    end_date: datetime
    positions: List[Any] = field(default_factory=list)
    trades: List[Any] = field(default_factory=list)
    snapshots: List[Any] = field(default_factory=list)
    alerts: List[Any] = field(default_factory=list)
    metrics: Dict[str, Any] = field(default_factory=dict)
    performance: Dict[str, Any] = field(default_factory=dict)
    risk_metrics: Dict[str, Any] = field(default_factory=dict)
    limit_status: Dict[str, Any] = field(default_factory=dict)
    recommendations: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class ChartData:
    """Data container for charts"""
    title: str
    chart_type: ChartType
    data: Dict[str, Any]
    x_label: Optional[str] = None
    y_label: Optional[str] = None
    legend: Optional[List[str]] = None
    colors: Optional[List[str]] = None


# =============================================================================
# RISK REPORTER
# =============================================================================

class RiskReporter:
    """
    Comprehensive Risk Reporter with full API integration.
    
    Generates various risk reports including:
    - Daily/Weekly/Monthly risk reports
    - Performance reports
    - Compliance reports
    - Executive summaries
    - Position reports
    - Limit reports
    - Stress test reports
    - VaR reports
    - Drawdown reports
    - Correlation reports
    
    Features:
    - Multiple output formats (PDF, HTML, JSON, CSV, Excel, Markdown)
    - Interactive charts
    - Scheduled reports
    - Email delivery
    - Customizable templates
    - Multi-language support
    """

    def __init__(
        self,
        config: Optional[RiskConfig] = None,
        position_repo: Optional[PositionRepository] = None,
        trade_repo: Optional[TradeRepository] = None,
        portfolio_repo: Optional[PortfolioRepository] = None,
        risk_repo: Optional[RiskRepository] = None,
        portfolio_risk: Optional[PortfolioRiskManager] = None,
        risk_limits: Optional[RiskLimitsManager] = None,
        risk_monitor: Optional[RiskMonitor] = None
    ):
        """
        Initialize RiskReporter.
        
        Args:
            config: Risk configuration
            position_repo: Position repository
            trade_repo: Trade repository
            portfolio_repo: Portfolio repository
            risk_repo: Risk repository
            portfolio_risk: Portfolio risk manager
            risk_limits: Risk limits manager
            risk_monitor: Risk monitor
        """
        self.config = config or RiskConfig()
        self.position_repo = position_repo or PositionRepository()
        self.trade_repo = trade_repo or TradeRepository()
        self.portfolio_repo = portfolio_repo or PortfolioRepository()
        self.risk_repo = risk_repo or RiskRepository()
        self.portfolio_risk = portfolio_risk or PortfolioRiskManager()
        self.risk_limits = risk_limits or RiskLimitsManager()
        self.risk_monitor = risk_monitor or RiskMonitor()
        
        # Report schedules
        self._schedules: Dict[str, ReportSchedule] = {}
        
        # Report cache
        self._report_cache: Dict[str, Dict[str, Any]] = {}
        
        # Chart defaults
        self._chart_defaults = {
            'figsize': (12, 8),
            'dpi': 100,
            'style': 'seaborn-v0_8-darkgrid',
            'font_size': 12,
            'title_size': 16
        }
        
        # Set style
        plt.style.use(self._chart_defaults['style'])
        sns.set_palette("husl")
        
        logger.info("RiskReporter initialized")

    # =========================================================================
    # Report Generation
    # =========================================================================

    @retry_async(max_attempts=3, delay=1.0)
    async def generate_report(
        self,
        request: ReportRequest
    ) -> ReportResponse:
        """
        Generate a risk report.
        
        Args:
            request: Report request
            
        Returns:
            ReportResponse: Generated report
        """
        try:
            # Validate request
            await self._validate_request(request)
            
            # Set date range
            start_date, end_date = self._get_date_range(request)
            
            # Collect report data
            report_data = await self._collect_report_data(
                request.portfolio_id,
                request.report_type,
                start_date,
                end_date
            )
            
            # Generate report
            report_id = f"report_{int(time.time() * 1000)}"
            
            # Generate content based on format
            content, file_size = await self._generate_content(
                report_data,
                request.format,
                request
            )
            
            # Create response
            response = ReportResponse(
                report_id=report_id,
                report_type=request.report_type,
                portfolio_id=request.portfolio_id,
                generated_at=datetime.utcnow(),
                start_date=start_date,
                end_date=end_date,
                format=request.format,
                file_size=file_size,
                summary=report_data.metrics.get('summary', {}),
                metrics=report_data.metrics,
                alerts_count=len(report_data.alerts),
                breaches_count=len(report_data.limit_status.get('breached_limits', [])),
                recommendations_count=len(report_data.recommendations),
                charts_generated=len(request.chart_types) if request.include_charts else 0
            )
            
            # Cache report
            self._report_cache[report_id] = {
                'response': response,
                'content': content,
                'data': report_data
            }
            
            # Send email if requested
            if request.email_report and request.email_recipients:
                await self._send_report_email(
                    response,
                    content,
                    request.email_recipients
                )
            
            logger.info(f"Report generated: {report_id}")
            return response
            
        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Report generation failed: {str(e)}"
            )

    async def _validate_request(self, request: ReportRequest) -> None:
        """Validate report request"""
        # Validate portfolio exists
        portfolio = await self.portfolio_repo.get_by_id(request.portfolio_id)
        if not portfolio:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Portfolio {request.portfolio_id} not found"
            )
        
        # Validate date range
        if request.start_date and request.end_date:
            if request.start_date > request.end_date:
                raise ValueError("Start date must be before end date")

    def _get_date_range(
        self,
        request: ReportRequest
    ) -> Tuple[datetime, datetime]:
        """Get date range for report"""
        end_date = request.end_date or datetime.utcnow()
        
        if request.start_date:
            start_date = request.start_date
        else:
            # Default based on report type
            if request.report_type == ReportType.DAILY:
                start_date = end_date - timedelta(days=1)
            elif request.report_type == ReportType.WEEKLY:
                start_date = end_date - timedelta(days=7)
            elif request.report_type == ReportType.MONTHLY:
                start_date = end_date - timedelta(days=30)
            elif request.report_type == ReportType.QUARTERLY:
                start_date = end_date - timedelta(days=90)
            elif request.report_type == ReportType.ANNUAL:
                start_date = end_date - timedelta(days=365)
            else:
                start_date = end_date - timedelta(days=30)
        
        return start_date, end_date

    async def _collect_report_data(
        self,
        portfolio_id: str,
        report_type: ReportType,
        start_date: datetime,
        end_date: datetime
    ) -> ReportData:
        """
        Collect data for report.
        
        Args:
            portfolio_id: Portfolio ID
            report_type: Type of report
            start_date: Start date
            end_date: End date
            
        Returns:
            ReportData: Collected data
        """
        # Get positions
        positions = await self.position_repo.get_by_portfolio_id(portfolio_id)
        
        # Get trades
        trades = await self.trade_repo.get_by_portfolio_id(
            portfolio_id,
            start_date=start_date,
            end_date=end_date
        )
        
        # Get snapshots
        snapshots = await self.portfolio_repo.get_snapshots(
            portfolio_id,
            start_date=start_date,
            end_date=end_date
        )
        
        # Get alerts
        alerts = await self.risk_repo.get_alerts(
            portfolio_id,
            start_date=start_date,
            end_date=end_date
        )
        
        # Get risk metrics
        metrics = await self._calculate_metrics(portfolio_id, trades, snapshots)
        
        # Get performance metrics
        performance = await self._calculate_performance_metrics(trades, snapshots)
        
        # Get limit status
        limit_status = await self.risk_limits.get_all_limits_status()
        
        # Generate recommendations
        recommendations = await self._generate_recommendations(
            portfolio_id,
            metrics,
            limit_status
        )
        
        return ReportData(
            portfolio_id=portfolio_id,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            positions=positions,
            trades=trades,
            snapshots=snapshots,
            alerts=alerts,
            metrics=metrics,
            performance=performance,
            risk_metrics=metrics.get('risk', {}),
            limit_status=limit_status.dict() if limit_status else {},
            recommendations=recommendations
        )

    async def _calculate_metrics(
        self,
        portfolio_id: str,
        trades: List[Any],
        snapshots: List[Any]
    ) -> Dict[str, Any]:
        """Calculate risk metrics"""
        metrics = {
            'summary': {},
            'risk': {},
            'performance': {},
            'position': {}
        }
        
        try:
            # Get portfolio risk metrics
            portfolio_risk = await self.portfolio_risk.analyze_portfolio_risk(
                portfolio_id
            )
            if portfolio_risk:
                metrics['risk'] = portfolio_risk.dict()
            
            # Calculate summary metrics
            if trades:
                total_pnl = sum(float(t.pnl) for t in trades if hasattr(t, 'pnl'))
                winning_trades = [t for t in trades if float(t.pnl) > 0]
                losing_trades = [t for t in trades if float(t.pnl) < 0]
                
                metrics['summary'] = {
                    'total_trades': len(trades),
                    'winning_trades': len(winning_trades),
                    'losing_trades': len(losing_trades),
                    'win_rate': len(winning_trades) / len(trades) if trades else 0,
                    'total_pnl': total_pnl,
                    'avg_pnl': total_pnl / len(trades) if trades else 0,
                    'max_win': max([float(t.pnl) for t in trades]) if trades else 0,
                    'max_loss': min([float(t.pnl) for t in trades]) if trades else 0,
                    'profit_factor': sum(float(t.pnl) for t in trades if float(t.pnl) > 0) / abs(sum(float(t.pnl) for t in trades if float(t.pnl) < 0)) if losing_trades else 0
                }
            
            # Position metrics
            positions = await self.position_repo.get_by_portfolio_id(portfolio_id)
            if positions:
                metrics['position'] = {
                    'total_positions': len(positions),
                    'total_exposure': sum(float(p.size) * float(p.entry_price) for p in positions),
                    'avg_position_size': sum(float(p.size) for p in positions) / len(positions) if positions else 0
                }
            
        except Exception as e:
            logger.error(f"Error calculating metrics: {e}")
        
        return metrics

    async def _calculate_performance_metrics(
        self,
        trades: List[Any],
        snapshots: List[Any]
    ) -> Dict[str, Any]:
        """Calculate performance metrics"""
        metrics = {}
        
        try:
            if snapshots and len(snapshots) > 1:
                # Calculate returns
                values = [float(s.total_value) for s in snapshots]
                returns = [(values[i] - values[i-1]) / values[i-1] 
                          for i in range(1, len(values))]
                
                if returns:
                    metrics['sharpe_ratio'] = np.mean(returns) / np.std(returns) if np.std(returns) > 0 else 0
                    metrics['max_drawdown'] = self._calculate_max_drawdown(values)
                    metrics['volatility'] = np.std(returns)
                    metrics['avg_return'] = np.mean(returns)
                    metrics['cumulative_return'] = (values[-1] - values[0]) / values[0] if values[0] > 0 else 0
                    
                    # Sortino ratio (downside risk)
                    downside_returns = [r for r in returns if r < 0]
                    if downside_returns:
                        downside_std = np.std(downside_returns)
                        metrics['sortino_ratio'] = np.mean(returns) / downside_std if downside_std > 0 else 0
                    
                    # Calmar ratio
                    max_dd = metrics.get('max_drawdown', 0)
                    if max_dd > 0:
                        annual_return = metrics['avg_return'] * 252  # Assuming daily returns
                        metrics['calmar_ratio'] = annual_return / max_dd
            
            # Win/loss metrics
            if trades:
                wins = [float(t.pnl) for t in trades if float(t.pnl) > 0]
                losses = [abs(float(t.pnl)) for t in trades if float(t.pnl) < 0]
                
                if wins and losses:
                    metrics['avg_win'] = np.mean(wins)
                    metrics['avg_loss'] = np.mean(losses)
                    metrics['profit_factor'] = sum(wins) / sum(losses) if sum(losses) > 0 else 0
                    
                    # Risk-reward ratio
                    metrics['risk_reward_ratio'] = np.mean(wins) / np.mean(losses) if np.mean(losses) > 0 else 0
                    
                metrics['win_rate'] = len(wins) / len(trades) if trades else 0
                
        except Exception as e:
            logger.error(f"Error calculating performance metrics: {e}")
        
        return metrics

    def _calculate_max_drawdown(self, values: List[float]) -> float:
        """Calculate maximum drawdown"""
        if not values:
            return 0.0
        
        peak = values[0]
        max_drawdown = 0.0
        
        for value in values:
            if value > peak:
                peak = value
            drawdown = (peak - value) / peak if peak > 0 else 0
            if drawdown > max_drawdown:
                max_drawdown = drawdown
        
        return max_drawdown

    async def _generate_recommendations(
        self,
        portfolio_id: str,
        metrics: Dict[str, Any],
        limit_status: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Generate risk recommendations"""
        recommendations = []
        
        # Check drawdown
        drawdown = metrics.get('risk', {}).get('drawdown', 0)
        if drawdown > 0.10:
            recommendations.append({
                'type': 'risk',
                'severity': 'high',
                'action': 'reduce_exposure',
                'description': f'Portfolio drawdown is {drawdown*100:.1f}%. Consider reducing risk exposure.',
                'expected_impact': 'Limits further drawdown'
            })
        
        # Check concentration
        concentration = metrics.get('risk', {}).get('concentration', 0)
        if concentration > 0.40:
            recommendations.append({
                'type': 'diversification',
                'severity': 'high',
                'action': 'diversify',
                'description': f'Portfolio concentration is {concentration*100:.1f}%. Consider adding uncorrelated assets.',
                'expected_impact': 'Improves risk-adjusted returns'
            })
        
        # Check leverage
        leverage = metrics.get('risk', {}).get('leverage', 0)
        if leverage > 1.5:
            recommendations.append({
                'type': 'risk',
                'severity': 'medium',
                'action': 'reduce_leverage',
                'description': f'Leverage is {leverage:.1f}x. Consider reducing leverage.',
                'expected_impact': 'Reduces liquidation risk'
            })
        
        # Check VaR
        var = metrics.get('risk', {}).get('var_95', 0)
        if var > 0.05:
            recommendations.append({
                'type': 'risk',
                'severity': 'medium',
                'action': 'reduce_position_size',
                'description': f'VaR is {var*100:.1f}%. Consider reducing position sizes.',
                'expected_impact': 'Reduces potential losses'
            })
        
        # Check win rate
        win_rate = metrics.get('performance', {}).get('win_rate', 0)
        if win_rate < 0.40:
            recommendations.append({
                'type': 'strategy',
                'severity': 'medium',
                'action': 'review_strategy',
                'description': f'Win rate is {win_rate*100:.1f}%. Consider reviewing trading strategy.',
                'expected_impact': 'Improves trade quality'
            })
        
        # Check limit breaches
        breaches = limit_status.get('breached_limits', [])
        if breaches:
            recommendations.append({
                'type': 'compliance',
                'severity': 'critical',
                'action': 'address_breaches',
                'description': f'{len(breaches)} risk limits are breached. Immediate action required.',
                'expected_impact': 'Restores compliance'
            })
        
        return recommendations

    # =========================================================================
    # Content Generation
    # =========================================================================

    async def _generate_content(
        self,
        report_data: ReportData,
        format: ReportFormat,
        request: ReportRequest
    ) -> Tuple[Any, int]:
        """
        Generate report content in specified format.
        
        Args:
            report_data: Report data
            format: Output format
            request: Original request
            
        Returns:
            Tuple[Any, int]: Content and file size
        """
        if format == ReportFormat.PDF:
            content = await self._generate_pdf(report_data, request)
        elif format == ReportFormat.HTML:
            content = await self._generate_html(report_data, request)
        elif format == ReportFormat.JSON:
            content = await self._generate_json(report_data, request)
        elif format == ReportFormat.CSV:
            content = await self._generate_csv(report_data, request)
        elif format == ReportFormat.EXCEL:
            content = await self._generate_excel(report_data, request)
        elif format == ReportFormat.MARKDOWN:
            content = await self._generate_markdown(report_data, request)
        else:
            content = await self._generate_json(report_data, request)
        
        # Calculate file size
        file_size = len(content) if isinstance(content, (bytes, str)) else 0
        
        return content, file_size

    # =========================================================================
    # PDF Generation
    # =========================================================================

    async def _generate_pdf(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate PDF report"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image, KeepTogether
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # Create PDF in memory
            buffer = BytesIO()
            
            # Create document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            # Story for document
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            heading_style = styles['Heading1']
            subheading_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Title
            story.append(Paragraph(f"Risk Report - {report_data.report_type.value.upper()}", title_style))
            story.append(Spacer(1, 0.25*inch))
            
            # Portfolio info
            story.append(Paragraph(f"Portfolio: {report_data.portfolio_id}", normal_style))
            story.append(Paragraph(f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}", normal_style))
            story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", normal_style))
            story.append(Spacer(1, 0.5*inch))
            
            # Summary section
            story.append(Paragraph("Executive Summary", heading_style))
            story.append(Spacer(1, 0.1*inch))
            
            summary = report_data.metrics.get('summary', {})
            summary_text = f"""
            Total Trades: {summary.get('total_trades', 0)}<br/>
            Win Rate: {summary.get('win_rate', 0)*100:.1f}%<br/>
            Total P&L: {summary.get('total_pnl', 0):.2f}<br/>
            Max Drawdown: {summary.get('max_drawdown', 0)*100:.1f}%<br/>
            Sharpe Ratio: {report_data.performance.get('sharpe_ratio', 0):.2f}<br/>
            """
            story.append(Paragraph(summary_text, normal_style))
            story.append(Spacer(1, 0.25*inch))
            
            # Risk Metrics
            story.append(Paragraph("Risk Metrics", heading_style))
            story.append(Spacer(1, 0.1*inch))
            
            risk_metrics = report_data.metrics.get('risk', {})
            risk_text = f"""
            Current Drawdown: {risk_metrics.get('drawdown', 0)*100:.1f}%<br/>
            VaR (95%): {risk_metrics.get('var_95', 0)*100:.1f}%<br/>
            Leverage: {risk_metrics.get('leverage', 0):.1f}x<br/>
            Concentration: {risk_metrics.get('concentration', 0)*100:.1f}%<br/>
            Diversification Score: {risk_metrics.get('diversification', 0):.2f}<br/>
            """
            story.append(Paragraph(risk_text, normal_style))
            story.append(Spacer(1, 0.25*inch))
            
            # Charts
            if request.include_charts and request.chart_types:
                story.append(PageBreak())
                story.append(Paragraph("Charts", heading_style))
                story.append(Spacer(1, 0.1*inch))
                
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            img = Image(chart_image, width=6*inch, height=4*inch)
                            story.append(KeepTogether([Paragraph(f"<b>{chart_data.title}</b>", heading_style), img, Spacer(1, 0.1*inch)]))
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                story.append(PageBreak())
                story.append(Paragraph("Recommendations", heading_style))
                story.append(Spacer(1, 0.1*inch))
                
                for rec in report_data.recommendations:
                    story.append(Paragraph(f"• <b>{rec.get('action', '').replace('_', ' ').title()}</b>: {rec.get('description', '')}", normal_style))
                    story.append(Spacer(1, 0.05*inch))
            
            # Build document
            doc.build(story)
            
            # Get PDF content
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return b""

    # =========================================================================
    # HTML Generation
    # =========================================================================

    async def _generate_html(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate HTML report"""
        try:
            html = []
            
            # Header
            html.append("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Risk Report</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
                    h2 { color: #34495e; margin-top: 30px; }
                    .summary-box { background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 10px 0; }
                    .metric { display: inline-block; margin: 10px 20px 10px 0; }
                    .metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }
                    .metric-label { font-size: 14px; color: #7f8c8d; }
                    .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                    .table th { background: #3498db; color: white; padding: 10px; text-align: left; }
                    .table td { padding: 10px; border-bottom: 1px solid #ddd; }
                    .table tr:hover { background: #f5f5f5; }
                    .alert-critical { color: #e74c3c; }
                    .alert-warning { color: #f39c12; }
                    .alert-info { color: #3498db; }
                    .recommendation { background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }
                    .chart-container { margin: 20px 0; text-align: center; }
                    img { max-width: 100%; border: 1px solid #ddd; border-radius: 5px; }
                </style>
            </head>
            <body>
            """)
            
            # Title
            html.append(f"<h1>Risk Report - {report_data.report_type.value.upper()}</h1>")
            html.append(f"<p><strong>Portfolio:</strong> {report_data.portfolio_id}</p>")
            html.append(f"<p><strong>Period:</strong> {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}</p>")
            html.append(f"<p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>")
            
            # Summary
            html.append("<h2>Executive Summary</h2>")
            html.append('<div class="summary-box">')
            
            summary = report_data.metrics.get('summary', {})
            html.append(f"""
            <div class="metric">
                <div class="metric-value">{summary.get('total_trades', 0)}</div>
                <div class="metric-label">Total Trades</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('win_rate', 0)*100:.1f}%</div>
                <div class="metric-label">Win Rate</div>
            </div>
            <div class="metric">
                <div class="metric-value">${summary.get('total_pnl', 0):.2f}</div>
                <div class="metric-label">Total P&L</div>
            </div>
            <div class="metric">
                <div class="metric-value">{report_data.performance.get('sharpe_ratio', 0):.2f}</div>
                <div class="metric-label">Sharpe Ratio</div>
            </div>
            """)
            html.append('</div>')
            
            # Risk Metrics
            html.append("<h2>Risk Metrics</h2>")
            html.append('<div class="summary-box">')
            
            risk_metrics = report_data.metrics.get('risk', {})
            html.append(f"""
            <div class="metric">
                <div class="metric-value">{risk_metrics.get('drawdown', 0)*100:.1f}%</div>
                <div class="metric-label">Current Drawdown</div>
            </div>
            <div class="metric">
                <div class="metric-value">{risk_metrics.get('var_95', 0)*100:.1f}%</div>
                <div class="metric-label">VaR (95%)</div>
            </div>
            <div class="metric">
                <div class="metric-value">{risk_metrics.get('leverage', 0):.1f}x</div>
                <div class="metric-label">Leverage</div>
            </div>
            <div class="metric">
                <div class="metric-value">{risk_metrics.get('concentration', 0)*100:.1f}%</div>
                <div class="metric-label">Concentration</div>
            </div>
            """)
            html.append('</div>')
            
            # Charts
            if request.include_charts and request.chart_types:
                html.append("<h2>Charts</h2>")
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            import base64
                            img_base64 = base64.b64encode(chart_image).decode('utf-8')
                            html.append(f"""
                            <div class="chart-container">
                                <h3>{chart_data.title}</h3>
                                <img src="data:image/png;base64,{img_base64}" />
                            </div>
                            """)
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                html.append("<h2>Recommendations</h2>")
                for rec in report_data.recommendations:
                    severity_class = {
                        'critical': 'alert-critical',
                        'high': 'alert-warning',
                        'medium': 'alert-info',
                        'low': 'alert-info'
                    }.get(rec.get('severity', 'low'), 'alert-info')
                    
                    html.append(f"""
                    <div class="recommendation {severity_class}">
                        <strong>{rec.get('action', '').replace('_', ' ').title()}</strong>
                        <p>{rec.get('description', '')}</p>
                        <p><em>Expected Impact: {rec.get('expected_impact', 'Not specified')}</em></p>
                    </div>
                    """)
            
            # Footer
            html.append("""
            <hr/>
            <p style="color: #7f8c8d; font-size: 12px;">
                Generated by NEXUS AI Trading System Risk Reporter
            </p>
            </body>
            </html>
            """)
            
            return ''.join(html)
            
        except Exception as e:
            logger.error(f"Error generating HTML: {e}")
            return ""

    # =========================================================================
    # JSON Generation
    # =========================================================================

    async def _generate_json(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate JSON report"""
        try:
            report = {
                'report_type': report_data.report_type.value,
                'portfolio_id': report_data.portfolio_id,
                'generated_at': datetime.utcnow().isoformat(),
                'start_date': report_data.start_date.isoformat(),
                'end_date': report_data.end_date.isoformat(),
                'metrics': report_data.metrics,
                'performance': report_data.performance,
                'risk_metrics': report_data.risk_metrics,
                'limit_status': report_data.limit_status,
                'recommendations': report_data.recommendations,
                'position_count': len(report_data.positions),
                'trade_count': len(report_data.trades),
                'alert_count': len(report_data.alerts)
            }
            
            return json.dumps(report, indent=2, default=str)
            
        except Exception as e:
            logger.error(f"Error generating JSON: {e}")
            return "{}"

    # =========================================================================
    # CSV Generation
    # =========================================================================

    async def _generate_csv(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate CSV report"""
        try:
            output = []
            
            # Trades CSV
            if report_data.trades:
                output.append("=== Trades ===")
                output.append("timestamp,symbol,side,size,price,pnl")
                for trade in report_data.trades:
                    output.append(f"{trade.execution_time},{trade.symbol},{trade.side},{trade.size},{trade.price},{trade.pnl}")
            
            # Snapshots CSV
            if report_data.snapshots:
                output.append("\n=== Portfolio Snapshots ===")
                output.append("timestamp,total_value,drawdown,var_95,leverage,sharpe_ratio")
                for snapshot in report_data.snapshots:
                    output.append(f"{snapshot.timestamp},{snapshot.total_value},{snapshot.drawdown},{snapshot.var_95},{snapshot.leverage},{snapshot.sharpe_ratio}")
            
            # Alerts CSV
            if report_data.alerts:
                output.append("\n=== Alerts ===")
                output.append("timestamp,category,severity,title,resolved")
                for alert in report_data.alerts:
                    output.append(f"{alert.timestamp},{alert.category},{alert.severity},{alert.title},{alert.resolved}")
            
            # Metrics CSV
            if report_data.metrics:
                output.append("\n=== Metrics ===")
                output.append("metric,value")
                for key, value in report_data.metrics.get('summary', {}).items():
                    output.append(f"{key},{value}")
            
            return '\n'.join(output)
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            return ""

    # =========================================================================
    # Excel Generation
    # =========================================================================

    async def _generate_excel(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            
            # Title
            ws['A1'] = f"Risk Report - {report_data.report_type.value.upper()}"
            ws['A1'].font = Font(size=16, bold=True)
            
            # Portfolio info
            ws['A3'] = f"Portfolio: {report_data.portfolio_id}"
            ws['A4'] = f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}"
            ws['A5'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            
            # Metrics
            row = 7
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for key, value in report_data.metrics.get('summary', {}).items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            # Trades sheet
            if report_data.trades:
                ws_trades = wb.create_sheet("Trades")
                headers = ['Timestamp', 'Symbol', 'Side', 'Size', 'Price', 'PnL']
                for col, header in enumerate(headers, 1):
                    ws_trades.cell(row=1, column=col, value=header)
                    ws_trades.cell(row=1, column=col).font = Font(bold=True)
                
                for row, trade in enumerate(report_data.trades, 2):
                    ws_trades.cell(row=row, column=1, value=trade.execution_time)
                    ws_trades.cell(row=row, column=2, value=trade.symbol)
                    ws_trades.cell(row=row, column=3, value=trade.side)
                    ws_trades.cell(row=row, column=4, value=trade.size)
                    ws_trades.cell(row=row, column=5, value=trade.price)
                    ws_trades.cell(row=row, column=6, value=trade.pnl)
            
            # Snapshots sheet
            if report_data.snapshots:
                ws_snap = wb.create_sheet("Snapshots")
                headers = ['Timestamp', 'Total Value', 'Drawdown', 'VaR 95%', 'Leverage', 'Sharpe Ratio']
                for col, header in enumerate(headers, 1):
                    ws_snap.cell(row=1, column=col, value=header)
                    ws_snap.cell(row=1, column=col).font = Font(bold=True)
                
                for row, snap in enumerate(report_data.snapshots, 2):
                    ws_snap.cell(row=row, column=1, value=snap.timestamp)
                    ws_snap.cell(row=row, column=2, value=snap.total_value)
                    ws_snap.cell(row=row, column=3, value=snap.drawdown)
                    ws_snap.cell(row=row, column=4, value=snap.var_95)
                    ws_snap.cell(row=row, column=5, value=snap.leverage)
                    ws_snap.cell(row=row, column=6, value=snap.sharpe_ratio)
            
            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    # =========================================================================
    # Markdown Generation
    # =========================================================================

    async def _generate_markdown(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate Markdown report"""
        try:
            md = []
            
            # Title
            md.append(f"# Risk Report - {report_data.report_type.value.upper()}\n")
            
            # Info
            md.append(f"**Portfolio:** {report_data.portfolio_id}")
            md.append(f"**Period:** {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}")
            md.append(f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n")
            
            # Summary
            md.append("## Executive Summary\n")
            
            summary = report_data.metrics.get('summary', {})
            md.append("| Metric | Value |")
            md.append("|--------|-------|")
            md.append(f"| Total Trades | {summary.get('total_trades', 0)} |")
            md.append(f"| Win Rate | {summary.get('win_rate', 0)*100:.1f}% |")
            md.append(f"| Total P&L | ${summary.get('total_pnl', 0):.2f} |")
            md.append(f"| Sharpe Ratio | {report_data.performance.get('sharpe_ratio', 0):.2f} |")
            md.append(f"| Max Drawdown | {report_data.performance.get('max_drawdown', 0)*100:.1f}% |\n")
            
            # Risk Metrics
            md.append("## Risk Metrics\n")
            
            risk_metrics = report_data.metrics.get('risk', {})
            md.append("| Metric | Value |")
            md.append("|--------|-------|")
            md.append(f"| Current Drawdown | {risk_metrics.get('drawdown', 0)*100:.1f}% |")
            md.append(f"| VaR (95%) | {risk_metrics.get('var_95', 0)*100:.1f}% |")
            md.append(f"| Leverage | {risk_metrics.get('leverage', 0):.1f}x |")
            md.append(f"| Concentration | {risk_metrics.get('concentration', 0)*100:.1f}% |")
            md.append(f"| Diversification Score | {risk_metrics.get('diversification', 0):.2f} |\n")
            
            # Recent Trades
            if report_data.trades:
                md.append("## Recent Trades\n")
                md.append("| Timestamp | Symbol | Side | Size | Price | PnL |")
                md.append("|-----------|--------|------|------|-------|-----|")
                for trade in report_data.trades[-20:]:
                    md.append(f"| {trade.execution_time} | {trade.symbol} | {trade.side} | {trade.size:.2f} | {trade.price:.2f} | {trade.pnl:.2f} |")
                md.append("")
            
            # Recommendations
            if report_data.recommendations:
                md.append("## Recommendations\n")
                for rec in report_data.recommendations:
                    md.append(f"### {rec.get('action', '').replace('_', ' ').title()}")
                    md.append(f"{rec.get('description', '')}")
                    if rec.get('expected_impact'):
                        md.append(f"*Expected Impact: {rec.get('expected_impact')}*")
                    md.append("")
            
            return '\n'.join(md)
            
        except Exception as e:
            logger.error(f"Error generating Markdown: {e}")
            return ""

    # =========================================================================
    # Chart Generation
    # =========================================================================

    async def _prepare_chart_data(
        self,
        report_data: ReportData,
        chart_type: ChartType
    ) -> Optional[ChartData]:
        """Prepare data for chart generation"""
        try:
            if chart_type == ChartType.EQUITY:
                # Equity curve
                if report_data.snapshots:
                    timestamps = [s.timestamp for s in report_data.snapshots]
                    values = [float(s.total_value) for s in report_data.snapshots]
                    
                    return ChartData(
                        title="Equity Curve",
                        chart_type=ChartType.LINE,
                        data={
                            'timestamps': timestamps,
                            'values': values
                        },
                        x_label="Date",
                        y_label="Portfolio Value ($)"
                    )
            
            elif chart_type == ChartType.DRAWDOWN:
                # Drawdown chart
                if report_data.snapshots:
                    timestamps = [s.timestamp for s in report_data.snapshots]
                    drawdowns = [float(s.drawdown) * 100 for s in report_data.snapshots]
                    
                    return ChartData(
                        title="Drawdown Chart",
                        chart_type=ChartType.AREA,
                        data={
                            'timestamps': timestamps,
                            'values': drawdowns
                        },
                        x_label="Date",
                        y_label="Drawdown (%)"
                    )
            
            elif chart_type == ChartType.RISK:
                # Risk metrics radar
                if report_data.risk_metrics:
                    metrics = {
                        'Drawdown': report_data.risk_metrics.get('drawdown', 0),
                        'VaR': report_data.risk_metrics.get('var_95', 0),
                        'Leverage': report_data.risk_metrics.get('leverage', 0) / 3,
                        'Concentration': report_data.risk_metrics.get('concentration', 0),
                        'Diversification': report_data.risk_metrics.get('diversification', 0)
                    }
                    
                    return ChartData(
                        title="Risk Metrics",
                        chart_type=ChartType.BAR,
                        data={
                            'labels': list(metrics.keys()),
                            'values': list(metrics.values())
                        }
                    )
            
            elif chart_type == ChartType.HISTOGRAM:
                # PnL distribution
                if report_data.trades:
                    pnls = [float(t.pnl) for t in report_data.trades]
                    
                    return ChartData(
                        title="PnL Distribution",
                        chart_type=ChartType.HISTOGRAM,
                        data={
                            'values': pnls
                        },
                        x_label="PnL ($)",
                        y_label="Frequency"
                    )
            
            return None
            
        except Exception as e:
            logger.error(f"Error preparing chart data: {e}")
            return None

    async def _create_chart_image(self, chart_data: ChartData) -> Optional[bytes]:
        """Create chart image"""
        try:
            fig, ax = plt.subplots(figsize=(12, 8))
            
            if chart_data.chart_type == ChartType.LINE:
                timestamps = chart_data.data.get('timestamps', [])
                values = chart_data.data.get('values', [])
                ax.plot(timestamps, values, linewidth=2, color='#3498db')
                ax.fill_between(timestamps, values, alpha=0.3, color='#3498db')
                
            elif chart_data.chart_type == ChartType.AREA:
                timestamps = chart_data.data.get('timestamps', [])
                values = chart_data.data.get('values', [])
                ax.fill_between(timestamps, values, alpha=0.5, color='#e74c3c')
                ax.plot(timestamps, values, linewidth=2, color='#e74c3c')
                
            elif chart_data.chart_type == ChartType.BAR:
                labels = chart_data.data.get('labels', [])
                values = chart_data.data.get('values', [])
                colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6']
                ax.bar(labels, values, color=colors[:len(labels)])
                
            elif chart_data.chart_type == ChartType.HISTOGRAM:
                values = chart_data.data.get('values', [])
                ax.hist(values, bins=30, color='#3498db', edgecolor='black', alpha=0.7)
                ax.axvline(np.mean(values), color='#e74c3c', linestyle='dashed', linewidth=2, label='Mean')
                ax.legend()
            
            # Formatting
            ax.set_xlabel(chart_data.x_label or '')
            ax.set_ylabel(chart_data.y_label or '')
            ax.set_title(chart_data.title, fontsize=16, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            # Rotate x labels if needed
            if chart_data.x_label == "Date":
                plt.xticks(rotation=45)
            
            # Tight layout
            plt.tight_layout()
            
            # Save to bytes
            buffer = BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating chart image: {e}")
            return None

    # =========================================================================
    # Report Management
    # =========================================================================

    async def get_report(
        self,
        report_id: str
    ) -> Optional[Dict[str, Any]]:
        """Get cached report"""
        return self._report_cache.get(report_id)

    async def delete_report(
        self,
        report_id: str
    ) -> bool:
        """Delete cached report"""
        if report_id in self._report_cache:
            del self._report_cache[report_id]
            return True
        return False

    async def list_reports(
        self,
        portfolio_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """List available reports"""
        reports = []
        
        for report_id, cache_data in self._report_cache.items():
            response = cache_data.get('response')
            if response:
                if portfolio_id and response.portfolio_id != portfolio_id:
                    continue
                reports.append({
                    'report_id': report_id,
                    'report_type': response.report_type.value,
                    'portfolio_id': response.portfolio_id,
                    'generated_at': response.generated_at,
                    'format': response.format.value,
                    'file_size': response.file_size
                })
        
        # Sort by generation time
        reports.sort(key=lambda x: x['generated_at'], reverse=True)
        
        return reports[:limit]

    # =========================================================================
    # Scheduled Reports
    # =========================================================================

    async def create_schedule(self, schedule: ReportSchedule) -> bool:
        """Create a report schedule"""
        try:
            # Calculate next generation time
            schedule.next_generation = self._calculate_next_generation(schedule)
            
            self._schedules[schedule.report_type.value] = schedule
            
            logger.info(f"Created schedule for {schedule.report_type.value}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating schedule: {e}")
            return False

    async def update_schedule(
        self,
        report_type: ReportType,
        updates: Dict[str, Any]
    ) -> bool:
        """Update a report schedule"""
        try:
            key = report_type.value
            if key not in self._schedules:
                return False
            
            schedule = self._schedules[key]
            for field, value in updates.items():
                if hasattr(schedule, field):
                    setattr(schedule, field, value)
            
            # Recalculate next generation
            schedule.next_generation = self._calculate_next_generation(schedule)
            
            logger.info(f"Updated schedule for {key}")
            return True
            
        except Exception as e:
            logger.error(f"Error updating schedule: {e}")
            return False

    async def delete_schedule(self, report_type: ReportType) -> bool:
        """Delete a report schedule"""
        try:
            key = report_type.value
            if key in self._schedules:
                del self._schedules[key]
                logger.info(f"Deleted schedule for {key}")
                return True
            return False
            
        except Exception as e:
            logger.error(f"Error deleting schedule: {e}")
            return False

    async def get_schedule(
        self,
        report_type: ReportType
    ) -> Optional[ReportSchedule]:
        """Get a report schedule"""
        return self._schedules.get(report_type.value)

    async def get_all_schedules(self) -> List[ReportSchedule]:
        """Get all report schedules"""
        return list(self._schedules.values())

    def _calculate_next_generation(self, schedule: ReportSchedule) -> datetime:
        """Calculate next generation time for schedule"""
        now = datetime.utcnow()
        
        if schedule.frequency == ReportFrequency.DAILY:
            # Next day at specified time
            next_time = now.replace(hour=int(schedule.time.split(':')[0]),
                                   minute=int(schedule.time.split(':')[1]),
                                   second=0, microsecond=0)
            if next_time <= now:
                next_time += timedelta(days=1)
            return next_time
            
        elif schedule.frequency == ReportFrequency.WEEKLY:
            # Next week on specified day
            days_until = schedule.day_of_week - now.weekday()
            if days_until <= 0:
                days_until += 7
            next_time = now + timedelta(days=days_until)
            next_time = next_time.replace(hour=int(schedule.time.split(':')[0]),
                                         minute=int(schedule.time.split(':')[1]),
                                         second=0, microsecond=0)
            return next_time
            
        elif schedule.frequency == ReportFrequency.MONTHLY:
            # Next month on specified day
            if schedule.day_of_month:
                if now.day < schedule.day_of_month:
                    next_time = now.replace(day=schedule.day_of_month,
                                          hour=int(schedule.time.split(':')[0]),
                                          minute=int(schedule.time.split(':')[1]),
                                          second=0, microsecond=0)
                else:
                    # Next month
                    next_month = now.month + 1
                    year = now.year
                    if next_month > 12:
                        next_month = 1
                        year += 1
                    next_time = now.replace(year=year, month=next_month,
                                          day=schedule.day_of_month,
                                          hour=int(schedule.time.split(':')[0]),
                                          minute=int(schedule.time.split(':')[1]),
                                          second=0, microsecond=0)
                return next_time
        
        # Default: next day
        return now + timedelta(days=1)

    # =========================================================================
    # Email Delivery
    # =========================================================================

    async def _send_report_email(
        self,
        response: ReportResponse,
        content: Any,
        recipients: List[str]
    ) -> None:
        """Send report via email"""
        try:
            # Implementation would use email service
            logger.info(f"Sending report {response.report_id} to {recipients}")
        except Exception as e:
            logger.error(f"Error sending report email: {e}")

    # =========================================================================
    # Cleanup
    # =========================================================================

    async def close(self) -> None:
        """Close the risk reporter"""
        # Clear cache
        self._report_cache.clear()
        self._schedules.clear()
        
        logger.info("RiskReporter closed")


# =============================================================================
# FASTAPI ROUTER
# =============================================================================

from fastapi import APIRouter, Depends, Query, Body, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse

router = APIRouter(prefix="/api/v1/risk-reports", tags=["Risk Reports"])


async def get_reporter() -> RiskReporter:
    """Dependency to get RiskReporter instance"""
    return RiskReporter()


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    background_tasks: BackgroundTasks,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Generate a risk report"""
    response = await reporter.generate_report(request)
    
    # Clean up old reports in background
    background_tasks.add_task(reporter.cleanup_old_reports, max_age_hours=24)
    
    return response


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Get a generated report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return report


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Delete a report"""
    success = await reporter.delete_report(report_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return {"success": True}


@router.get("/")
async def list_reports(
    portfolio_id: Optional[str] = Query(None),
    limit: int = Query(50, le=500),
    reporter: RiskReporter = Depends(get_reporter)
):
    """List available reports"""
    return await reporter.list_reports(portfolio_id, limit)


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Download a report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    
    response_data = report.get('response')
    content = report.get('content')
    
    if not content:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report content not found"
        )
    
    # Determine content type
    content_type = {
        'pdf': 'application/pdf',
        'html': 'text/html',
        'json': 'application/json',
        'csv': 'text/csv',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'markdown': 'text/markdown'
    }.get(response_data.format.value, 'application/octet-stream')
    
    # Return file
    return Response(
        content=content if isinstance(content, bytes) else content.encode('utf-8'),
        media_type=content_type,
        headers={
            'Content-Disposition': f'attachment; filename="risk_report_{report_id}.{response_data.format.value}"'
        }
    )


@router.post("/schedule")
async def create_schedule(
    schedule: ReportSchedule,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Create a report schedule"""
    success = await reporter.create_schedule(schedule)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create schedule"
        )
    return {"success": True}


@router.put("/schedule/{report_type}")
async def update_schedule(
    report_type: ReportType,
    updates: Dict[str, Any],
    reporter: RiskReporter = Depends(get_reporter)
):
    """Update a report schedule"""
    success = await reporter.update_schedule(report_type, updates)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {report_type.value} not found"
        )
    return {"success": True}


@router.delete("/schedule/{report_type}")
async def delete_schedule(
    report_type: ReportType,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Delete a report schedule"""
    success = await reporter.delete_schedule(report_type)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {report_type.value} not found"
        )
    return {"success": True}


@router.get("/schedule/{report_type}")
async def get_schedule(
    report_type: ReportType,
    reporter: RiskReporter = Depends(get_reporter)
):
    """Get a report schedule"""
    schedule = await reporter.get_schedule(report_type)
    if not schedule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {report_type.value} not found"
        )
    return schedule


@router.get("/schedule")
async def get_all_schedules(
    reporter: RiskReporter = Depends(get_reporter)
):
    """Get all report schedules"""
    return await reporter.get_all_schedules()


@router.get("/chart-types")
async def get_chart_types():
    """Get available chart types"""
    return {
        'chart_types': [
            {'name': ct.value, 'description': ct.name.replace('_', ' ').title()}
            for ct in ChartType
        ]
    }


@router.get("/report-types")
async def get_report_types():
    """Get available report types"""
    return {
        'report_types': [
            {'name': rt.value, 'description': rt.name.replace('_', ' ').title()}
            for rt in ReportType
        ]
    }


@router.get("/formats")
async def get_report_formats():
    """Get available report formats"""
    return {
        'formats': [
            {'name': f.value, 'description': f.name.upper()}
            for f in ReportFormat
        ]
    }


# =============================================================================
# EXPORTS
# =============================================================================

__all__ = [
    'RiskReporter',
    'ReportRequest',
    'ReportResponse',
    'ReportSchedule',
    'ReportFormat',
    'ReportType',
    'ReportFrequency',
    'ChartType',
    'ReportData',
    'ChartData',
    'router'
]
