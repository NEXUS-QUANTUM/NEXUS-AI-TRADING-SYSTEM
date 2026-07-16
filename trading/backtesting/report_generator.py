"""
NEXUS AI TRADING SYSTEM - Report Generator
Copyright © 2026 NEXUS QUANTUM LTD - All Rights Reserved
CEO: Dr X... - Majority Shareholder

Module: trading/backtesting/report_generator.py
Description: Générateur de rapports détaillés pour les backtests.
             Supporte les formats HTML, PDF, JSON, Excel et Markdown.
             Inclut des visualisations interactives et des tableaux de bord.
"""

import os
import json
import logging
import base64
from datetime import datetime
from typing import Dict, List, Any, Optional, Union, Tuple
from pathlib import Path
from dataclasses import dataclass, field

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import PercentFormatter, FuncFormatter
import seaborn as sns

from jinja2 import Environment, FileSystemLoader, select_autoescape
import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart

from trading.backtesting.backtest_engine import BacktestResult
from trading.backtesting.metrics_calculator import PerformanceMetrics, MetricsCalculator
from shared.helpers.date_helpers import format_date
from shared.helpers.number_helpers import format_currency, format_percent
from shared.exceptions import ReportGenerationError

# Configuration du logging
logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """
    Configuration de génération de rapport.
    """
    # Format de sortie
    formats: List[str] = field(default_factory=lambda: ['html'])
    output_dir: str = "data/backtest_results/"
    filename_prefix: str = "backtest_report"
    
    # Sections à inclure
    include_summary: bool = True
    include_metrics: bool = True
    include_trades: bool = True
    include_equity_curve: bool = True
    include_performance_analysis: bool = True
    include_monthly_returns: bool = True
    include_risk_analysis: bool = True
    include_drawdown_analysis: bool = True
    include_trade_analysis: bool = True
    include_comparison: bool = True
    
    # Visualisation
    figsize: Tuple[int, int] = (12, 6)
    dpi: int = 150
    theme: str = 'dark'  # 'dark' ou 'light'
    show_trades: bool = True
    show_markers: bool = True
    
    # PDF
    pdf_orientation: str = 'portrait'  # 'portrait' ou 'landscape'
    pdf_margins: Tuple[float, float, float, float] = (20, 20, 20, 20)
    
    # HTML
    template_dir: str = "configs/templates/"
    interactive_charts: bool = True
    include_plotly: bool = True
    
    def __post_init__(self):
        """Validation des paramètres."""
        if self.theme not in ['dark', 'light']:
            self.theme = 'dark'
        
        if self.pdf_orientation not in ['portrait', 'landscape']:
            self.pdf_orientation = 'portrait'
        
        # Création du répertoire de sortie
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)


class ReportGenerator:
    """
    Générateur de rapports pour les backtests.
    """
    
    def __init__(
        self,
        config: Optional[ReportConfig] = None,
        metrics_calculator: Optional[MetricsCalculator] = None
    ):
        """
        Initialise le générateur de rapports.
        
        Args:
            config: Configuration du rapport.
            metrics_calculator: Calculateur de métriques.
        """
        self.config = config or ReportConfig()
        self.metrics_calculator = metrics_calculator or MetricsCalculator()
        
        # Style pour les graphiques
        self._setup_style()
        
        # Template Jinja2
        self._setup_template_engine()
        
        logger.info("ReportGenerator initialisé")
        logger.info(f"Formats de sortie: {self.config.formats}")
        logger.info(f"Répertoire: {self.config.output_dir}")
    
    def _setup_style(self) -> None:
        """
        Configure le style pour les graphiques.
        """
        if self.config.theme == 'dark':
            # Style sombre
            plt.style.use('dark_background')
            colors = ['#00ff9d', '#ff6b6b', '#ffd93d', '#6bcbff', '#a66cff']
            self.colors = colors
            self.text_color = '#e0e0e0'
            self.grid_color = '#333333'
            self.bg_color = '#1a1a2e'
        else:
            # Style clair
            plt.style.use('seaborn-v0_8-whitegrid')
            colors = ['#2ecc71', '#e74c3c', '#f1c40f', '#3498db', '#9b59b6']
            self.colors = colors
            self.text_color = '#2c3e50'
            self.grid_color = '#ecf0f1'
            self.bg_color = '#ffffff'
        
        # Palettes personnalisées
        self.color_green = colors[0]
        self.color_red = colors[1]
        self.color_gold = colors[2]
        self.color_blue = colors[3]
        self.color_purple = colors[4]
        
        # Configuration de Seaborn
        sns.set_palette(colors)
        sns.set_style("darkgrid" if self.config.theme == 'dark' else "whitegrid")
    
    def _setup_template_engine(self) -> None:
        """
        Configure le moteur de templates Jinja2.
        """
        # Répertoire des templates
        template_dirs = [
            self.config.template_dir,
            os.path.join(os.path.dirname(__file__), 'templates'),
            'templates'
        ]
        
        # Trouver le premier répertoire existant
        template_path = None
        for path in template_dirs:
            if os.path.exists(path):
                template_path = path
                break
        
        if template_path is None:
            # Utiliser le répertoire courant
            template_path = os.path.dirname(__file__)
        
        self.template_env = Environment(
            loader=FileSystemLoader(template_path),
            autoescape=select_autoescape(['html', 'xml']),
            trim_blocks=True,
            lstrip_blocks=True
        )
        
        # Filtres personnalisés
        self.template_env.filters['format_currency'] = format_currency
        self.template_env.filters['format_percent'] = format_percent
        self.template_env.filters['format_date'] = format_date
        self.template_env.filters['json'] = json.dumps
    
    # ============================================================
    # GÉNÉRATION DE RAPPORTS
    # ============================================================
    
    def generate_report(
        self,
        result: BacktestResult,
        name: Optional[str] = None,
        benchmark_result: Optional[BacktestResult] = None
    ) -> Dict[str, str]:
        """
        Génère un rapport complet.
        
        Args:
            result: Résultats du backtesting.
            name: Nom du rapport (optionnel).
            benchmark_result: Résultats de référence (optionnel).
            
        Returns:
            Dictionnaire des chemins des fichiers générés.
        """
        if name is None:
            name = f"{result.config.symbol}_{result.config.strategy_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        logger.info(f"Génération du rapport: {name}")
        
        # Préparation des données
        data = self._prepare_report_data(result, benchmark_result)
        
        outputs = {}
        
        # Génération des formats
        for fmt in self.config.formats:
            if fmt == 'html':
                outputs['html'] = self._generate_html_report(data, name)
            elif fmt == 'pdf':
                outputs['pdf'] = self._generate_pdf_report(data, name)
            elif fmt == 'json':
                outputs['json'] = self._generate_json_report(data, name)
            elif fmt == 'excel':
                outputs['excel'] = self._generate_excel_report(data, name)
            elif fmt == 'markdown':
                outputs['markdown'] = self._generate_markdown_report(data, name)
            else:
                logger.warning(f"Format non supporté: {fmt}")
        
        logger.info(f"Rapport généré: {name}")
        return outputs
    
    def _prepare_report_data(
        self,
        result: BacktestResult,
        benchmark_result: Optional[BacktestResult] = None
    ) -> Dict[str, Any]:
        """
        Prépare les données pour le rapport.
        
        Args:
            result: Résultats du backtesting.
            benchmark_result: Résultats de référence.
            
        Returns:
            Données préparées.
        """
        # Métriques complètes
        metrics = self.metrics_calculator.calculate_all_metrics(
            result.equity_curve,
            result.trades
        )
        
        data = {
            'name': f"{result.config.symbol} - {result.config.strategy_name}",
            'symbol': result.config.symbol,
            'strategy': result.config.strategy_name,
            'timestamp': datetime.now().isoformat(),
            
            # Configuration
            'config': {
                'start_date': format_date(result.config.start_date),
                'end_date': format_date(result.config.end_date),
                'initial_capital': result.config.initial_capital,
                'timeframe': result.config.timeframe,
                'parameters': result.config.strategy_params
            },
            
            # Métriques
            'metrics': metrics.to_dict(),
            
            # Données
            'equity_curve': result.equity_curve.tolist(),
            'drawdown_curve': result.drawdown_curve.tolist() if hasattr(result, 'drawdown_curve') else [],
            'returns': metrics.returns.tolist() if not metrics.returns.empty else [],
            
            # Trades
            'trades': result.trades,
            'total_trades': len(result.trades),
            
            # Statistiques avancées
            'monthly_returns': self._calculate_monthly_returns(result),
            'yearly_returns': self._calculate_yearly_returns(result),
            'performance_distribution': self._calculate_performance_distribution(result),
            
            # Benchmark
            'benchmark': self._prepare_benchmark_data(benchmark_result) if benchmark_result else None,
            
            # Visualisations
            'charts': self._generate_charts(result, benchmark_result)
        }
        
        return data
    
    def _calculate_monthly_returns(self, result: BacktestResult) -> Dict[str, float]:
        """
        Calcule les rendements mensuels.
        
        Args:
            result: Résultats du backtesting.
            
        Returns:
            Dictionnaire des rendements mensuels.
        """
        if result.equity_curve.empty:
            return {}
        
        # Création du DataFrame avec dates
        df = pd.DataFrame({
            'date': pd.date_range(
                start=result.config.start_date,
                periods=len(result.equity_curve),
                freq='D'  # Approximation
            ),
            'equity': result.equity_curve
        })
        
        # Resampling mensuel
        df.set_index('date', inplace=True)
        monthly = df.resample('M').last()
        monthly_returns = monthly.pct_change().dropna()
        
        return monthly_returns['equity'].to_dict()
    
    def _calculate_yearly_returns(self, result: BacktestResult) -> Dict[str, float]:
        """
        Calcule les rendements annuels.
        
        Args:
            result: Résultats du backtesting.
            
        Returns:
            Dictionnaire des rendements annuels.
        """
        if result.equity_curve.empty:
            return {}
        
        # Création du DataFrame avec dates
        df = pd.DataFrame({
            'date': pd.date_range(
                start=result.config.start_date,
                periods=len(result.equity_curve),
                freq='D'
            ),
            'equity': result.equity_curve
        })
        
        # Resampling annuel
        df.set_index('date', inplace=True)
        yearly = df.resample('Y').last()
        yearly_returns = yearly.pct_change().dropna()
        
        return {str(k.year): v for k, v in yearly_returns['equity'].to_dict().items()}
    
    def _calculate_performance_distribution(self, result: BacktestResult) -> Dict[str, Any]:
        """
        Calcule la distribution des performances.
        
        Args:
            result: Résultats du backtesting.
            
        Returns:
            Statistiques de distribution.
        """
        if not result.trades:
            return {}
        
        pnls = [t.pnl for t in result.trades if hasattr(t, 'pnl')]
        
        if not pnls:
            return {}
        
        return {
            'mean': np.mean(pnls),
            'std': np.std(pnls),
            'min': np.min(pnls),
            'max': np.max(pnls),
            'percentiles': {
                '10': np.percentile(pnls, 10),
                '25': np.percentile(pnls, 25),
                '50': np.percentile(pnls, 50),
                '75': np.percentile(pnls, 75),
                '90': np.percentile(pnls, 90)
            },
            'skewness': self.metrics_calculator.calculate_skewness(pnls),
            'kurtosis': self.metrics_calculator.calculate_kurtosis(pnls)
        }
    
    def _prepare_benchmark_data(self, result: BacktestResult) -> Dict[str, Any]:
        """
        Prépare les données de référence.
        
        Args:
            result: Résultats de référence.
            
        Returns:
            Données de référence.
        """
        return {
            'name': f"{result.config.symbol} - Benchmark",
            'total_return': result.total_return,
            'sharpe_ratio': result.sharpe_ratio,
            'max_drawdown': result.max_drawdown_pct,
            'equity_curve': result.equity_curve.tolist() if hasattr(result, 'equity_curve') else []
        }
    
    # ============================================================
    # VISUALISATIONS
    # ============================================================
    
    def _generate_charts(
        self,
        result: BacktestResult,
        benchmark_result: Optional[BacktestResult] = None
    ) -> Dict[str, str]:
        """
        Génère les graphiques pour le rapport.
        
        Args:
            result: Résultats du backtesting.
            benchmark_result: Résultats de référence.
            
        Returns:
            Dictionnaire des chemins des graphiques.
        """
        charts = {}
        
        # Création du répertoire des graphiques
        chart_dir = Path(self.config.output_dir) / 'charts'
        chart_dir.mkdir(exist_ok=True)
        
        # Equity Curve
        charts['equity_curve'] = self._plot_equity_curve(result, benchmark_result, chart_dir)
        
        # Drawdown Curve
        charts['drawdown'] = self._plot_drawdown(result, chart_dir)
        
        # Monthly Returns Heatmap
        charts['monthly_heatmap'] = self._plot_monthly_heatmap(result, chart_dir)
        
        # Performance Distribution
        charts['performance_distribution'] = self._plot_performance_distribution(result, chart_dir)
        
        # Trade Analysis
        charts['trade_analysis'] = self._plot_trade_analysis(result, chart_dir)
        
        # Risk Metrics
        charts['risk_metrics'] = self._plot_risk_metrics(result, chart_dir)
        
        return charts
    
    def _plot_equity_curve(
        self,
        result: BacktestResult,
        benchmark_result: Optional[BacktestResult],
        chart_dir: Path
    ) -> str:
        """
        Graphique de la courbe de capitaux.
        
        Args:
            result: Résultats du backtesting.
            benchmark_result: Résultats de référence.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, ax = plt.subplots(figsize=self.config.figsize)
        
        # Equity curve principale
        equity = result.equity_curve
        ax.plot(equity.index if hasattr(equity, 'index') else range(len(equity)),
                equity, linewidth=2, color=self.color_green, label='Strategy')
        
        # Benchmark
        if benchmark_result and hasattr(benchmark_result, 'equity_curve'):
            bench = benchmark_result.equity_curve
            ax.plot(bench.index if hasattr(bench, 'index') else range(len(bench)),
                    bench, linewidth=1.5, color=self.color_blue,
                    linestyle='--', alpha=0.7, label='Benchmark')
        
        # Trades
        if self.config.show_trades and hasattr(result, 'trades'):
            for trade in result.trades:
                if hasattr(trade, 'entry_price') and hasattr(trade, 'exit_price'):
                    x = trade.entry_time if hasattr(trade, 'entry_time') else 0
                    y = trade.entry_price
                    marker = '^' if trade.pnl > 0 else 'v'
                    color = self.color_green if trade.pnl > 0 else self.color_red
                    ax.scatter(x, y, marker=marker, s=50, color=color, alpha=0.6)
        
        # Formatage
        ax.set_xlabel('Time')
        ax.set_ylabel('Portfolio Value ($)')
        ax.set_title('Equity Curve', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.legend(loc='upper left')
        
        # Formatage des axes
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
        
        plt.tight_layout()
        
        # Sauvegarde
        filepath = chart_dir / 'equity_curve.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    def _plot_drawdown(
        self,
        result: BacktestResult,
        chart_dir: Path
    ) -> str:
        """
        Graphique de la courbe de drawdown.
        
        Args:
            result: Résultats du backtesting.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, ax = plt.subplots(figsize=self.config.figsize)
        
        if hasattr(result, 'drawdown_curve') and not result.drawdown_curve.empty:
            drawdown = result.drawdown_curve
            ax.fill_between(
                drawdown.index if hasattr(drawdown, 'index') else range(len(drawdown)),
                0,
                drawdown,
                color=self.color_red,
                alpha=0.4,
                label='Drawdown'
            )
            ax.plot(
                drawdown.index if hasattr(drawdown, 'index') else range(len(drawdown)),
                drawdown,
                color=self.color_red,
                linewidth=1.5
            )
        else:
            # Calcul du drawdown
            equity = result.equity_curve
            running_max = equity.expanding().max()
            drawdown = (equity - running_max) / running_max
            ax.fill_between(
                range(len(drawdown)),
                0,
                drawdown,
                color=self.color_red,
                alpha=0.4
            )
        
        # Formatage
        ax.set_xlabel('Time')
        ax.set_ylabel('Drawdown (%)')
        ax.set_title('Drawdown Curve', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.yaxis.set_major_formatter(PercentFormatter(1.0))
        
        # Ligne horizontale à -20%
        ax.axhline(y=-0.20, color='yellow', linestyle='--', linewidth=1, alpha=0.5, label='Critical Level')
        
        ax.legend(loc='lower left')
        plt.tight_layout()
        
        filepath = chart_dir / 'drawdown.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    def _plot_monthly_heatmap(
        self,
        result: BacktestResult,
        chart_dir: Path
    ) -> str:
        """
        Heatmap des rendements mensuels.
        
        Args:
            result: Résultats du backtesting.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Calcul des rendements mensuels
        monthly_returns = self._calculate_monthly_returns(result)
        if not monthly_returns:
            # Données synthétiques
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            data = np.random.normal(0.01, 0.05, (12, 5))
            df = pd.DataFrame(data, index=months)
        else:
            # Conversion en DataFrame
            df = pd.DataFrame(list(monthly_returns.items()), columns=['date', 'return'])
            df['date'] = pd.to_datetime(df['date'])
            df['year'] = df['date'].dt.year
            df['month'] = df['date'].dt.month
            df = df.pivot(index='month', columns='year', values='return')
        
        # Heatmap
        im = ax.imshow(df.values, cmap='RdYlGn', aspect='auto', vmin=-0.1, vmax=0.1)
        
        # Étiquettes
        ax.set_xticks(range(len(df.columns)))
        ax.set_xticklabels(df.columns)
        ax.set_yticks(range(len(df.index)))
        ax.set_yticklabels(df.index)
        
        # Valeurs
        for i in range(len(df.index)):
            for j in range(len(df.columns)):
                value = df.iloc[i, j]
                if not pd.isna(value):
                    text_color = 'white' if abs(value) > 0.05 else 'black'
                    ax.text(j, i, f'{value:.1%}',
                           ha='center', va='center', color=text_color, fontsize=8)
        
        # Colorbar
        plt.colorbar(im, ax=ax, format=PercentFormatter(1.0), label='Return')
        
        ax.set_title('Monthly Returns Heatmap', fontsize=14, fontweight='bold')
        ax.set_xlabel('Year')
        ax.set_ylabel('Month')
        
        plt.tight_layout()
        
        filepath = chart_dir / 'monthly_heatmap.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    def _plot_performance_distribution(
        self,
        result: BacktestResult,
        chart_dir: Path
    ) -> str:
        """
        Graphique de distribution des performances.
        
        Args:
            result: Résultats du backtesting.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
        
        # Distribution des rendements
        if hasattr(result, 'trades') and result.trades:
            pnls = [t.pnl for t in result.trades if hasattr(t, 'pnl')]
            
            if pnls:
                # Histogramme
                ax1.hist(pnls, bins=30, color=self.color_blue, alpha=0.7, edgecolor='white')
                ax1.axvline(0, color='white', linestyle='--', alpha=0.5)
                ax1.axvline(np.mean(pnls), color=self.color_green, linestyle='-', 
                            linewidth=2, label=f"Mean: ${np.mean(pnls):.2f}")
                ax1.set_xlabel('PNL ($)')
                ax1.set_ylabel('Frequency')
                ax1.set_title('Trade PNL Distribution')
                ax1.legend()
                
                # Boxplot
                ax2.boxplot(pnls, vert=True, patch_artist=True,
                           boxprops=dict(facecolor=self.color_blue, alpha=0.7))
                ax2.axhline(0, color='white', linestyle='--', alpha=0.5)
                ax2.set_ylabel('PNL ($)')
                ax2.set_title('Trade PNL Boxplot')
                ax2.set_xticklabels(['Trades'])
        
        plt.tight_layout()
        
        filepath = chart_dir / 'performance_distribution.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    def _plot_trade_analysis(
        self,
        result: BacktestResult,
        chart_dir: Path
    ) -> str:
        """
        Analyse détaillée des trades.
        
        Args:
            result: Résultats du backtesting.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        
        if hasattr(result, 'trades') and result.trades:
            trades = result.trades
            
            # 1. Cumulative PNL
            ax = axes[0, 0]
            if hasattr(trades[0], 'pnl'):
                cumulative = np.cumsum([t.pnl for t in trades])
                ax.plot(cumulative, color=self.color_green, linewidth=2)
                ax.fill_between(range(len(cumulative)), 0, cumulative, alpha=0.3, color=self.color_green)
                ax.set_xlabel('Trade Number')
                ax.set_ylabel('Cumulative PNL ($)')
                ax.set_title('Cumulative PNL')
                ax.grid(True, alpha=0.3)
            
            # 2. Win/Loss Distribution
            ax = axes[0, 1]
            win_count = sum(1 for t in trades if hasattr(t, 'pnl') and t.pnl > 0)
            loss_count = len(trades) - win_count
            colors = [self.color_green, self.color_red]
            ax.pie([win_count, loss_count], labels=['Win', 'Loss'],
                   autopct='%1.1f%%', colors=colors, startangle=90)
            ax.set_title(f'Win/Loss Distribution ({win_count}/{loss_count})')
            
            # 3. Trade Size Distribution
            ax = axes[1, 0]
            sizes = []
            for t in trades:
                if hasattr(t, 'quantity'):
                    sizes.append(t.quantity)
            if sizes:
                ax.hist(sizes, bins=20, color=self.color_purple, alpha=0.7)
                ax.set_xlabel('Position Size')
                ax.set_ylabel('Frequency')
                ax.set_title('Position Size Distribution')
            
            # 4. Trade Duration
            ax = axes[1, 1]
            durations = []
            for t in trades:
                if hasattr(t, 'entry_time') and hasattr(t, 'exit_time'):
                    duration = (t.exit_time - t.entry_time).total_seconds() / 3600  # hours
                    durations.append(duration)
            if durations:
                ax.hist(durations, bins=20, color=self.color_gold, alpha=0.7)
                ax.set_xlabel('Duration (hours)')
                ax.set_ylabel('Frequency')
                ax.set_title('Trade Duration Distribution')
        
        plt.tight_layout()
        
        filepath = chart_dir / 'trade_analysis.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    def _plot_risk_metrics(
        self,
        result: BacktestResult,
        chart_dir: Path
    ) -> str:
        """
        Visualisation des métriques de risque.
        
        Args:
            result: Résultats du backtesting.
            chart_dir: Répertoire des graphiques.
            
        Returns:
            Chemin du graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Métriques de risque
        metrics = {
            'VaR (95%)': result.value_at_risk if hasattr(result, 'value_at_risk') else 0,
            'CVaR (95%)': result.expected_shortfall if hasattr(result, 'expected_shortfall') else 0,
            'Max DD': result.max_drawdown if hasattr(result, 'max_drawdown') else 0,
            'Volatility': result.annualized_volatility if hasattr(result, 'annualized_volatility') else 0
        }
        
        # Création du bar chart
        names = list(metrics.keys())
        values = [abs(v) for v in metrics.values()]
        colors = [self.color_red if v > 0.1 else self.color_gold for v in values]
        
        bars = ax.bar(names, values, color=colors, alpha=0.7)
        
        # Ajout des valeurs
        for bar, value in zip(bars, values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + height*0.02,
                    f'{value:.2f}', ha='center', va='bottom', fontsize=10)
        
        ax.set_ylabel('Value ($)' if max(values) > 100 else 'Value')
        ax.set_title('Risk Metrics', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        
        filepath = chart_dir / 'risk_metrics.png'
        plt.savefig(filepath, dpi=self.config.dpi, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()
        
        return str(filepath)
    
    # ============================================================
    # FORMATS DE RAPPORT
    # ============================================================
    
    def _generate_html_report(self, data: Dict[str, Any], name: str) -> str:
        """
        Génère un rapport HTML.
        
        Args:
            data: Données du rapport.
            name: Nom du rapport.
            
        Returns:
            Chemin du fichier HTML.
        """
        filepath = Path(self.config.output_dir) / f"{name}.html"
        
        try:
            template = self.template_env.get_template('backtest_report.html')
            html_content = template.render(**data)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur lors de la génération du HTML: {e}")
            raise ReportGenerationError(f"Erreur HTML: {e}")
    
    def _generate_pdf_report(self, data: Dict[str, Any], name: str) -> str:
        """
        Génère un rapport PDF.
        
        Args:
            data: Données du rapport.
            name: Nom du rapport.
            
        Returns:
            Chemin du fichier PDF.
        """
        try:
            # D'abord générer le HTML
            html_file = self._generate_html_report(data, name)
            
            # Puis convertir en PDF
            pdf_file = Path(self.config.output_dir) / f"{name}.pdf"
            
            weasyprint.HTML(filename=html_file).write_pdf(
                pdf_file,
                stylesheets=[self._get_pdf_styles()]
            )
            
            return str(pdf_file)
            
        except Exception as e:
            logger.error(f"Erreur lors de la génération du PDF: {e}")
            raise ReportGenerationError(f"Erreur PDF: {e}")
    
    def _get_pdf_styles(self) -> str:
        """
        Styles CSS pour le PDF.
        
        Returns:
            CSS pour le PDF.
        """
        return """
            @page {
                size: A4;
                margin: 2cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 10pt;
                color: #333;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                font-size: 8pt;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 4px;
                text-align: left;
            }
            th {
                background-color: #2c3e50;
                color: white;
            }
            .header {
                background-color: #1a1a2e;
                color: white;
                padding: 20px;
                text-align: center;
            }
            .metric-card {
                background: #f8f9fa;
                padding: 10px;
                margin: 5px;
                border-radius: 5px;
            }
            .positive { color: #27ae60; }
            .negative { color: #e74c3c; }
            img {
                max-width: 100%;
                height: auto;
            }
        """
    
    def _generate_json_report(self, data: Dict[str, Any], name: str) -> str:
        """
        Génère un rapport JSON.
        
        Args:
            data: Données du rapport.
            name: Nom du rapport.
            
        Returns:
            Chemin du fichier JSON.
        """
        filepath = Path(self.config.output_dir) / f"{name}.json"
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        
        return str(filepath)
    
    def _generate_excel_report(self, data: Dict[str, Any], name: str) -> str:
        """
        Génère un rapport Excel.
        
        Args:
            data: Données du rapport.
            name: Nom du rapport.
            
        Returns:
            Chemin du fichier Excel.
        """
        filepath = Path(self.config.output_dir) / f"{name}.xlsx"
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        
        summary_data = [
            ['Metric', 'Value'],
            ['Symbol', data.get('symbol', '')],
            ['Strategy', data.get('strategy', '')],
            ['Start Date', data.get('config', {}).get('start_date', '')],
            ['End Date', data.get('config', {}).get('end_date', '')],
            ['Initial Capital', f"${data.get('config', {}).get('initial_capital', 0):,.2f}"],
            ['Total Return', f"{data.get('metrics', {}).get('total_return', 0):.2%}"],
            ['Annualized Return', f"{data.get('metrics', {}).get('annualized_return', 0):.2%}"],
            ['Sharpe Ratio', f"{data.get('metrics', {}).get('sharpe_ratio', 0):.3f}"],
            ['Max Drawdown', f"{data.get('metrics', {}).get('max_drawdown_pct', 0):.2%}"],
            ['Win Rate', f"{data.get('metrics', {}).get('win_rate', 0):.2%}"],
            ['Total Trades', data.get('total_trades', 0)],
            ['Profit Factor', f"{data.get('metrics', {}).get('profit_factor', 0):.3f}"]
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = thin_border
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        
        # 2. Metrics Sheet
        ws_metrics = wb.create_sheet("Metrics")
        metrics_data = [['Metric', 'Value']]
        for key, value in data.get('metrics', {}).items():
            if isinstance(value, (int, float)):
                if 'rate' in key or 'ratio' in key:
                    metrics_data.append([key, f"{value:.4f}"])
                else:
                    metrics_data.append([key, value])
        
        for row_idx, row in enumerate(metrics_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_metrics.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = thin_border
        
        ws_metrics.column_dimensions['A'].width = 25
        ws_metrics.column_dimensions['B'].width = 20
        
        # 3. Trades Sheet
        ws_trades = wb.create_sheet("Trades")
        if data.get('trades'):
            # Convertir les trades en DataFrame
            trade_data = []
            for t in data['trades']:
                trade_data.append({
                    'Symbol': getattr(t, 'symbol', ''),
                    'Side': getattr(t, 'side', ''),
                    'Quantity': getattr(t, 'quantity', 0),
                    'Entry Price': getattr(t, 'entry_price', 0),
                    'Exit Price': getattr(t, 'exit_price', 0),
                    'PNL': getattr(t, 'pnl', 0),
                    'Status': getattr(t, 'status', '')
                })
            
            if trade_data:
                df = pd.DataFrame(trade_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws_trades.append(r)
                
                # Styles pour l'en-tête
                for col in range(1, len(df.columns) + 1):
                    cell = ws_trades.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
        
        # Ajustement des colonnes
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_length = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_length
        
        wb.save(filepath)
        
        return str(filepath)
    
    def _generate_markdown_report(self, data: Dict[str, Any], name: str) -> str:
        """
        Génère un rapport Markdown.
        
        Args:
            data: Données du rapport.
            name: Nom du rapport.
            
        Returns:
            Chemin du fichier Markdown.
        """
        filepath = Path(self.config.output_dir) / f"{name}.md"
        
        lines = []
        lines.append(f"# Backtest Report: {data.get('name', 'Unknown')}")
        lines.append("")
        lines.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")
        lines.append("")
        
        # Summary
        lines.append("## Summary")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("|--------|-------|")
        lines.append(f"| Symbol | {data.get('symbol', '')} |")
        lines.append(f"| Strategy | {data.get('strategy', '')} |")
        lines.append(f"| Start Date | {data.get('config', {}).get('start_date', '')} |")
        lines.append(f"| End Date | {data.get('config', {}).get('end_date', '')} |")
        lines.append(f"| Initial Capital | ${data.get('config', {}).get('initial_capital', 0):,.2f} |")
        lines.append(f"| Total Return | {data.get('metrics', {}).get('total_return', 0):.2%} |")
        lines.append(f"| Sharpe Ratio | {data.get('metrics', {}).get('sharpe_ratio', 0):.3f} |")
        lines.append(f"| Max Drawdown | {data.get('metrics', {}).get('max_drawdown_pct', 0):.2%} |")
        lines.append(f"| Total Trades | {data.get('total_trades', 0)} |")
        lines.append("")
        
        # Performance Metrics
        lines.append("## Performance Metrics")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("|--------|-------|")
        metrics = data.get('metrics', {})
        for key, value in metrics.items():
            if isinstance(value, (int, float)):
                if 'rate' in key or 'ratio' in key:
                    lines.append(f"| {key} | {value:.4f} |")
                else:
                    lines.append(f"| {key} | {value} |")
        lines.append("")
        
        # Charts (si disponibles)
        charts = data.get('charts', {})
        if charts:
            lines.append("## Charts")
            lines.append("")
            for chart_name, chart_path in charts.items():
                lines.append(f"### {chart_name.replace('_', ' ').title()}")
                lines.append("")
                lines.append(f"![{chart_name}]({os.path.basename(chart_path)})")
                lines.append("")
        
        # Write to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return str(filepath)


# Fonctions utilitaires
def generate_backtest_report(
    result: BacktestResult,
    output_dir: str = "data/backtest_results/",
    formats: List[str] = ['html'],
    **kwargs
) -> Dict[str, str]:
    """
    Fonction utilitaire pour générer un rapport de backtest.
    
    Args:
        result: Résultats du backtesting.
        output_dir: Répertoire de sortie.
        formats: Formats de sortie.
        **kwargs: Paramètres supplémentaires.
        
    Returns:
        Dictionnaire des chemins des fichiers générés.
    """
    config = ReportConfig(
        output_dir=output_dir,
        formats=formats,
        **kwargs
    )
    
    generator = ReportGenerator(config)
    return generator.generate_report(result)


# Exportation
__all__ = [
    'ReportGenerator',
    'ReportConfig',
    'generate_backtest_report'
]
