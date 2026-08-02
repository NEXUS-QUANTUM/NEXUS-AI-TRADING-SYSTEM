# trading/bots/hedge_bot/hedge_bot_data_report.py

import asyncio
import logging
import time
import json
import hashlib
import base64
import csv
import io
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from typing import Any, Dict, List, Optional, Set, Union, Tuple, Callable
from decimal import Decimal
from collections import defaultdict, deque
import pandas as pd
import numpy as np

try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib.units import inch, mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)


class ReportType(str, Enum):
    PERFORMANCE = "performance"
    RISK = "risk"
    TRADING = "trading"
    PORTFOLIO = "portfolio"
    ANALYTICS = "analytics"
    AUDIT = "audit"
    COMPLIANCE = "compliance"
    SUMMARY = "summary"
    DETAILED = "detailed"
    EXECUTIVE = "executive"
    TECHNICAL = "technical"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    CUSTOM = "custom"


class ReportFormat(str, Enum):
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"
    TEXT = "text"
    PNG = "png"
    SVG = "svg"


class ReportStatus(str, Enum):
    DRAFT = "draft"
    GENERATING = "generating"
    COMPLETED = "completed"
    FAILED = "failed"
    DELIVERED = "delivered"
    VIEWED = "viewed"
    EXPIRED = "expired"


@dataclass
class ReportSection:
    id: str
    title: str
    type: str
    data: Any
    order: int
    metadata: Dict[str, Any] = field(default_factory=dict)
    charts: List[str] = field(default_factory=list)
    tables: List[str] = field(default_factory=list)


@dataclass
class Report:
    id: str
    name: str
    type: ReportType
    format: ReportFormat
    status: ReportStatus
    created_at: float
    updated_at: float
    generated_at: Optional[float] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    sections: List[ReportSection] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    tags: List[str] = field(default_factory=list)
    version: str = "1.0.0"
    content: Optional[bytes] = None
    content_type: Optional[str] = None
    size: int = 0
    expires_at: Optional[float] = None
    delivery_methods: List[str] = field(default_factory=list)
    recipients: List[str] = field(default_factory=list)
    parameters: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportTemplate:
    id: str
    name: str
    description: str
    type: ReportType
    sections: List[Dict[str, Any]]
    styles: Dict[str, Any]
    layout: Dict[str, Any]
    metadata: Dict[str, Any] = field(default_factory=dict)
    created_at: float = field(default_factory=time.time)
    updated_at: float = field(default_factory=time.time)


@dataclass
class ReportSchedule:
    id: str
    report_id: str
    frequency: str
    interval: int
    next_run: float
    last_run: Optional[float] = None
    active: bool = True
    recipients: List[str] = field(default_factory=list)
    parameters: Dict[str, Any] = field(default_factory=dict)
    created_at: float = field(default_factory=time.time)
    updated_at: float = field(default_factory=time.time)


class ReportManager:
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self._lock = asyncio.Lock()
        self._reports: Dict[str, Report] = {}
        self._templates: Dict[str, ReportTemplate] = {}
        self._schedules: Dict[str, ReportSchedule] = {}
        self._generators: Dict[ReportType, Callable] = {}
        self._formatters: Dict[ReportFormat, Callable] = {}
        self._validators: List[Callable] = []
        self._observers: List[Callable] = []
        self._running = False
        
        self._initialize_generators()
        self._initialize_formatters()
        self._initialize_templates()

    def _initialize_generators(self) -> None:
        self.register_generator(ReportType.PERFORMANCE, self._generate_performance_report)
        self.register_generator(ReportType.RISK, self._generate_risk_report)
        self.register_generator(ReportType.TRADING, self._generate_trading_report)
        self.register_generator(ReportType.PORTFOLIO, self._generate_portfolio_report)
        self.register_generator(ReportType.ANALYTICS, self._generate_analytics_report)
        self.register_generator(ReportType.AUDIT, self._generate_audit_report)
        self.register_generator(ReportType.COMPLIANCE, self._generate_compliance_report)
        self.register_generator(ReportType.SUMMARY, self._generate_summary_report)
        self.register_generator(ReportType.DAILY, self._generate_daily_report)
        self.register_generator(ReportType.WEEKLY, self._generate_weekly_report)
        self.register_generator(ReportType.MONTHLY, self._generate_monthly_report)

    def _initialize_formatters(self) -> None:
        self.register_formatter(ReportFormat.PDF, self._format_pdf)
        self.register_formatter(ReportFormat.HTML, self._format_html)
        self.register_formatter(ReportFormat.JSON, self._format_json)
        self.register_formatter(ReportFormat.CSV, self._format_csv)
        self.register_formatter(ReportFormat.EXCEL, self._format_excel)
        self.register_formatter(ReportFormat.MARKDOWN, self._format_markdown)
        self.register_formatter(ReportFormat.TEXT, self._format_text)

    def _initialize_templates(self) -> None:
        default_template = ReportTemplate(
            id="default",
            name="Default Report Template",
            description="Standard report template",
            type=ReportType.SUMMARY,
            sections=[
                {"title": "Executive Summary", "type": "summary", "order": 0},
                {"title": "Key Metrics", "type": "metrics", "order": 1},
                {"title": "Performance Analysis", "type": "performance", "order": 2},
                {"title": "Risk Analysis", "type": "risk", "order": 3},
                {"title": "Conclusion", "type": "conclusion", "order": 4}
            ],
            styles={
                "primary_color": "#1a1a2e",
                "secondary_color": "#16213e",
                "accent_color": "#0f3460",
                "highlight_color": "#e94560",
                "font_family": "Arial, Helvetica, sans-serif",
                "font_size": 11,
                "header_font_size": 18,
                "subheader_font_size": 14
            },
            layout={
                "page_size": "A4",
                "margin_top": 72,
                "margin_bottom": 72,
                "margin_left": 72,
                "margin_right": 72,
                "show_header": True,
                "show_footer": True,
                "show_page_numbers": True,
                "show_table_of_contents": True
            }
        )
        
        self._templates[default_template.id] = default_template

    def register_generator(self, report_type: ReportType, generator: Callable) -> None:
        self._generators[report_type] = generator

    def register_formatter(self, format_type: ReportFormat, formatter: Callable) -> None:
        self._formatters[format_type] = formatter

    def register_validator(self, validator: Callable) -> None:
        self._validators.append(validator)

    def register_observer(self, observer: Callable) -> None:
        self._observers.append(observer)

    async def create_report(
        self,
        name: str,
        report_type: ReportType,
        format_type: ReportFormat,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        template_id: str = "default",
        parameters: Optional[Dict[str, Any]] = None,
        metadata: Optional[Dict[str, Any]] = None,
        tags: Optional[List[str]] = None,
        recipients: Optional[List[str]] = None
    ) -> Report:
        async with self._lock:
            report_id = hashlib.md5(f"{name}_{time.time()}".encode()).hexdigest()
            
            template = self._templates.get(template_id, self._templates["default"])
            
            sections = []
            for i, section_template in enumerate(template.sections):
                section = ReportSection(
                    id=hashlib.md5(f"{section_template['title']}_{i}".encode()).hexdigest(),
                    title=section_template.get("title", f"Section {i}"),
                    type=section_template.get("type", "generic"),
                    data={},
                    order=section_template.get("order", i),
                    metadata=section_template.get("metadata", {})
                )
                sections.append(section)
            
            report = Report(
                id=report_id,
                name=name,
                type=report_type,
                format=format_type,
                status=ReportStatus.DRAFT,
                created_at=time.time(),
                updated_at=time.time(),
                start_date=start_date,
                end_date=end_date,
                sections=sections,
                metadata=metadata or {},
                tags=tags or [],
                parameters=parameters or {},
                recipients=recipients or []
            )
            
            self._reports[report_id] = report
            await self._notify_observers("report_created", report)
            
            return report

    async def generate_report(self, report_id: str, force: bool = False) -> Optional[Report]:
        async with self._lock:
            if report_id not in self._reports:
                return None
            
            report = self._reports[report_id]
            
            if report.status == ReportStatus.GENERATING:
                return None
            
            if report.status == ReportStatus.COMPLETED and not force:
                return report
            
            if report.type not in self._generators:
                logger.error(f"No generator for report type: {report.type}")
                return None
            
            report.status = ReportStatus.GENERATING
            report.updated_at = time.time()
            
            try:
                generator = self._generators[report.type]
                data = await generator(report)
                
                for section in report.sections:
                    if section.type in data:
                        section.data = data[section.type]
                
                formatter = self._formatters.get(report.format)
                if formatter:
                    content = await formatter(report)
                    report.content = content
                    report.content_type = self._get_content_type(report.format)
                    report.size = len(content)
                
                report.status = ReportStatus.COMPLETED
                report.generated_at = time.time()
                report.updated_at = time.time()
                
                await self._notify_observers("report_generated", report)
                return report
                
            except Exception as e:
                logger.error(f"Error generating report: {e}")
                report.status = ReportStatus.FAILED
                report.metadata["error"] = str(e)
                return None

    async def get_report(self, report_id: str) -> Optional[Report]:
        return self._reports.get(report_id)

    async def get_reports(
        self,
        report_type: Optional[ReportType] = None,
        status: Optional[ReportStatus] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        limit: int = 100
    ) -> List[Report]:
        reports = list(self._reports.values())
        
        if report_type:
            reports = [r for r in reports if r.type == report_type]
        
        if status:
            reports = [r for r in reports if r.status == status]
        
        if start_date:
            reports = [r for r in reports if r.created_at >= start_date.timestamp()]
        
        if end_date:
            reports = [r for r in reports if r.created_at <= end_date.timestamp()]
        
        reports.sort(key=lambda r: r.created_at, reverse=True)
        return reports[:limit]

    async def delete_report(self, report_id: str) -> bool:
        async with self._lock:
            if report_id not in self._reports:
                return False
            
            del self._reports[report_id]
            await self._notify_observers("report_deleted", report_id)
            return True

    async def create_template(
        self,
        name: str,
        description: str,
        report_type: ReportType,
        sections: List[Dict[str, Any]],
        styles: Dict[str, Any],
        layout: Dict[str, Any],
        metadata: Optional[Dict[str, Any]] = None
    ) -> ReportTemplate:
        template = ReportTemplate(
            id=hashlib.md5(f"{name}_{time.time()}".encode()).hexdigest(),
            name=name,
            description=description,
            type=report_type,
            sections=sections,
            styles=styles,
            layout=layout,
            metadata=metadata or {}
        )
        
        self._templates[template.id] = template
        return template

    async def get_template(self, template_id: str) -> Optional[ReportTemplate]:
        return self._templates.get(template_id)

    async def delete_template(self, template_id: str) -> bool:
        if template_id in self._templates and template_id != "default":
            del self._templates[template_id]
            return True
        return False

    async def _generate_performance_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "total_pnl": 0.0,
                "win_rate": 0.0,
                "sharpe_ratio": 0.0,
                "max_drawdown": 0.0,
                "total_trades": 0
            },
            "metrics": {
                "daily_pnl": [],
                "weekly_pnl": [],
                "monthly_pnl": [],
                "cumulative_pnl": []
            },
            "performance": {
                "returns": [],
                "volatility": 0.0,
                "benchmark_returns": []
            },
            "trades": {
                "winning_trades": 0,
                "losing_trades": 0,
                "avg_win": 0.0,
                "avg_loss": 0.0,
                "profit_factor": 0.0
            }
        }
        return data

    async def _generate_risk_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "var_95": 0.0,
                "var_99": 0.0,
                "cvar_95": 0.0,
                "cvar_99": 0.0,
                "expected_shortfall": 0.0
            },
            "metrics": {
                "daily_var": [],
                "weekly_var": [],
                "monthly_var": []
            },
            "risk_factors": {
                "market_risk": 0.0,
                "credit_risk": 0.0,
                "liquidity_risk": 0.0,
                "operational_risk": 0.0
            },
            "stress_tests": {
                "worst_case": 0.0,
                "best_case": 0.0,
                "scenarios": []
            }
        }
        return data

    async def _generate_trading_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "total_trades": 0,
                "open_positions": 0,
                "closed_positions": 0,
                "volume": 0.0
            },
            "trades": [],
            "orders": [],
            "positions": [],
            "execution": {
                "avg_execution_time": 0.0,
                "slippage": 0.0,
                "fill_rate": 0.0
            }
        }
        return data

    async def _generate_portfolio_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "total_value": 0.0,
                "cash": 0.0,
                "invested": 0.0,
                "returns": 0.0
            },
            "allocation": {
                "by_asset": {},
                "by_strategy": {},
                "by_risk": {}
            },
            "performance": {
                "daily": [],
                "monthly": [],
                "yearly": []
            },
            "holdings": []
        }
        return data

    async def _generate_analytics_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "total_visits": 0,
                "unique_users": 0,
                "avg_session": 0.0,
                "conversion_rate": 0.0
            },
            "metrics": {
                "daily_active_users": [],
                "weekly_active_users": [],
                "monthly_active_users": []
            },
            "analytics": {
                "user_behavior": {},
                "trading_patterns": {},
                "market_trends": {}
            },
            "insights": []
        }
        return data

    async def _generate_audit_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "total_events": 0,
                "critical_events": 0,
                "warning_events": 0,
                "info_events": 0
            },
            "events": [],
            "compliance": {
                "passed": 0,
                "failed": 0,
                "warnings": 0
            },
            "recommendations": []
        }
        return data

    async def _generate_compliance_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "status": "compliant",
                "issues": 0,
                "critical_issues": 0
            },
            "checks": [],
            "violations": [],
            "actions": []
        }
        return data

    async def _generate_summary_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "summary": {
                "overall_status": "positive",
                "key_metrics": {},
                "highlights": [],
                "risks": [],
                "recommendations": []
            },
            "details": {}
        }
        return data

    async def _generate_daily_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "date": datetime.now().isoformat(),
            "summary": {
                "total_pnl": 0.0,
                "trades": 0,
                "volume": 0.0
            },
            "highlights": [],
            "issues": []
        }
        return data

    async def _generate_weekly_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "week": datetime.now().strftime("%Y-W%W"),
            "summary": {
                "total_pnl": 0.0,
                "trades": 0,
                "volume": 0.0,
                "win_rate": 0.0
            },
            "daily_breakdown": [],
            "insights": []
        }
        return data

    async def _generate_monthly_report(self, report: Report) -> Dict[str, Any]:
        data = {
            "month": datetime.now().strftime("%Y-%m"),
            "summary": {
                "total_pnl": 0.0,
                "trades": 0,
                "volume": 0.0,
                "win_rate": 0.0,
                "sharpe_ratio": 0.0
            },
            "weekly_breakdown": [],
            "performance": {},
            "insights": []
        }
        return data

    async def _format_pdf(self, report: Report) -> bytes:
        if not REPORTLAB_AVAILABLE:
            logger.warning("ReportLab not installed, falling back to text format")
            return await self._format_text(report)
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            styles = getSampleStyleSheet()
            story = []
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a2e'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            story.append(Paragraph(f"{report.name}", title_style))
            story.append(Spacer(1, 0.25 * inch))
            
            info_style = styles['Normal']
            story.append(Paragraph(f"Report Type: {report.type.value}", info_style))
            story.append(Paragraph(f"Generated: {datetime.fromtimestamp(report.generated_at or time.time()).strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Spacer(1, 0.25 * inch))
            
            for section in sorted(report.sections, key=lambda s: s.order):
                section_style = ParagraphStyle(
                    'SectionTitle',
                    parent=styles['Heading2'],
                    fontSize=16,
                    textColor=colors.HexColor('#16213e'),
                    spaceAfter=12
                )
                
                story.append(Paragraph(section.title, section_style))
                
                if isinstance(section.data, dict):
                    data_items = []
                    for key, value in section.data.items():
                        data_items.append([str(key), str(value)])
                    
                    if data_items:
                        data_table = Table(data_items, colWidths=[3 * inch, 3 * inch])
                        data_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
                            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('LEFTPADDING', (0, 0), (-1, -1), 6),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
                        ]))
                        story.append(data_table)
                        story.append(Spacer(1, 0.25 * inch))
                
                story.append(Spacer(1, 0.15 * inch))
            
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"PDF generation error: {e}")
            return await self._format_text(report)

    async def _format_html(self, report: Report) -> bytes:
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{report.name}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
                h1 {{ color: #1a1a2e; border-bottom: 2px solid #1a1a2e; padding-bottom: 10px; }}
                h2 {{ color: #16213e; margin-top: 30px; }}
                .section {{ margin-bottom: 30px; padding: 20px; background: #f8f9fa; border-radius: 5px; }}
                .section-title {{ font-size: 18px; font-weight: bold; color: #16213e; }}
                .data {{ margin-top: 10px; }}
                .data-item {{ display: flex; justify-content: space-between; padding: 5px 0; border-bottom: 1px solid #eee; }}
                .data-key {{ font-weight: bold; color: #555; }}
                .data-value {{ color: #333; }}
                .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #999; }}
            </style>
        </head>
        <body>
            <h1>{report.name}</h1>
            <p><strong>Type:</strong> {report.type.value}</p>
            <p><strong>Generated:</strong> {datetime.fromtimestamp(report.generated_at or time.time()).strftime('%Y-%m-%d %H:%M:%S')}</p>
        """
        
        for section in sorted(report.sections, key=lambda s: s.order):
            html += f"""
            <div class="section">
                <div class="section-title">{section.title}</div>
                <div class="data">
            """
            
            if isinstance(section.data, dict):
                for key, value in section.data.items():
                    html += f"""
                    <div class="data-item">
                        <span class="data-key">{key}</span>
                        <span class="data-value">{value}</span>
                    </div>
                    """
            
            html += """
                </div>
            </div>
            """
        
        html += f"""
            <div class="footer">
                Generated by Report Manager v{report.version}
            </div>
        </body>
        </html>
        """
        
        return html.encode()

    async def _format_json(self, report: Report) -> bytes:
        data = {
            "id": report.id,
            "name": report.name,
            "type": report.type.value,
            "format": report.format.value,
            "status": report.status.value,
            "created_at": datetime.fromtimestamp(report.created_at).isoformat(),
            "updated_at": datetime.fromtimestamp(report.updated_at).isoformat(),
            "sections": [
                {
                    "id": s.id,
                    "title": s.title,
                    "type": s.type,
                    "data": s.data,
                    "order": s.order
                }
                for s in report.sections
            ],
            "metadata": report.metadata,
            "tags": report.tags,
            "version": report.version
        }
        
        return json.dumps(data, indent=2, default=str).encode()

    async def _format_csv(self, report: Report) -> bytes:
        output = io.StringIO()
        
        for section in sorted(report.sections, key=lambda s: s.order):
            output.write(f"Section: {section.title}\n")
            
            if isinstance(section.data, dict):
                writer = csv.writer(output)
                for key, value in section.data.items():
                    if isinstance(value, (list, dict)):
                        writer.writerow([key, json.dumps(value, default=str)])
                    else:
                        writer.writerow([key, str(value)])
            
            output.write("\n")
        
        return output.getvalue().encode()

    async def _format_excel(self, report: Report) -> bytes:
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            
            for i, section in enumerate(sorted(report.sections, key=lambda s: s.order)):
                ws = wb.create_sheet(title=section.title[:31], index=i)
                
                ws['A1'] = section.title
                ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
                
                if isinstance(section.data, dict):
                    row = 3
                    for key, value in section.data.items():
                        ws[f'A{row}'] = key
                        ws[f'B{row}'] = str(value)
                        row += 1
                
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 50
            
            wb.remove(wb['Sheet'])
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except ImportError:
            return await self._format_csv(report)

    async def _format_markdown(self, report: Report) -> bytes:
        md = f"# {report.name}\n\n"
        md += f"**Type:** {report.type.value}\n\n"
        md += f"**Generated:** {datetime.fromtimestamp(report.generated_at or time.time()).strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        for section in sorted(report.sections, key=lambda s: s.order):
            md += f"## {section.title}\n\n"
            
            if isinstance(section.data, dict):
                for key, value in section.data.items():
                    md += f"- **{key}:** {value}\n"
            
            md += "\n"
        
        return md.encode()

    async def _format_text(self, report: Report) -> bytes:
        text = f"{report.name}\n"
        text += f"{'=' * len(report.name)}\n\n"
        text += f"Type: {report.type.value}\n"
        text += f"Generated: {datetime.fromtimestamp(report.generated_at or time.time()).strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        for section in sorted(report.sections, key=lambda s: s.order):
            text += f"{section.title}\n"
            text += f"{'-' * len(section.title)}\n"
            
            if isinstance(section.data, dict):
                for key, value in section.data.items():
                    text += f"{key}: {value}\n"
            
            text += "\n"
        
        return text.encode()

    def _get_content_type(self, format_type: ReportFormat) -> str:
        content_types = {
            ReportFormat.PDF: "application/pdf",
            ReportFormat.HTML: "text/html",
            ReportFormat.JSON: "application/json",
            ReportFormat.CSV: "text/csv",
            ReportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ReportFormat.MARKDOWN: "text/markdown",
            ReportFormat.TEXT: "text/plain",
            ReportFormat.PNG: "image/png",
            ReportFormat.SVG: "image/svg+xml"
        }
        return content_types.get(format_type, "application/octet-stream")

    async def _notify_observers(self, event: str, *args) -> None:
        for observer in self._observers:
            try:
                if asyncio.iscoroutinefunction(observer):
                    await observer(event, *args)
                else:
                    observer(event, *args)
            except Exception as e:
                logger.error(f"Observer error: {e}")

    async def schedule_report(
        self,
        report_id: str,
        frequency: str,
        interval: int,
        recipients: List[str],
        parameters: Optional[Dict[str, Any]] = None
    ) -> Optional[ReportSchedule]:
        if report_id not in self._reports:
            return None
        
        schedule = ReportSchedule(
            id=hashlib.md5(f"{report_id}_{time.time()}".encode()).hexdigest(),
            report_id=report_id,
            frequency=frequency,
            interval=interval,
            next_run=time.time() + self._get_interval_seconds(frequency, interval),
            recipients=recipients,
            parameters=parameters or {}
        )
        
        self._schedules[schedule.id] = schedule
        return schedule

    def _get_interval_seconds(self, frequency: str, interval: int) -> int:
        if frequency == "minutes":
            return interval * 60
        elif frequency == "hours":
            return interval * 3600
        elif frequency == "days":
            return interval * 86400
        elif frequency == "weeks":
            return interval * 604800
        elif frequency == "months":
            return interval * 2592000
        elif frequency == "years":
            return interval * 31536000
        else:
            return interval * 86400

    async def process_schedules(self) -> None:
        now = time.time()
        
        for schedule in self._schedules.values():
            if not schedule.active:
                continue
            
            if schedule.next_run > now:
                continue
            
            report = self._reports.get(schedule.report_id)
            if not report:
                continue
            
            await self.generate_report(report.id)
            
            schedule.last_run = now
            schedule.next_run = now + self._get_interval_seconds(schedule.frequency, schedule.interval)
            schedule.updated_at = time.time()

    def get_stats(self) -> Dict[str, Any]:
        return {
            "reports": len(self._reports),
            "templates": len(self._templates),
            "schedules": len(self._schedules),
            "generators": len(self._generators),
            "formatters": len(self._formatters),
            "validators": len(self._validators),
            "observers": len(self._observers),
            "running": self._running
        }


__all__ = [
    "ReportType",
    "ReportFormat",
    "ReportStatus",
    "ReportSection",
    "Report",
    "ReportTemplate",
    "ReportSchedule",
    "ReportManager"
]
