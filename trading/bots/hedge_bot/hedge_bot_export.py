# trading/bots/hedge_bot/hedge_bot_export.py

import asyncio
import logging
import time
import json
import csv
import io
import os
import hashlib
import base64
import zipfile
import tarfile
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from typing import Any, Dict, List, Optional, Set, Union, Tuple, Callable, BinaryIO
from decimal import Decimal
from collections import defaultdict
import pandas as pd
import numpy as np

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import aiohttp
    AIOHTTP_AVAILABLE = True
except ImportError:
    AIOHTTP_AVAILABLE = False

try:
    import boto3
    BOTO3_AVAILABLE = True
except ImportError:
    BOTO3_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    PARQUET = "parquet"
    FEATHER = "feather"
    HDF5 = "hdf5"
    PICKLE = "pickle"
    ORC = "orc"
    XML = "xml"
    YAML = "yaml"
    HTML = "html"
    MARKDOWN = "markdown"
    PDF = "pdf"
    TEXT = "text"
    PROTOBUF = "protobuf"
    ARROW = "arrow"
    SQL = "sql"


class ExportCompression(str, Enum):
    NONE = "none"
    GZIP = "gzip"
    ZIP = "zip"
    TAR = "tar"
    TAR_GZ = "tar_gz"
    BZ2 = "bz2"
    XZ = "xz"


class ExportDestination(str, Enum):
    FILE = "file"
    S3 = "s3"
    GCS = "gcs"
    AZURE = "azure"
    FTP = "ftp"
    SFTP = "sftp"
    WEBHOOK = "webhook"
    DATABASE = "database"
    EMAIL = "email"
    PUSH = "push"
    STREAM = "stream"


class ExportStatus(str, Enum):
    PENDING = "pending"
    RUNNING = "running"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"
    PARTIAL = "partial"


@dataclass
class ExportConfig:
    id: str
    name: str
    format: ExportFormat
    compression: ExportCompression
    destination: ExportDestination
    destination_config: Dict[str, Any]
    include_headers: bool = True
    include_index: bool = False
    batch_size: int = 10000
    max_file_size: int = 100 * 1024 * 1024
    delimiter: str = ","
    encoding: str = "utf-8"
    date_format: str = "%Y-%m-%d %H:%M:%S"
    metadata: Dict[str, Any] = field(default_factory=dict)
    created_at: float = field(default_factory=time.time)
    updated_at: float = field(default_factory=time.time)


@dataclass
class ExportJob:
    id: str
    config_id: str
    status: ExportStatus
    total_rows: int = 0
    exported_rows: int = 0
    failed_rows: int = 0
    total_size: int = 0
    exported_size: int = 0
    start_time: float = field(default_factory=time.time)
    end_time: Optional[float] = None
    error: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    files: List[str] = field(default_factory=list)
    progress: float = 0.0


@dataclass
class ExportFile:
    id: str
    job_id: str
    name: str
    path: str
    size: int
    format: ExportFormat
    compression: ExportCompression
    row_count: int
    checksum: str
    url: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    created_at: float = field(default_factory=time.time)


class DataExporter:
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self._lock = asyncio.Lock()
        self._configs: Dict[str, ExportConfig] = {}
        self._jobs: Dict[str, ExportJob] = {}
        self._files: Dict[str, ExportFile] = {}
        self._formatters: Dict[ExportFormat, Callable] = {}
        self._destinations: Dict[ExportDestination, Callable] = {}
        self._observers: List[Callable] = []
        self._running = False
        self._temp_dir = tempfile.mkdtemp()
        
        self._initialize_formatters()
        self._initialize_destinations()

    def _initialize_formatters(self) -> None:
        self.register_formatter(ExportFormat.CSV, self._format_csv)
        self.register_formatter(ExportFormat.JSON, self._format_json)
        self.register_formatter(ExportFormat.EXCEL, self._format_excel)
        self.register_formatter(ExportFormat.PARQUET, self._format_parquet)
        self.register_formatter(ExportFormat.FEATHER, self._format_feather)
        self.register_formatter(ExportFormat.HDF5, self._format_hdf5)
        self.register_formatter(ExportFormat.PICKLE, self._format_pickle)
        self.register_formatter(ExportFormat.ORC, self._format_orc)
        self.register_formatter(ExportFormat.XML, self._format_xml)
        self.register_formatter(ExportFormat.YAML, self._format_yaml)
        self.register_formatter(ExportFormat.HTML, self._format_html)
        self.register_formatter(ExportFormat.MARKDOWN, self._format_markdown)
        self.register_formatter(ExportFormat.TEXT, self._format_text)
        self.register_formatter(ExportFormat.SQL, self._format_sql)

    def _initialize_destinations(self) -> None:
        self.register_destination(ExportDestination.FILE, self._destination_file)
        self.register_destination(ExportDestination.S3, self._destination_s3)
        self.register_destination(ExportDestination.GCS, self._destination_gcs)
        self.register_destination(ExportDestination.AZURE, self._destination_azure)
        self.register_destination(ExportDestination.FTP, self._destination_ftp)
        self.register_destination(ExportDestination.SFTP, self._destination_sftp)
        self.register_destination(ExportDestination.WEBHOOK, self._destination_webhook)
        self.register_destination(ExportDestination.DATABASE, self._destination_database)
        self.register_destination(ExportDestination.EMAIL, self._destination_email)

    def register_formatter(self, format_type: ExportFormat, formatter: Callable) -> None:
        self._formatters[format_type] = formatter

    def register_destination(self, destination: ExportDestination, dest_handler: Callable) -> None:
        self._destinations[destination] = dest_handler

    def register_observer(self, observer: Callable) -> None:
        self._observers.append(observer)

    async def create_config(
        self,
        name: str,
        format: ExportFormat,
        destination: ExportDestination,
        destination_config: Dict[str, Any],
        compression: ExportCompression = ExportCompression.NONE,
        include_headers: bool = True,
        include_index: bool = False,
        batch_size: int = 10000,
        max_file_size: int = 100 * 1024 * 1024,
        delimiter: str = ",",
        encoding: str = "utf-8",
        metadata: Optional[Dict[str, Any]] = None
    ) -> ExportConfig:
        async with self._lock:
            config_id = hashlib.md5(f"{name}_{time.time()}".encode()).hexdigest()
            
            config = ExportConfig(
                id=config_id,
                name=name,
                format=format,
                compression=compression,
                destination=destination,
                destination_config=destination_config,
                include_headers=include_headers,
                include_index=include_index,
                batch_size=batch_size,
                max_file_size=max_file_size,
                delimiter=delimiter,
                encoding=encoding,
                metadata=metadata or {}
            )
            
            self._configs[config_id] = config
            await self._notify_observers("config_created", config)
            return config

    async def export_data(
        self,
        config_id: str,
        data: Union[pd.DataFrame, List[Dict], Dict],
        job_metadata: Optional[Dict[str, Any]] = None
    ) -> ExportJob:
        async with self._lock:
            if config_id not in self._configs:
                raise ValueError(f"Config not found: {config_id}")
            
            config = self._configs[config_id]
            
            if isinstance(data, list):
                data = pd.DataFrame(data)
            elif isinstance(data, dict):
                data = pd.DataFrame([data])
            elif not isinstance(data, pd.DataFrame):
                raise ValueError("Data must be DataFrame, list, or dict")
            
            job_id = hashlib.md5(f"{config_id}_{time.time()}".encode()).hexdigest()
            
            job = ExportJob(
                id=job_id,
                config_id=config_id,
                status=ExportStatus.RUNNING,
                total_rows=len(data),
                metadata=job_metadata or {}
            )
            
            self._jobs[job_id] = job
            await self._notify_observers("job_started", job)
            
            try:
                formatter = self._formatters.get(config.format)
                if not formatter:
                    raise ValueError(f"Unsupported format: {config.format}")
                
                destination = self._destinations.get(config.destination)
                if not destination:
                    raise ValueError(f"Unsupported destination: {config.destination}")
                
                total_rows = len(data)
                rows_processed = 0
                
                if total_rows > config.batch_size:
                    for start in range(0, total_rows, config.batch_size):
                        end = min(start + config.batch_size, total_rows)
                        batch = data.iloc[start:end]
                        
                        formatted_data = await formatter(batch, config)
                        compressed_data = await self._compress_data(formatted_data, config.compression)
                        
                        file_name = await self._save_file(job_id, start, end, formatted_data, config)
                        
                        if config.max_file_size > 0 and len(compressed_data) > config.max_file_size:
                            await self._handle_large_file(job, batch, config)
                        
                        rows_processed += len(batch)
                        job.exported_rows = rows_processed
                        job.progress = (rows_processed / total_rows) * 100
                        job.exported_size += len(compressed_data)
                        
                        await self._notify_observers("job_progress", job)
                else:
                    formatted_data = await formatter(data, config)
                    compressed_data = await self._compress_data(formatted_data, config.compression)
                    
                    file_path = await self._save_file(job_id, 0, total_rows, formatted_data, config)
                    job.files.append(file_path)
                    job.exported_rows = total_rows
                    job.progress = 100.0
                    job.exported_size = len(compressed_data)
                
                result = await destination(job, config)
                
                job.status = ExportStatus.COMPLETED
                job.end_time = time.time()
                
                await self._notify_observers("job_completed", job)
                
            except Exception as e:
                logger.error(f"Export failed: {e}")
                job.status = ExportStatus.FAILED
                job.error = str(e)
                job.end_time = time.time()
                await self._notify_observers("job_failed", job)
            
            return job

    async def _format_csv(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        output = io.StringIO()
        data.to_csv(
            output,
            sep=config.delimiter,
            header=config.include_headers,
            index=config.include_index,
            date_format=config.date_format,
            encoding=config.encoding
        )
        return output.getvalue().encode(config.encoding)

    async def _format_json(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        records = data.to_dict(orient='records')
        json_str = json.dumps(records, default=str, indent=2)
        return json_str.encode(config.encoding)

    async def _format_excel(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not available")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        if config.include_headers:
            for col_idx, col_name in enumerate(data.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True)
        
        for row_idx, row in enumerate(data.values, start=2 if config.include_headers else 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def _format_parquet(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        import pyarrow as pa
        import pyarrow.parquet as pq
        
        table = pa.Table.from_pandas(data)
        buffer = io.BytesIO()
        pq.write_table(table, buffer)
        return buffer.getvalue()

    async def _format_feather(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        import pyarrow.feather as feather
        
        buffer = io.BytesIO()
        feather.write_feather(data, buffer)
        return buffer.getvalue()

    async def _format_hdf5(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        buffer = io.BytesIO()
        data.to_hdf(buffer, key='data', mode='w')
        return buffer.getvalue()

    async def _format_pickle(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        import pickle
        return pickle.dumps(data)

    async def _format_orc(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        import pyarrow.orc as orc
        
        table = pa.Table.from_pandas(data)
        buffer = io.BytesIO()
        orc.write_table(table, buffer)
        return buffer.getvalue()

    async def _format_xml(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        xml = '<?xml version="1.0" encoding="UTF-8"?>\n<data>\n'
        for _, row in data.iterrows():
            xml += '  <row>\n'
            for col in data.columns:
                xml += f'    <{col}>{row[col]}</{col}>\n'
            xml += '  </row>\n'
        xml += '</data>'
        return xml.encode(config.encoding)

    async def _format_yaml(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        import yaml
        records = data.to_dict(orient='records')
        yaml_str = yaml.dump(records, default_flow_style=False)
        return yaml_str.encode(config.encoding)

    async def _format_html(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        html = data.to_html(
            classes='table table-striped',
            index=config.include_index,
            max_rows=None,
            max_cols=None
        )
        return html.encode(config.encoding)

    async def _format_markdown(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        markdown = data.to_markdown(index=config.include_index)
        return markdown.encode(config.encoding)

    async def _format_text(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        text = data.to_string(index=config.include_index)
        return text.encode(config.encoding)

    async def _format_sql(self, data: pd.DataFrame, config: ExportConfig) -> bytes:
        table_name = config.metadata.get("table_name", "data")
        sql = f"CREATE TABLE {table_name} (\n"
        
        for col in data.columns:
            dtype = data[col].dtype
            if dtype == 'object':
                sql += f"  {col} TEXT,\n"
            elif dtype == 'int64':
                sql += f"  {col} INTEGER,\n"
            elif dtype == 'float64':
                sql += f"  {col} REAL,\n"
            elif dtype == 'datetime64[ns]':
                sql += f"  {col} DATETIME,\n"
            else:
                sql += f"  {col} TEXT,\n"
        sql = sql[:-2] + "\n);\n\n"
        
        sql += f"INSERT INTO {table_name} ("
        sql += ", ".join(data.columns) + ") VALUES\n"
        
        for _, row in data.iterrows():
            values = []
            for col in data.columns:
                val = row[col]
                if pd.isna(val):
                    values.append("NULL")
                elif isinstance(val, str):
                    values.append(f"'{val.replace("'", "''")}'")
                elif isinstance(val, datetime):
                    values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                else:
                    values.append(str(val))
            sql += "(" + ", ".join(values) + "),\n"
        
        sql = sql[:-2] + ";"
        return sql.encode(config.encoding)

    async def _compress_data(self, data: bytes, compression: ExportCompression) -> bytes:
        if compression == ExportCompression.GZIP:
            import gzip
            return gzip.compress(data)
        elif compression == ExportCompression.ZIP:
            import zipfile
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('data', data)
            return buffer.getvalue()
        elif compression == ExportCompression.BZ2:
            import bz2
            return bz2.compress(data)
        elif compression == ExportCompression.XZ:
            import lzma
            return lzma.compress(data)
        elif compression in [ExportCompression.TAR, ExportCompression.TAR_GZ]:
            import tarfile
            buffer = io.BytesIO()
            mode = 'w:gz' if compression == ExportCompression.TAR_GZ else 'w'
            with tarfile.open(fileobj=buffer, mode=mode) as tf:
                tf.add(tempfile.NamedTemporaryFile(delete=False, suffix='.data'), arcname='data')
            return buffer.getvalue()
        else:
            return data

    async def _save_file(
        self,
        job_id: str,
        start: int,
        end: int,
        data: bytes,
        config: ExportConfig
    ) -> str:
        file_id = hashlib.md5(f"{job_id}_{start}_{end}".encode()).hexdigest()
        file_name = f"{file_id}.{config.format.value}"
        if config.compression != ExportCompression.NONE:
            file_name += f".{config.compression.value}"
        
        file_path = os.path.join(self._temp_dir, file_name)
        
        with open(file_path, 'wb') as f:
            f.write(data)
        
        file_info = ExportFile(
            id=file_id,
            job_id=job_id,
            name=file_name,
            path=file_path,
            size=len(data),
            format=config.format,
            compression=config.compression,
            row_count=end - start,
            checksum=hashlib.sha256(data).hexdigest()
        )
        
        self._files[file_id] = file_info
        return file_path

    async def _handle_large_file(self, job: ExportJob, batch: pd.DataFrame, config: ExportConfig) -> None:
        part_count = len(batch) // (config.batch_size // 2)
        for i in range(part_count):
            start = i * (config.batch_size // 2)
            end = min(start + config.batch_size // 2, len(batch))
            part = batch.iloc[start:end]
            
            formatted = await self._formatters[config.format](part, config)
            compressed = await self._compress_data(formatted, config.compression)
            await self._save_file(job.id, start, end, compressed, config)

    async def _destination_file(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        output_path = config.destination_config.get("output_path", "./exports")
        os.makedirs(output_path, exist_ok=True)
        
        file_paths = []
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                dest_path = os.path.join(output_path, file_info.name)
                shutil.copy2(file_info.path, dest_path)
                file_paths.append(dest_path)
        
        return {"files": file_paths, "path": output_path}

    async def _destination_s3(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        if not BOTO3_AVAILABLE:
            raise RuntimeError("boto3 not available")
        
        bucket = config.destination_config.get("bucket")
        prefix = config.destination_config.get("prefix", "exports/")
        
        if not bucket:
            raise ValueError("Bucket required for S3 export")
        
        s3_client = boto3.client('s3')
        
        uploaded_files = []
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                key = f"{prefix}{file_info.name}"
                s3_client.upload_file(file_info.path, bucket, key)
                uploaded_files.append(key)
        
        return {"bucket": bucket, "files": uploaded_files}

    async def _destination_gcs(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        try:
            from google.cloud import storage
        except ImportError:
            raise RuntimeError("google-cloud-storage not available")
        
        bucket_name = config.destination_config.get("bucket")
        prefix = config.destination_config.get("prefix", "exports/")
        
        if not bucket_name:
            raise ValueError("Bucket required for GCS export")
        
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        
        uploaded_files = []
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                blob = bucket.blob(f"{prefix}{file_info.name}")
                blob.upload_from_filename(file_info.path)
                uploaded_files.append(blob.name)
        
        return {"bucket": bucket_name, "files": uploaded_files}

    async def _destination_azure(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        try:
            from azure.storage.blob import BlobServiceClient
        except ImportError:
            raise RuntimeError("azure-storage-blob not available")
        
        connection_string = config.destination_config.get("connection_string")
        container = config.destination_config.get("container")
        prefix = config.destination_config.get("prefix", "exports/")
        
        if not connection_string or not container:
            raise ValueError("Connection string and container required")
        
        blob_service = BlobServiceClient.from_connection_string(connection_string)
        container_client = blob_service.get_container_client(container)
        
        uploaded_files = []
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                blob_client = container_client.get_blob_client(f"{prefix}{file_info.name}")
                with open(file_info.path, "rb") as f:
                    blob_client.upload_blob(f, overwrite=True)
                uploaded_files.append(blob_client.blob_name)
        
        return {"container": container, "files": uploaded_files}

    async def _destination_ftp(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        try:
            import aioftp
        except ImportError:
            raise RuntimeError("aioftp not available")
        
        host = config.destination_config.get("host")
        username = config.destination_config.get("username")
        password = config.destination_config.get("password")
        path = config.destination_config.get("path", "/")
        
        if not host:
            raise ValueError("Host required for FTP export")
        
        async with aioftp.Client() as client:
            await client.connect(host)
            if username and password:
                await client.login(username, password)
            
            uploaded_files = []
            for file_info in self._files.values():
                if file_info.job_id == job.id:
                    remote_path = f"{path}/{file_info.name}"
                    await client.upload(file_info.path, remote_path)
                    uploaded_files.append(remote_path)
        
        return {"host": host, "files": uploaded_files}

    async def _destination_sftp(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        try:
            import asyncssh
        except ImportError:
            raise RuntimeError("asyncssh not available")
        
        host = config.destination_config.get("host")
        username = config.destination_config.get("username")
        password = config.destination_config.get("password")
        path = config.destination_config.get("path", "/")
        port = config.destination_config.get("port", 22)
        
        if not host:
            raise ValueError("Host required for SFTP export")
        
        async with asyncssh.connect(host, username=username, password=password, port=port) as conn:
            async with conn.open_sftp() as sftp:
                uploaded_files = []
                for file_info in self._files.values():
                    if file_info.job_id == job.id:
                        remote_path = f"{path}/{file_info.name}"
                        await sftp.put(file_info.path, remote_path)
                        uploaded_files.append(remote_path)
        
        return {"host": host, "files": uploaded_files}

    async def _destination_webhook(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        if not AIOHTTP_AVAILABLE:
            raise RuntimeError("aiohttp not available")
        
        url = config.destination_config.get("url")
        method = config.destination_config.get("method", "POST")
        
        if not url:
            raise ValueError("URL required for webhook export")
        
        files = []
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                with open(file_info.path, 'rb') as f:
                    file_data = f.read()
                files.append(('file', (file_info.name, file_data)))
        
        async with aiohttp.ClientSession() as session:
            async with session.request(method=method, url=url, data=files) as response:
                if response.status >= 400:
                    raise ValueError(f"Webhook error: {response.status}")
        
        return {"status": "sent", "url": url}

    async def _destination_database(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        import sqlalchemy
        
        connection_string = config.destination_config.get("connection_string")
        table_name = config.destination_config.get("table", "exports")
        
        if not connection_string:
            raise ValueError("Connection string required")
        
        engine = sqlalchemy.create_engine(connection_string)
        
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                df = pd.read_csv(file_info.path)
                df.to_sql(table_name, engine, if_exists='append', index=False)
        
        return {"table": table_name, "rows": job.total_rows}

    async def _destination_email(self, job: ExportJob, config: ExportConfig) -> Dict[str, Any]:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email import encoders
        
        smtp_server = config.destination_config.get("smtp_server")
        smtp_port = config.destination_config.get("smtp_port", 587)
        username = config.destination_config.get("username")
        password = config.destination_config.get("password")
        sender = config.destination_config.get("sender")
        recipients = config.destination_config.get("recipients", [])
        
        if not smtp_server or not sender:
            raise ValueError("SMTP server and sender required")
        
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = f"Export {job.id} - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        body = f"Export completed successfully.\nRows: {job.exported_rows}\nFiles: {len(job.files)}"
        msg.attach(MIMEText(body, 'plain'))
        
        for file_info in self._files.values():
            if file_info.job_id == job.id:
                with open(file_info.path, "rb") as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename="{file_info.name}"'
                    )
                    msg.attach(part)
        
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            if username and password:
                server.login(username, password)
            server.send_message(msg)
        
        return {"sent_to": recipients, "count": len(recipients)}

    async def get_config(self, config_id: str) -> Optional[ExportConfig]:
        return self._configs.get(config_id)

    async def get_job(self, job_id: str) -> Optional[ExportJob]:
        return self._jobs.get(job_id)

    async def get_file(self, file_id: str) -> Optional[ExportFile]:
        return self._files.get(file_id)

    async def get_jobs_by_config(self, config_id: str) -> List[ExportJob]:
        return [j for j in self._jobs.values() if j.config_id == config_id]

    async def cleanup(self, job_id: str) -> None:
        for file_info in list(self._files.values()):
            if file_info.job_id == job_id:
                if os.path.exists(file_info.path):
                    os.remove(file_info.path)
                del self._files[file_info.id]

    async def _notify_observers(self, event: str, *args) -> None:
        for observer in self._observers:
            try:
                if asyncio.iscoroutinefunction(observer):
                    await observer(event, *args)
                else:
                    observer(event, *args)
            except Exception as e:
                logger.error(f"Observer error: {e}")

    def get_stats(self) -> Dict[str, Any]:
        return {
            "configs": len(self._configs),
            "jobs": len(self._jobs),
            "files": len(self._files),
            "formatters": len(self._formatters),
            "destinations": len(self._destinations),
            "observers": len(self._observers),
            "running": self._running
        }


__all__ = [
    "ExportFormat",
    "ExportCompression",
    "ExportDestination",
    "ExportStatus",
    "ExportConfig",
    "ExportJob",
    "ExportFile",
    "DataExporter"
]
