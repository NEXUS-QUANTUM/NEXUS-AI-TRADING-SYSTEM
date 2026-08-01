"""
NEXUS AI TRADING SYSTEM
Hedge Bot Reporting Module

Copyright © 2026 NEXUS QUANTUM LTD
CEO: Dr X... - Majority Shareholder

File: trading/bots/hedge_bot/hedge_bot_reporting.py
Description: Advanced reporting system for hedge bot with real-time analytics,
             performance metrics, risk reports, and multi-format export capabilities.
"""

import asyncio
import json
import logging
import math
import os
import tempfile
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple, Union, Set, Callable, Awaitable
from collections import defaultdict, deque
import io
import base64

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns
from scipy import stats

from shared.utilities.logger import get_logger
from shared.utilities.retry import retry_async, RetryConfig

logger = get_logger(__name__)


class ReportType(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    CUSTOM = "custom"
    REAL_TIME = "real_time"
    PERFORMANCE = "performance"
    RISK = "risk"
    PORTFOLIO = "portfolio"
    TRADING = "trading"
    EXECUTION = "execution"
    STRATEGY = "strategy"
    COMPLIANCE = "compliance"
    AUDIT = "audit"


class ReportFormat(str, Enum):
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"
    TEXT = "text"
    PNG = "png"
    SVG = "svg"
    DASHBOARD = "dashboard"


class TimePeriod(str, Enum):
    TODAY = "today"
    YESTERDAY = "yesterday"
    LAST_7_DAYS = "last_7_days"
    LAST_30_DAYS = "last_30_days"
    LAST_90_DAYS = "last_90_days"
    LAST_180_DAYS = "last_180_days"
    LAST_365_DAYS = "last_365_days"
    YEAR_TO_DATE = "ytd"
    QUARTER_TO_DATE = "qtd"
    MONTH_TO_DATE = "mtd"
    CUSTOM = "custom"


@dataclass
class ReportConfig:
    report_type: ReportType
    time_period: TimePeriod
    format: ReportFormat
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_details: bool = True
    include_recommendations: bool = True
    include_risk_metrics: bool = True
    include_performance_metrics: bool = True
    include_trade_history: bool = True
    include_portfolio_snapshot: bool = True
    max_trades: int = 1000
    chart_style: str = "dark_background"
    color_palette: str = "viridis"
    custom_attributes: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportData:
    title: str
    generated_at: datetime
    report_type: ReportType
    time_period: TimePeriod
    summary: Dict[str, Any]
    performance: Dict[str, Any]
    risk: Dict[str, Any]
    portfolio: Dict[str, Any]
    trading: Dict[str, Any]
    strategies: Dict[str, Any]
    charts: Dict[str, str] = field(default_factory=dict)
    tables: Dict[str, pd.DataFrame] = field(default_factory=dict)
    recommendations: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    custom_sections: Dict[str, Any] = field(default_factory=dict)


@dataclass
class PerformanceSnapshot:
    total_pnl: float
    total_pnl_percent: float
    daily_pnl: List[float]
    weekly_pnl: List[float]
    monthly_pnl: List[float]
    win_rate: float
    losing_rate: float
    avg_win: float
    avg_loss: float
    profit_factor: float
    max_drawdown: float
    current_drawdown: float
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    total_trades: int
    winning_trades: int
    losing_trades: int
    breakeven_trades: int
    avg_trade_duration: float
    timestamp: datetime = field(default_factory=datetime.now)


@dataclass
class RiskSnapshot:
    var_95: float
    var_99: float
    cvar_95: float
    cvar_99: float
    expected_shortfall: float
    max_drawdown: float
    current_drawdown: float
    volatility: float
    beta: float
    alpha: float
    correlation_matrix: Dict[str, Dict[str, float]]
    concentration_ratio: float
    sector_exposures: Dict[str, float]
    asset_class_exposures: Dict[str, float]
    risk_score: float
    risk_level: str
    stress_test_results: Dict[str, float]
    timestamp: datetime = field(default_factory=datetime.now)


@dataclass
class PortfolioSnapshot:
    total_value: float
    total_exposure: float
    total_risk: float
    cash_balance: float
    positions: List[Dict[str, Any]]
    sector_allocation: Dict[str, float]
    asset_class_allocation: Dict[str, float]
    top_performers: List[Dict[str, Any]]
    worst_performers: List[Dict[str, Any]]
    timestamp: datetime = field(default_factory=datetime.now)


@dataclass
class TradingSnapshot:
    total_trades: int
    open_positions: int
    closed_positions: int
    pending_orders: int
    average_order_size: float
    average_slippage: float
    average_latency: float
    fill_rate: float
    order_book_depth: Dict[str, Any]
    recent_trades: List[Dict[str, Any]]
    timestamp: datetime = field(default_factory=datetime.now)


@dataclass
class StrategySnapshot:
    strategy_name: str
    total_pnl: float
    win_rate: float
    profit_factor: float
    total_trades: int
    avg_return: float
    max_drawdown: float
    sharpe_ratio: float
    sortino_ratio: float
    calmar_ratio: float
    active_positions: int
    timestamp: datetime = field(default_factory=datetime.now)


class ReportingService:
    """
    Advanced reporting service for hedge bot.
    
    Features:
    - Multi-format report generation
    - Real-time analytics
    - Performance metrics
    - Risk metrics
    - Portfolio analytics
    - Trading analytics
    - Strategy analytics
    - Interactive dashboards
    - Chart generation
    - Export capabilities
    - Scheduled reports
    - Email delivery
    - Webhook integration
    - Custom report templates
    - Data aggregation
    - Time series analysis
    - Benchmark comparison
    - Anomaly detection
    """
    
    def __init__(
        self,
        config: Dict[str, Any],
        risk_manager: Optional[Any] = None,
        portfolio_manager: Optional[Any] = None,
        trading_engine: Optional[Any] = None,
    ):
        self.config = config
        self.risk_manager = risk_manager
        self.portfolio_manager = portfolio_manager
        self.trading_engine = trading_engine
        
        self._report_config = ReportConfig(**config.get("reporting", {}))
        
        self._reports: Dict[str, ReportData] = {}
        self._snapshots: Dict[str, Any] = {}
        self._report_history: List[Dict[str, Any]] = []
        
        self._data_cache: Dict[str, pd.DataFrame] = {}
        self._chart_cache: Dict[str, str] = {}
        
        self._is_running = False
        self._scheduler_task: Optional[asyncio.Task] = None
        
        # Initialize matplotlib style
        plt.style.use(self._report_config.chart_style)
        sns.set_palette(self._report_config.color_palette)
        sns.set_style("darkgrid")
        
        # Set defaults for charts
        plt.rcParams['figure.figsize'] = (12, 6)
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.labelsize'] = 12
        plt.rcParams['axes.titlesize'] = 14
        plt.rcParams['legend.fontsize'] = 10
        
        logger.info("ReportingService initialized")
    
    # ========================================================================
    # REPORT GENERATION
    # ========================================================================
    
    async def generate_report(
        self,
        report_config: Optional[ReportConfig] = None,
        custom_data: Optional[Dict[str, Any]] = None,
    ) -> ReportData:
        """
        Generate a comprehensive report.
        
        Args:
            report_config: Report configuration
            custom_data: Custom data to include
            
        Returns:
            ReportData object
        """
        config = report_config or self._report_config
        
        logger.info(f"Generating {config.report_type.value} report")
        
        # Gather data
        summary = await self._gather_summary_data(config)
        performance = await self._gather_performance_data(config)
        risk = await self._gather_risk_data(config)
        portfolio = await self._gather_portfolio_data(config)
        trading = await self._gather_trading_data(config)
        strategies = await self._gather_strategy_data(config)
        
        # Generate charts
        charts = {}
        if config.include_charts:
            charts = await self._generate_charts(
                performance,
                risk,
                portfolio,
                trading,
                config,
            )
        
        # Generate tables
        tables = {}
        if config.include_tables:
            tables = await self._generate_tables(
                performance,
                risk,
                portfolio,
                trading,
                config,
            )
        
        # Generate recommendations
        recommendations = []
        if config.include_recommendations:
            recommendations = await self._generate_recommendations(
                performance,
                risk,
                portfolio,
                trading,
            )
        
        # Create report
        report = ReportData(
            title=f"{config.report_type.value.title()} Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            generated_at=datetime.now(),
            report_type=config.report_type,
            time_period=config.time_period,
            summary=summary,
            performance=performance,
            risk=risk,
            portfolio=portfolio,
            trading=trading,
            strategies=strategies,
            charts=charts,
            tables=tables,
            recommendations=recommendations,
            metadata={
                "config": asdict(config),
                "version": "1.0",
                "system": "NexusTradingIA",
            },
            custom_sections=custom_data or {},
        )
        
        # Store report
        report_id = f"{config.report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self._reports[report_id] = report
        self._report_history.append({
            "id": report_id,
            "type": config.report_type.value,
            "generated_at": report.generated_at,
            "format": config.format.value,
        })
        
        logger.info(f"Report generated: {report_id}")
        
        return report
    
    async def export_report(
        self,
        report: ReportData,
        format: ReportFormat,
        output_path: Optional[str] = None,
    ) -> Union[str, bytes]:
        """
        Export a report in the specified format.
        
        Args:
            report: ReportData object
            format: Export format
            output_path: Output file path
            
        Returns:
            Exported data (string or bytes)
        """
        logger.info(f"Exporting report to {format.value}")
        
        if format == ReportFormat.PDF:
            return await self._export_pdf(report, output_path)
        elif format == ReportFormat.HTML:
            return await self._export_html(report, output_path)
        elif format == ReportFormat.JSON:
            return await self._export_json(report, output_path)
        elif format == ReportFormat.CSV:
            return await self._export_csv(report, output_path)
        elif format == ReportFormat.EXCEL:
            return await self._export_excel(report, output_path)
        elif format == ReportFormat.MARKDOWN:
            return await self._export_markdown(report, output_path)
        elif format == ReportFormat.TEXT:
            return await self._export_text(report, output_path)
        elif format == ReportFormat.DASHBOARD:
            return await self._export_dashboard(report, output_path)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    # ========================================================================
    # DATA GATHERING
    # ========================================================================
    
    async def _gather_summary_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather summary data for the report."""
        summary = {
            "generated_at": datetime.now().isoformat(),
            "time_period": config.time_period.value,
            "report_type": config.report_type.value,
        }
        
        # Get portfolio summary
        if self.portfolio_manager:
            try:
                portfolio = await self.portfolio_manager.get_summary()
                summary["portfolio"] = portfolio
            except Exception as e:
                logger.error(f"Error gathering portfolio summary: {e}")
        
        # Get risk summary
        if self.risk_manager:
            try:
                risk = await self.risk_manager.get_portfolio_summary()
                summary["risk"] = risk
            except Exception as e:
                logger.error(f"Error gathering risk summary: {e}")
        
        # Get trading summary
        if self.trading_engine:
            try:
                trading = await self.trading_engine.get_summary()
                summary["trading"] = trading
            except Exception as e:
                logger.error(f"Error gathering trading summary: {e}")
        
        return summary
    
    async def _gather_performance_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather performance metrics."""
        performance = {}
        
        try:
            # Get performance metrics
            if self.portfolio_manager:
                performance["portfolio"] = await self.portfolio_manager.get_performance_metrics()
            
            # Get trade statistics
            if self.trading_engine:
                performance["trading"] = await self.trading_engine.get_performance_metrics()
            
            # Calculate derived metrics
            if self.portfolio_manager and self.risk_manager:
                performance["risk_adjusted"] = await self._calculate_risk_adjusted_metrics()
            
            # Get benchmark comparison
            performance["benchmark"] = await self._get_benchmark_comparison()
            
        except Exception as e:
            logger.error(f"Error gathering performance data: {e}")
        
        return performance
    
    async def _gather_risk_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather risk metrics."""
        risk_data = {}
        
        try:
            if self.risk_manager:
                # Get current risk snapshot
                risk_snapshot = await self.risk_manager.calculate_portfolio_risk()
                if risk_snapshot:
                    risk_data = asdict(risk_snapshot)
            
            # Get historical risk metrics
            risk_data["historical"] = await self._get_historical_risk_metrics()
            
            # Get stress test results
            if self.risk_manager:
                risk_data["stress_tests"] = await self.risk_manager._run_stress_tests([])
            
        except Exception as e:
            logger.error(f"Error gathering risk data: {e}")
        
        return risk_data
    
    async def _gather_portfolio_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather portfolio data."""
        portfolio_data = {}
        
        try:
            if self.portfolio_manager:
                # Get current portfolio
                portfolio = await self.portfolio_manager.get_portfolio()
                if portfolio:
                    portfolio_data["current"] = asdict(portfolio)
                
                # Get allocation
                allocation = await self.portfolio_manager.get_allocation()
                if allocation:
                    portfolio_data["allocation"] = allocation
                
                # Get historical portfolio values
                portfolio_data["history"] = await self._get_portfolio_history()
            
        except Exception as e:
            logger.error(f"Error gathering portfolio data: {e}")
        
        return portfolio_data
    
    async def _gather_trading_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather trading data."""
        trading_data = {}
        
        try:
            if self.trading_engine:
                # Get trade history
                trades = await self.trading_engine.get_trade_history(
                    limit=config.max_trades
                )
                if trades:
                    trading_data["trades"] = trades
                
                # Get open positions
                positions = await self.trading_engine.get_open_positions()
                if positions:
                    trading_data["open_positions"] = positions
                
                # Get execution metrics
                execution = await self.trading_engine.get_execution_metrics()
                if execution:
                    trading_data["execution"] = execution
            
            # Get trade statistics
            trading_data["statistics"] = await self._calculate_trade_statistics()
            
        except Exception as e:
            logger.error(f"Error gathering trading data: {e}")
        
        return trading_data
    
    async def _gather_strategy_data(self, config: ReportConfig) -> Dict[str, Any]:
        """Gather strategy data."""
        strategy_data = {}
        
        try:
            if self.trading_engine:
                strategies = await self.trading_engine.get_strategy_performance()
                if strategies:
                    strategy_data["strategies"] = strategies
            
            # Get strategy rankings
            strategy_data["rankings"] = await self._get_strategy_rankings()
            
        except Exception as e:
            logger.error(f"Error gathering strategy data: {e}")
        
        return strategy_data
    
    # ========================================================================
    # CHART GENERATION
    # ========================================================================
    
    async def _generate_charts(
        self,
        performance: Dict[str, Any],
        risk: Dict[str, Any],
        portfolio: Dict[str, Any],
        trading: Dict[str, Any],
        config: ReportConfig,
    ) -> Dict[str, str]:
        """Generate charts for the report."""
        charts = {}
        
        try:
            # PNL Chart
            charts["pnl"] = await self._generate_pnl_chart(performance, config)
            
            # Drawdown Chart
            charts["drawdown"] = await self._generate_drawdown_chart(risk, config)
            
            # Portfolio Allocation Chart
            charts["allocation"] = await self._generate_allocation_chart(portfolio, config)
            
            # Performance Metrics Chart
            charts["performance_metrics"] = await self._generate_metrics_chart(performance, config)
            
            # Risk Metrics Chart
            charts["risk_metrics"] = await self._generate_risk_chart(risk, config)
            
            # Trade Distribution Chart
            charts["trade_distribution"] = await self._generate_trade_distribution_chart(trading, config)
            
            # Correlation Heatmap
            charts["correlation"] = await self._generate_correlation_heatmap(risk, config)
            
            # Strategy Performance Chart
            charts["strategy_performance"] = await self._generate_strategy_chart(performance, config)
            
            # Equity Curve
            charts["equity_curve"] = await self._generate_equity_curve_chart(performance, config)
            
            # Monthly Returns Heatmap
            charts["monthly_returns"] = await self._generate_monthly_returns_chart(performance, config)
            
        except Exception as e:
            logger.error(f"Error generating charts: {e}")
        
        return charts
    
    async def _generate_pnl_chart(
        self,
        performance: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate PNL chart."""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # Get PNL data
        pnl_data = performance.get("portfolio", {}).get("daily_pnl", [])
        if not pnl_data:
            return ""
        
        dates = pd.date_range(end=datetime.now(), periods=len(pnl_data))
        
        # Plot PNL
        ax.bar(dates, pnl_data, color=['green' if x >= 0 else 'red' for x in pnl_data])
        
        # Formatting
        ax.set_title('Daily PNL', fontsize=14, fontweight='bold')
        ax.set_xlabel('Date')
        ax.set_ylabel('PNL ($)')
        ax.axhline(y=0, color='white', linestyle='-', alpha=0.3)
        
        # Add rolling average
        if len(pnl_data) > 20:
            rolling_avg = pd.Series(pnl_data).rolling(20).mean()
            ax.plot(dates, rolling_avg, color='yellow', linewidth=2, label='20-day MA')
            ax.legend()
        
        ax.grid(True, alpha=0.3)
        
        return self._figure_to_base64(fig)
    
    async def _generate_drawdown_chart(
        self,
        risk: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate drawdown chart."""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # Get drawdown data
        drawdown_data = risk.get("historical", {}).get("drawdowns", [])
        if not drawdown_data:
            return ""
        
        dates = pd.date_range(end=datetime.now(), periods=len(drawdown_data))
        
        # Plot drawdown
        ax.fill_between(dates, 0, drawdown_data, color='red', alpha=0.5)
        ax.plot(dates, drawdown_data, color='darkred', linewidth=1)
        
        # Formatting
        ax.set_title('Portfolio Drawdown', fontsize=14, fontweight='bold')
        ax.set_xlabel('Date')
        ax.set_ylabel('Drawdown (%)')
        ax.axhline(y=-0.1, color='yellow', linestyle='--', alpha=0.5, label='10% Warning')
        ax.axhline(y=-0.2, color='orange', linestyle='--', alpha=0.5, label='20% Critical')
        
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Format y-axis as percentage
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y:.0%}'))
        
        return self._figure_to_base64(fig)
    
    async def _generate_allocation_chart(
        self,
        portfolio: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate portfolio allocation chart."""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        # Get allocation data
        allocation = portfolio.get("allocation", {})
        sector_allocation = allocation.get("sectors", {})
        asset_allocation = allocation.get("asset_classes", {})
        
        # Sector allocation pie chart
        if sector_allocation:
            labels = list(sector_allocation.keys())
            sizes = list(sector_allocation.values())
            ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax1.set_title('Sector Allocation', fontweight='bold')
        
        # Asset allocation pie chart
        if asset_allocation:
            labels = list(asset_allocation.keys())
            sizes = list(asset_allocation.values())
            ax2.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax2.set_title('Asset Class Allocation', fontweight='bold')
        
        plt.tight_layout()
        return self._figure_to_base64(fig)
    
    async def _generate_metrics_chart(
        self,
        performance: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate performance metrics chart."""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Get metrics
        metrics = performance.get("risk_adjusted", {})
        
        if not metrics:
            return ""
        
        # Select key metrics
        metric_names = ['Sharpe Ratio', 'Sortino Ratio', 'Calmar Ratio', 'Win Rate', 'Profit Factor']
        metric_values = [
            metrics.get('sharpe_ratio', 0),
            metrics.get('sortino_ratio', 0),
            metrics.get('calmar_ratio', 0),
            metrics.get('win_rate', 0) * 100,
            metrics.get('profit_factor', 0),
        ]
        
        # Create bar chart
        bars = ax.bar(metric_names, metric_values)
        
        # Color bars based on values
        colors = ['green' if v > 0 else 'red' for v in metric_values]
        for bar, color in zip(bars, colors):
            bar.set_color(color)
        
        ax.set_title('Key Performance Metrics', fontsize=14, fontweight='bold')
        ax.set_ylabel('Value')
        ax.grid(True, alpha=0.3)
        
        # Add value labels on bars
        for bar, value in zip(bars, metric_values):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                    f'{value:.2f}', ha='center', va='bottom')
        
        return self._figure_to_base64(fig)
    
    async def _generate_risk_chart(
        self,
        risk: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate risk metrics chart."""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Get risk metrics
        risk_metrics = risk.get("risk", {})
        
        if not risk_metrics:
            return ""
        
        # Select key metrics
        metric_names = ['VaR 95%', 'CVaR 95%', 'VaR 99%', 'CVaR 99%', 'Volatility']
        metric_values = [
            abs(risk_metrics.get('var_95', 0)) * 100,
            abs(risk_metrics.get('cvar_95', 0)) * 100,
            abs(risk_metrics.get('var_99', 0)) * 100,
            abs(risk_metrics.get('cvar_99', 0)) * 100,
            risk_metrics.get('volatility', 0) * 100,
        ]
        
        # Create bar chart
        colors = ['red' if v > 5 else 'orange' if v > 2 else 'green' for v in metric_values]
        ax.bar(metric_names, metric_values, color=colors)
        
        ax.set_title('Risk Metrics', fontsize=14, fontweight='bold')
        ax.set_ylabel('Value (%)')
        ax.grid(True, alpha=0.3)
        
        # Add value labels
        for i, v in enumerate(metric_values):
            ax.text(i, v + 0.5, f'{v:.1f}%', ha='center', va='bottom')
        
        return self._figure_to_base64(fig)
    
    async def _generate_trade_distribution_chart(
        self,
        trading: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate trade distribution chart."""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        # Get trade data
        trades = trading.get("statistics", {})
        
        if not trades:
            return ""
        
        # Win/Loss distribution
        win_rate = trades.get('win_rate', 0)
        loss_rate = trades.get('loss_rate', 0)
        breakeven_rate = trades.get('breakeven_rate', 0)
        
        labels = ['Wins', 'Losses', 'Breakeven']
        sizes = [win_rate * 100, loss_rate * 100, breakeven_rate * 100]
        colors = ['green', 'red', 'gray']
        
        ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax1.set_title('Trade Outcome Distribution', fontweight='bold')
        
        # PNL distribution histogram
        pnl_values = trading.get('trades', {}).get('pnl', [])
        if pnl_values:
            ax2.hist(pnl_values, bins=30, color='blue', alpha=0.7, edgecolor='black')
            ax2.axvline(x=0, color='red', linestyle='--', alpha=0.5)
            ax2.set_title('PNL Distribution')
            ax2.set_xlabel('PNL ($)')
            ax2.set_ylabel('Frequency')
            ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        return self._figure_to_base64(fig)
    
    async def _generate_correlation_heatmap(
        self,
        risk: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate correlation heatmap."""
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Get correlation matrix
        corr_matrix = risk.get("correlation_matrix", {})
        
        if not corr_matrix:
            return ""
        
        # Convert to DataFrame
        symbols = list(corr_matrix.keys())
        matrix = np.zeros((len(symbols), len(symbols)))
        
        for i, s1 in enumerate(symbols):
            for j, s2 in enumerate(symbols):
                matrix[i, j] = corr_matrix.get(s1, {}).get(s2, 0)
        
        df_corr = pd.DataFrame(matrix, index=symbols, columns=symbols)
        
        # Create heatmap
        sns.heatmap(df_corr, annot=True, cmap='RdBu_r', center=0, ax=ax,
                    fmt='.2f', vmin=-1, vmax=1)
        ax.set_title('Asset Correlation Matrix', fontsize=14, fontweight='bold')
        
        plt.tight_layout()
        return self._figure_to_base64(fig)
    
    async def _generate_strategy_chart(
        self,
        performance: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate strategy performance chart."""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # Get strategy data
        strategies = performance.get("strategies", {})
        
        if not strategies:
            return ""
        
        # Extract data
        names = list(strategies.keys())
        pnl = [s.get('total_pnl', 0) for s in strategies.values()]
        win_rates = [s.get('win_rate', 0) * 100 for s in strategies.values()]
        
        x = np.arange(len(names))
        width = 0.35
        
        # Create grouped bar chart
        bars1 = ax.bar(x - width/2, pnl, width, label='Total PNL', color='green')
        ax2 = ax.twinx()
        bars2 = ax2.bar(x + width/2, win_rates, width, label='Win Rate', color='blue')
        
        ax.set_xlabel('Strategy')
        ax.set_ylabel('PNL ($)')
        ax2.set_ylabel('Win Rate (%)')
        ax.set_title('Strategy Performance', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=45, ha='right')
        
        # Add legend
        lines1, labels1 = ax.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
        
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        return self._figure_to_base64(fig)
    
    async def _generate_equity_curve_chart(
        self,
        performance: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate equity curve chart."""
        fig, ax = plt.subplots(figsize=(12, 6))
        
        # Get equity data
        equity = performance.get("portfolio", {}).get("equity_curve", [])
        
        if not equity:
            return ""
        
        dates = pd.date_range(end=datetime.now(), periods=len(equity))
        
        # Plot equity curve
        ax.plot(dates, equity, color='blue', linewidth=2)
        ax.fill_between(dates, min(equity), equity, alpha=0.3)
        
        # Add moving average
        if len(equity) > 20:
            ma = pd.Series(equity).rolling(20).mean()
            ax.plot(dates, ma, color='orange', linewidth=1.5, label='20-day MA')
            ax.legend()
        
        ax.set_title('Portfolio Equity Curve', fontsize=14, fontweight='bold')
        ax.set_xlabel('Date')
        ax.set_ylabel('Portfolio Value ($)')
        ax.grid(True, alpha=0.3)
        
        # Format y-axis
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'${y:,.0f}'))
        
        return self._figure_to_base64(fig)
    
    async def _generate_monthly_returns_chart(
        self,
        performance: Dict[str, Any],
        config: ReportConfig,
    ) -> str:
        """Generate monthly returns heatmap."""
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Get monthly returns data
        monthly_returns = performance.get("portfolio", {}).get("monthly_returns", [])
        
        if not monthly_returns:
            return ""
        
        # Create matrix
        years = []
        months = []
        returns = []
        
        for entry in monthly_returns:
            years.append(entry['year'])
            months.append(entry['month'])
            returns.append(entry['return'])
        
        # Create DataFrame
        df_returns = pd.DataFrame(returns, index=pd.MultiIndex.from_arrays([years, months]))
        df_returns = df_returns.unstack()
        
        # Create heatmap
        sns.heatmap(df_returns, annot=True, cmap='RdYlGn', center=0, ax=ax,
                    fmt='.2%', cbar_kws={'label': 'Monthly Return'})
        
        ax.set_title('Monthly Returns Heatmap', fontsize=14, fontweight='bold')
        ax.set_xlabel('Month')
        ax.set_ylabel('Year')
        
        plt.tight_layout()
        return self._figure_to_base64(fig)
    
    def _figure_to_base64(self, fig: Figure) -> str:
        """Convert matplotlib figure to base64 string."""
        canvas = FigureCanvasAgg(fig)
        buf = io.BytesIO()
        canvas.print_png(buf)
        plt.close(fig)
        return base64.b64encode(buf.getvalue()).decode('utf-8')
    
    # ========================================================================
    # TABLE GENERATION
    # ========================================================================
    
    async def _generate_tables(
        self,
        performance: Dict[str, Any],
        risk: Dict[str, Any],
        portfolio: Dict[str, Any],
        trading: Dict[str, Any],
        config: ReportConfig,
    ) -> Dict[str, pd.DataFrame]:
        """Generate data tables for the report."""
        tables = {}
        
        try:
            # Performance table
            tables["performance"] = await self._generate_performance_table(performance)
            
            # Risk table
            tables["risk"] = await self._generate_risk_table(risk)
            
            # Portfolio table
            tables["portfolio"] = await self._generate_portfolio_table(portfolio)
            
            # Trading table
            tables["trading"] = await self._generate_trading_table(trading)
            
            # Strategy table
            tables["strategy"] = await self._generate_strategy_table(performance)
            
        except Exception as e:
            logger.error(f"Error generating tables: {e}")
        
        return tables
    
    async def _generate_performance_table(self, performance: Dict[str, Any]) -> pd.DataFrame:
        """Generate performance metrics table."""
        data = {
            "Metric": [],
            "Value": [],
        }
        
        metrics = performance.get("risk_adjusted", {})
        
        metric_mapping = {
            "Total PNL": metrics.get("total_pnl", 0),
            "Total PNL %": metrics.get("total_pnl_percent", 0),
            "Win Rate": metrics.get("win_rate", 0),
            "Profit Factor": metrics.get("profit_factor", 0),
            "Sharpe Ratio": metrics.get("sharpe_ratio", 0),
            "Sortino Ratio": metrics.get("sortino_ratio", 0),
            "Calmar Ratio": metrics.get("calmar_ratio", 0),
            "Max Drawdown": metrics.get("max_drawdown", 0),
            "Avg Win": metrics.get("avg_win", 0),
            "Avg Loss": metrics.get("avg_loss", 0),
            "Total Trades": metrics.get("total_trades", 0),
        }
        
        for name, value in metric_mapping.items():
            data["Metric"].append(name)
            if isinstance(value, float):
                if "Ratio" in name or "Rate" in name or "%" in name:
                    data["Value"].append(f"{value:.2%}" if abs(value) < 1 else f"{value:.2f}")
                else:
                    data["Value"].append(f"{value:,.2f}")
            else:
                data["Value"].append(str(value))
        
        return pd.DataFrame(data)
    
    async def _generate_risk_table(self, risk: Dict[str, Any]) -> pd.DataFrame:
        """Generate risk metrics table."""
        data = {
            "Metric": [],
            "Value": [],
        }
        
        risk_metrics = risk.get("risk", {})
        
        metric_mapping = {
            "VaR 95%": risk_metrics.get("var_95", 0),
            "CVaR 95%": risk_metrics.get("cvar_95", 0),
            "VaR 99%": risk_metrics.get("var_99", 0),
            "CVaR 99%": risk_metrics.get("cvar_99", 0),
            "Volatility": risk_metrics.get("volatility", 0),
            "Beta": risk_metrics.get("beta", 0),
            "Alpha": risk_metrics.get("alpha", 0),
            "Concentration": risk_metrics.get("concentration_ratio", 0),
            "Risk Score": risk_metrics.get("risk_score", 0),
            "Risk Level": risk_metrics.get("risk_level", "Unknown"),
        }
        
        for name, value in metric_mapping.items():
            data["Metric"].append(name)
            if isinstance(value, float):
                if "Ratio" in name or "Risk Score" in name:
                    data["Value"].append(f"{value:.2f}")
                elif "Risk Level" not in name:
                    data["Value"].append(f"{value:.2%}")
            else:
                data["Value"].append(str(value))
        
        return pd.DataFrame(data)
    
    async def _generate_portfolio_table(self, portfolio: Dict[str, Any]) -> pd.DataFrame:
        """Generate portfolio table."""
        data = {
            "Metric": [],
            "Value": [],
        }
        
        portfolio_data = portfolio.get("current", {})
        
        if portfolio_data:
            metric_mapping = {
                "Total Value": portfolio_data.get("total_value", 0),
                "Total Exposure": portfolio_data.get("total_exposure", 0),
                "Total Risk": portfolio_data.get("total_risk", 0),
                "Cash Balance": portfolio_data.get("cash_balance", 0),
                "Positions": portfolio_data.get("position_count", 0),
                "Leverage": portfolio_data.get("leverage", 0),
            }
            
            for name, value in metric_mapping.items():
                data["Metric"].append(name)
                if isinstance(value, float):
                    if "Leverage" in name:
                        data["Value"].append(f"{value:.2f}x")
                    else:
                        data["Value"].append(f"${value:,.2f}")
                else:
                    data["Value"].append(str(value))
        
        return pd.DataFrame(data)
    
    async def _generate_trading_table(self, trading: Dict[str, Any]) -> pd.DataFrame:
        """Generate trading table."""
        data = {
            "Metric": [],
            "Value": [],
        }
        
        stats = trading.get("statistics", {})
        
        metric_mapping = {
            "Total Trades": stats.get("total_trades", 0),
            "Open Positions": stats.get("open_positions", 0),
            "Closed Positions": stats.get("closed_positions", 0),
            "Winning Trades": stats.get("winning_trades", 0),
            "Losing Trades": stats.get("losing_trades", 0),
            "Average Order Size": stats.get("avg_order_size", 0),
            "Average Slippage": stats.get("avg_slippage", 0),
            "Average Latency": stats.get("avg_latency", 0),
            "Fill Rate": stats.get("fill_rate", 0),
        }
        
        for name, value in metric_mapping.items():
            data["Metric"].append(name)
            if isinstance(value, float):
                if "Rate" in name or "Slippage" in name:
                    data["Value"].append(f"{value:.2%}")
                elif "Latency" in name:
                    data["Value"].append(f"{value:.2f}ms")
                else:
                    data["Value"].append(f"{value:,.2f}")
            else:
                data["Value"].append(str(value))
        
        return pd.DataFrame(data)
    
    async def _generate_strategy_table(self, performance: Dict[str, Any]) -> pd.DataFrame:
        """Generate strategy performance table."""
        strategies = performance.get("strategies", {})
        
        if not strategies:
            return pd.DataFrame()
        
        data = {
            "Strategy": [],
            "PNL": [],
            "Win Rate": [],
            "Profit Factor": [],
            "Trades": [],
            "Sharpe": [],
            "Drawdown": [],
        }
        
        for name, strategy in strategies.items():
            data["Strategy"].append(name)
            data["PNL"].append(f"${strategy.get('total_pnl', 0):,.2f}")
            data["Win Rate"].append(f"{strategy.get('win_rate', 0):.1%}")
            data["Profit Factor"].append(f"{strategy.get('profit_factor', 0):.2f}")
            data["Trades"].append(str(strategy.get('total_trades', 0)))
            data["Sharpe"].append(f"{strategy.get('sharpe_ratio', 0):.2f}")
            data["Drawdown"].append(f"{strategy.get('max_drawdown', 0):.1%}")
        
        return pd.DataFrame(data)
    
    # ========================================================================
    # EXPORT FUNCTIONS
    # ========================================================================
    
    async def _export_pdf(self, report: ReportData, output_path: Optional[str] = None) -> bytes:
        """Export report as PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            import io
            from PIL import Image as PILImage
            
            if output_path:
                doc = SimpleDocTemplate(output_path, pagesize=letter)
            else:
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
            
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0066CC'),
                spaceAfter=30,
            )
            story.append(Paragraph(report.title, title_style))
            
            # Metadata
            metadata_style = styles['Normal']
            story.append(Paragraph(f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}", metadata_style))
            story.append(Paragraph(f"Report Type: {report.report_type.value}", metadata_style))
            story.append(Paragraph(f"Time Period: {report.time_period.value}", metadata_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Summary Section
            story.append(Paragraph("Summary", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            # Summary table
            summary_data = [['Metric', 'Value']]
            for key, value in report.summary.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        summary_data.append([f"{key}_{sub_key}", str(sub_value)])
                else:
                    summary_data.append([key, str(value)])
            
            summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(summary_table)
            story.append(PageBreak())
            
            # Performance Section
            story.append(Paragraph("Performance Metrics", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            # Performance table
            if 'performance' in report.tables:
                perf_df = report.tables['performance']
                perf_data = [perf_df.columns.tolist()] + perf_df.values.tolist()
                perf_table = Table(perf_data, colWidths=[2*inch, 3*inch])
                perf_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(perf_table)
            
            story.append(Spacer(1, 0.2*inch))
            
            # Charts
            if report.charts:
                story.append(Paragraph("Charts", styles['Heading2']))
                story.append(Spacer(1, 0.1*inch))
                
                for chart_name, chart_data in report.charts.items():
                    # Decode base64 image
                    image_data = base64.b64decode(chart_data)
                    img_buffer = io.BytesIO(image_data)
                    img = PILImage.open(img_buffer)
                    
                    # Convert to reportlab Image
                    img_draw = Drawing(400, 200)
                    img_draw.add(Image(0, 0, 400, 200, img_buffer))
                    story.append(img_draw)
                    story.append(Spacer(1, 0.1*inch))
            
            # Risk Section
            story.append(PageBreak())
            story.append(Paragraph("Risk Metrics", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            if 'risk' in report.tables:
                risk_df = report.tables['risk']
                risk_data = [risk_df.columns.tolist()] + risk_df.values.tolist()
                risk_table = Table(risk_data, colWidths=[2*inch, 3*inch])
                risk_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(risk_table)
            
            # Recommendations
            if report.recommendations:
                story.append(PageBreak())
                story.append(Paragraph("Recommendations", styles['Heading2']))
                story.append(Spacer(1, 0.1*inch))
                
                for i, rec in enumerate(report.recommendations, 1):
                    story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
            
            # Build PDF
            doc.build(story)
            
            if output_path:
                return b""
            else:
                return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}")
            return b""
    
    async def _export_html(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as HTML."""
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title}</title>
            <style>
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
                    margin: 40px;
                    background: #f5f7fa;
                    color: #2c3e50;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: white;
                    padding: 30px;
                    border-radius: 12px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                h1 {{
                    color: #2c3e50;
                    border-bottom: 3px solid #3498db;
                    padding-bottom: 10px;
                }}
                h2 {{
                    color: #34495e;
                    margin-top: 30px;
                    border-bottom: 2px solid #ecf0f1;
                    padding-bottom: 8px;
                }}
                .metadata {{
                    background: #ecf0f1;
                    padding: 15px;
                    border-radius: 8px;
                    margin-bottom: 20px;
                }}
                .section {{
                    margin-bottom: 30px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }}
                th {{
                    background: #3498db;
                    color: white;
                    padding: 12px;
                    text-align: left;
                }}
                td {{
                    padding: 10px;
                    border-bottom: 1px solid #ecf0f1;
                }}
                tr:hover {{
                    background: #f8f9fa;
                }}
                .chart-container {{
                    margin: 20px 0;
                    text-align: center;
                }}
                .chart-container img {{
                    max-width: 100%;
                    border: 1px solid #ecf0f1;
                    border-radius: 8px;
                }}
                .recommendation {{
                    background: #f8f9fa;
                    padding: 15px;
                    margin: 10px 0;
                    border-left: 4px solid #3498db;
                    border-radius: 4px;
                }}
                .badge {{
                    display: inline-block;
                    padding: 4px 12px;
                    border-radius: 20px;
                    font-size: 12px;
                    font-weight: bold;
                }}
                .badge-green {{
                    background: #27ae60;
                    color: white;
                }}
                .badge-red {{
                    background: #e74c3c;
                    color: white;
                }}
                .badge-yellow {{
                    background: #f39c12;
                    color: white;
                }}
                @media print {{
                    body {{
                        background: white;
                        margin: 20px;
                    }}
                    .container {{
                        box-shadow: none;
                        padding: 20px;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>{title}</h1>
                <div class="metadata">
                    <strong>Generated:</strong> {generated_at}<br>
                    <strong>Report Type:</strong> {report_type}<br>
                    <strong>Time Period:</strong> {time_period}
                </div>
                
                <div class="section">
                    <h2>Summary</h2>
                    {summary_table}
                </div>
                
                <div class="section">
                    <h2>Performance Metrics</h2>
                    {performance_table}
                </div>
                
                <div class="section">
                    <h2>Charts</h2>
                    {charts}
                </div>
                
                <div class="section">
                    <h2>Risk Metrics</h2>
                    {risk_table}
                </div>
                
                <div class="section">
                    <h2>Recommendations</h2>
                    {recommendations}
                </div>
            </div>
        </body>
        </html>
        """
        
        # Build summary table
        summary_rows = []
        for key, value in report.summary.items():
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    summary_rows.append(f"<tr><td>{key}_{sub_key}</td><td>{sub_value}</td></tr>")
            else:
                summary_rows.append(f"<tr><td>{key}</td><td>{value}</td></tr>")
        summary_table = f"<table>{''.join(summary_rows)}</table>"
        
        # Build performance table
        if 'performance' in report.tables:
            perf_df = report.tables['performance']
            perf_rows = []
            for _, row in perf_df.iterrows():
                perf_rows.append(f"<tr><td>{row.iloc[0]}</td><td>{row.iloc[1]}</td></tr>")
            performance_table = f"<table><tr><th>Metric</th><th>Value</th></tr>{''.join(perf_rows)}</table>"
        else:
            performance_table = "<p>No performance data available</p>"
        
        # Build risk table
        if 'risk' in report.tables:
            risk_df = report.tables['risk']
            risk_rows = []
            for _, row in risk_df.iterrows():
                risk_rows.append(f"<tr><td>{row.iloc[0]}</td><td>{row.iloc[1]}</td></tr>")
            risk_table = f"<table><tr><th>Metric</th><th>Value</th></tr>{''.join(risk_rows)}</table>"
        else:
            risk_table = "<p>No risk data available</p>"
        
        # Build charts
        charts_html = ""
        for chart_name, chart_data in report.charts.items():
            charts_html += f"""
            <div class="chart-container">
                <h3>{chart_name.replace('_', ' ').title()}</h3>
                <img src="data:image/png;base64,{chart_data}" alt="{chart_name}">
            </div>
            """
        
        # Build recommendations
        recommendations_html = ""
        for i, rec in enumerate(report.recommendations, 1):
            recommendations_html += f'<div class="recommendation">{i}. {rec}</div>'
        
        # Fill template
        html = html_template.format(
            title=report.title,
            generated_at=report.generated_at.strftime('%Y-%m-%d %H:%M:%S'),
            report_type=report.report_type.value,
            time_period=report.time_period.value,
            summary_table=summary_table,
            performance_table=performance_table,
            risk_table=risk_table,
            charts=charts_html,
            recommendations=recommendations_html or "<p>No recommendations</p>",
        )
        
        if output_path:
            with open(output_path, 'w') as f:
                f.write(html)
        
        return html
    
    async def _export_json(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as JSON."""
        data = asdict(report)
        
        # Convert datetime objects to strings
        data['generated_at'] = data['generated_at'].isoformat()
        
        # Convert DataFrames to dicts
        for key, df in data.get('tables', {}).items():
            if isinstance(df, pd.DataFrame):
                data['tables'][key] = df.to_dict('records')
        
        # Convert charts to strings
        for key, chart in data.get('charts', {}).items():
            if isinstance(chart, bytes):
                data['charts'][key] = base64.b64encode(chart).decode('utf-8')
        
        json_data = json.dumps(data, indent=2, default=str)
        
        if output_path:
            with open(output_path, 'w') as f:
                f.write(json_data)
        
        return json_data
    
    async def _export_csv(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as CSV."""
        csv_data = {}
        
        for table_name, df in report.tables.items():
            if isinstance(df, pd.DataFrame):
                csv_data[table_name] = df.to_csv(index=False)
        
        if output_path:
            with open(output_path, 'w') as f:
                for name, data in csv_data.items():
                    f.write(f"=== {name} ===\n")
                    f.write(data)
                    f.write("\n\n")
        
        return "\n".join(csv_data.values())
    
    async def _export_excel(self, report: ReportData, output_path: Optional[str] = None) -> bytes:
        """Export report as Excel."""
        try:
            import openpyxl
            from openpyxl.drawing.image import Image
            from openpyxl.utils.dataframe import dataframe_to_rows
            import io
            from PIL import Image as PILImage
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            row = 1
            for key, value in report.summary.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        ws_summary.cell(row, 1, f"{key}_{sub_key}")
                        ws_summary.cell(row, 2, str(sub_value))
                        row += 1
                else:
                    ws_summary.cell(row, 1, key)
                    ws_summary.cell(row, 2, str(value))
                    row += 1
            
            # Sheets for each table
            for table_name, df in report.tables.items():
                if isinstance(df, pd.DataFrame):
                    ws = wb.create_sheet(title=table_name[:31])
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
            
            # Charts sheet
            if report.charts:
                ws_charts = wb.create_sheet(title="Charts")
                row = 1
                for chart_name, chart_data in report.charts.items():
                    # Decode image
                    image_data = base64.b64decode(chart_data)
                    img_buffer = io.BytesIO(image_data)
                    img = PILImage.open(img_buffer)
                    
                    # Save to temp file
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                    img.save(temp_file.name)
                    temp_file.close()
                    
                    # Add to workbook
                    img_openpyxl = Image(temp_file.name)
                    ws_charts.add_image(img_openpyxl, f'A{row}')
                    ws_charts.cell(row, 1, chart_name)
                    row += 20
                    
                    # Clean up
                    os.unlink(temp_file.name)
            
            # Recommendations sheet
            if report.recommendations:
                ws_rec = wb.create_sheet(title="Recommendations")
                ws_rec.append(["#", "Recommendation"])
                for i, rec in enumerate(report.recommendations, 1):
                    ws_rec.append([i, rec])
            
            # Save
            if output_path:
                wb.save(output_path)
                return b""
            else:
                buffer = io.BytesIO()
                wb.save(buffer)
                return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            return b""
    
    async def _export_markdown(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as Markdown."""
        md = []
        
        # Title
        md.append(f"# {report.title}")
        md.append("")
        md.append(f"**Generated:** {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}")
        md.append(f"**Report Type:** {report.report_type.value}")
        md.append(f"**Time Period:** {report.time_period.value}")
        md.append("")
        
        # Summary
        md.append("## Summary")
        md.append("")
        md.append("| Metric | Value |")
        md.append("|--------|-------|")
        for key, value in report.summary.items():
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    md.append(f"| {key}_{sub_key} | {sub_value} |")
            else:
                md.append(f"| {key} | {value} |")
        md.append("")
        
        # Performance
        if 'performance' in report.tables:
            md.append("## Performance Metrics")
            md.append("")
            df = report.tables['performance']
            md.append(df.to_markdown(index=False))
            md.append("")
        
        # Risk
        if 'risk' in report.tables:
            md.append("## Risk Metrics")
            md.append("")
            df = report.tables['risk']
            md.append(df.to_markdown(index=False))
            md.append("")
        
        # Recommendations
        if report.recommendations:
            md.append("## Recommendations")
            md.append("")
            for i, rec in enumerate(report.recommendations, 1):
                md.append(f"{i}. {rec}")
            md.append("")
        
        # Charts (as links since markdown can't embed images)
        if report.charts:
            md.append("## Charts")
            md.append("")
            for chart_name in report.charts.keys():
                md.append(f"- {chart_name.replace('_', ' ').title()}")
            md.append("")
        
        output = "\n".join(md)
        
        if output_path:
            with open(output_path, 'w') as f:
                f.write(output)
        
        return output
    
    async def _export_text(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as plain text."""
        text = []
        
        # Title
        text.append(f"{'='*60}")
        text.append(f"{report.title}")
        text.append(f"{'='*60}")
        text.append("")
        text.append(f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}")
        text.append(f"Report Type: {report.report_type.value}")
        text.append(f"Time Period: {report.time_period.value}")
        text.append("")
        
        # Summary
        text.append("SUMMARY")
        text.append("-" * 40)
        for key, value in report.summary.items():
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    text.append(f"{key}_{sub_key}: {sub_value}")
            else:
                text.append(f"{key}: {value}")
        text.append("")
        
        # Performance
        if 'performance' in report.tables:
            text.append("PERFORMANCE METRICS")
            text.append("-" * 40)
            df = report.tables['performance']
            for _, row in df.iterrows():
                text.append(f"{row.iloc[0]}: {row.iloc[1]}")
            text.append("")
        
        # Risk
        if 'risk' in report.tables:
            text.append("RISK METRICS")
            text.append("-" * 40)
            df = report.tables['risk']
            for _, row in df.iterrows():
                text.append(f"{row.iloc[0]}: {row.iloc[1]}")
            text.append("")
        
        # Recommendations
        if report.recommendations:
            text.append("RECOMMENDATIONS")
            text.append("-" * 40)
            for i, rec in enumerate(report.recommendations, 1):
                text.append(f"{i}. {rec}")
            text.append("")
        
        output = "\n".join(text)
        
        if output_path:
            with open(output_path, 'w') as f:
                f.write(output)
        
        return output
    
    async def _export_dashboard(self, report: ReportData, output_path: Optional[str] = None) -> str:
        """Export report as dashboard (HTML with embedded charts)."""
        # Use HTML export with enhanced styling
        html = await self._export_html(report, output_path)
        
        # Add dashboard-specific styling
        dashboard_html = html.replace(
            '<div class="container">',
            '''<div class="container">
            <div class="dashboard-header">
                <div class="status-indicator">
                    <span class="badge badge-green">● LIVE</span>
                    <span class="badge">v1.0</span>
                </div>
            </div>
            ''',
        )
        
        return dashboard_html
    
    # ========================================================================
    # RECOMMENDATIONS
    # ========================================================================
    
    async def _generate_recommendations(
        self,
        performance: Dict[str, Any],
        risk: Dict[str, Any],
        portfolio: Dict[str, Any],
        trading: Dict[str, Any],
    ) -> List[str]:
        """Generate recommendations based on data."""
        recommendations = []
        
        # Performance recommendations
        if performance:
            metrics = performance.get("risk_adjusted", {})
            
            if metrics.get("sharpe_ratio", 0) < 0.5:
                recommendations.append("Consider improving risk-adjusted returns - Sharpe ratio below 0.5")
            
            if metrics.get("win_rate", 0) < 0.3:
                recommendations.append("Low win rate detected - consider reviewing strategy selection")
            
            if metrics.get("profit_factor", 0) < 1.0:
                recommendations.append("Profit factor below 1.0 - strategy may not be profitable")
        
        # Risk recommendations
        if risk:
            risk_metrics = risk.get("risk", {})
            
            if risk_metrics.get("max_drawdown", 0) > 0.2:
                recommendations.append("High drawdown detected - implement additional risk controls")
            
            if risk_metrics.get("var_95", 0) < -0.05:
                recommendations.append("VaR exceeds 5% - consider reducing position sizes")
            
            if risk_metrics.get("concentration_ratio", 0) > 0.25:
                recommendations.append("High portfolio concentration - consider diversification")
        
        # Portfolio recommendations
        if portfolio:
            portfolio_data = portfolio.get("current", {})
            
            if portfolio_data.get("leverage", 0) > 5:
                recommendations.append("High leverage detected - consider reducing exposure")
            
            if portfolio_data.get("position_count", 0) < 3:
                recommendations.append("Portfolio under-diversified - consider adding more positions")
        
        # Trading recommendations
        if trading:
            stats = trading.get("statistics", {})
            
            if stats.get("avg_slippage", 0) > 0.01:
                recommendations.append("High slippage - consider limit orders or different execution timing")
            
            if stats.get("fill_rate", 0) < 0.9:
                recommendations.append("Low fill rate - review order placement strategy")
        
        # Add general recommendations
        if not recommendations:
            recommendations.append("All metrics within acceptable ranges - continue monitoring")
        
        return recommendations
    
    # ========================================================================
    # SCHEDULED REPORTS
    # ========================================================================
    
    async def start_scheduler(self) -> None:
        """Start the report scheduler."""
        if self._is_running:
            logger.warning("Scheduler already running")
            return
        
        self._is_running = True
        self._scheduler_task = asyncio.create_task(self._scheduler_loop())
        logger.info("Report scheduler started")
    
    async def stop_scheduler(self) -> None:
        """Stop the report scheduler."""
        self._is_running = False
        
        if self._scheduler_task and not self._scheduler_task.done():
            self._scheduler_task.cancel()
            try:
                await self._scheduler_task
            except asyncio.CancelledError:
                pass
        
        logger.info("Report scheduler stopped")
    
    async def _scheduler_loop(self) -> None:
        """Background loop for scheduled reports."""
        schedule = self.config.get("schedule", {})
        
        while self._is_running:
            try:
                now = datetime.now()
                
                # Check daily report
                if schedule.get("daily", {}).get("enabled", False):
                    hour = schedule["daily"].get("hour", 23)
                    minute = schedule["daily"].get("minute", 59)
                    
                    if now.hour == hour and now.minute == minute:
                        await self._generate_scheduled_report(ReportType.DAILY)
                        await asyncio.sleep(60)  # Prevent multiple runs
                
                # Check weekly report
                if schedule.get("weekly", {}).get("enabled", False):
                    day = schedule["weekly"].get("day", "sunday")
                    hour = schedule["weekly"].get("hour", 23)
                    minute = schedule["weekly"].get("minute", 59)
                    
                    if now.strftime("%A").lower() == day.lower() and now.hour == hour and now.minute == minute:
                        await self._generate_scheduled_report(ReportType.WEEKLY)
                        await asyncio.sleep(60)
                
                # Check monthly report
                if schedule.get("monthly", {}).get("enabled", False):
                    day = schedule["monthly"].get("day", 1)
                    hour = schedule["monthly"].get("hour", 23)
                    minute = schedule["monthly"].get("minute", 59)
                    
                    if now.day == day and now.hour == hour and now.minute == minute:
                        await self._generate_scheduled_report(ReportType.MONTHLY)
                        await asyncio.sleep(60)
                
                await asyncio.sleep(30)
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Error in scheduler loop: {e}")
                await asyncio.sleep(60)
    
    async def _generate_scheduled_report(self, report_type: ReportType) -> None:
        """Generate a scheduled report."""
        config = ReportConfig(
            report_type=report_type,
            time_period=TimePeriod.CUSTOM,
            format=ReportFormat.HTML,
            include_charts=True,
            include_tables=True,
            include_summary=True,
            include_details=True,
            include_recommendations=True,
            include_risk_metrics=True,
            include_performance_metrics=True,
            include_trade_history=True,
            include_portfolio_snapshot=True,
        )
        
        report = await self.generate_report(config)
        
        # Export and deliver
        html = await self.export_report(report, ReportFormat.HTML)
        
        # Send email if configured
        email_config = self.config.get("email", {})
        if email_config.get("enabled", False):
            await self._send_report_email(report, html)
        
        # Send webhook if configured
        webhook_config = self.config.get("webhook", {})
        if webhook_config.get("enabled", False):
            await self._send_report_webhook(report)
        
        logger.info(f"Scheduled report generated: {report_type.value}")
    
    async def _send_report_email(self, report: ReportData, content: str) -> None:
        """Send report via email."""
        email_config = self.config.get("email", {})
        
        try:
            import aiohttp
            
            # Use email service (example)
            async with aiohttp.ClientSession() as session:
                data = {
                    "to": email_config.get("to", []),
                    "subject": f"{report.title}",
                    "body": content,
                    "html": True,
                }
                await session.post(email_config.get("url"), json=data)
                
        except Exception as e:
            logger.error(f"Error sending report email: {e}")
    
    async def _send_report_webhook(self, report: ReportData) -> None:
        """Send report via webhook."""
        webhook_config = self.config.get("webhook", {})
        
        try:
            import aiohttp
            
            data = asdict(report)
            data['generated_at'] = data['generated_at'].isoformat()
            
            async with aiohttp.ClientSession() as session:
                await session.post(webhook_config.get("url"), json=data)
                
        except Exception as e:
            logger.error(f"Error sending report webhook: {e}")
    
    # ========================================================================
    # UTILITY METHODS
    # ========================================================================
    
    def get_report(self, report_id: str) -> Optional[ReportData]:
        """Get a report by ID."""
        return self._reports.get(report_id)
    
    def get_reports(self, limit: int = 100) -> List[Dict[str, Any]]:
        """Get recent reports."""
        return self._report_history[-limit:]
    
    def clear_cache(self) -> None:
        """Clear report cache."""
        self._data_cache.clear()
        self._chart_cache.clear()
        logger.info("Report cache cleared")
    
    def get_stats(self) -> Dict[str, Any]:
        """Get service statistics."""
        return {
            "total_reports": len(self._reports),
            "report_history": len(self._report_history),
            "cache_size": len(self._data_cache),
            "chart_cache_size": len(self._chart_cache),
            "is_running": self._is_running,
        }


# ========================================================================
# FACTORY FUNCTION
# ========================================================================

def create_reporting_service(
    config: Dict[str, Any],
    risk_manager: Optional[Any] = None,
    portfolio_manager: Optional[Any] = None,
    trading_engine: Optional[Any] = None,
) -> ReportingService:
    """Factory function to create a ReportingService instance."""
    return ReportingService(
        config=config,
        risk_manager=risk_manager,
        portfolio_manager=portfolio_manager,
        trading_engine=trading_engine,
    )
