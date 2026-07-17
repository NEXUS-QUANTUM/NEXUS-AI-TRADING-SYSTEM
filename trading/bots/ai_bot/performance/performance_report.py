"""
NEXUS AI TRADING SYSTEM - Performance Report
Copyright © 2026 NEXUS QUANTUM LTD - All Rights Reserved
CEO: Dr X... - Majority Shareholder

Advanced performance reporting system for trading bots, models, and strategies
with comprehensive report generation, templating, and multi-format export.
"""

import asyncio
import json
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import pandas as pd
import yaml
from jinja2 import Environment, FileSystemLoader, Template
from prometheus_client import Counter, Histogram

from shared.utilities.logger import get_logger
from shared.utilities.metrics import MetricsCollector

from .performance_metrics import PerformanceMetrics, PerformanceMetricsCalculator
from .performance_analyzer import PerformanceAnalyzer, PerformanceReport

logger = get_logger(__name__)

# Prometheus metrics
REPORT_COUNTER = Counter(
    "nexus_performance_reports_total",
    "Total number of performance reports generated",
    ["report_type", "format", "status"],
)
REPORT_GENERATION_DURATION = Histogram(
    "nexus_performance_report_duration_seconds",
    "Duration of report generation",
    ["report_type"],
)


class ReportType(Enum):
    """Types of performance reports."""

    EXECUTIVE = "executive"
    DETAILED = "detailed"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    COMPARATIVE = "comparative"
    RISK = "risk"
    MODEL = "model"
    CUSTOM = "custom"


class ReportFormat(Enum):
    """Report export formats."""

    HTML = "html"
    PDF = "pdf"
    JSON = "json"
    YAML = "yaml"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"
    TEXT = "text"


@dataclass
class ReportConfig:
    """Configuration for report generation."""

    name: str
    type: ReportType
    title: str
    description: str = ""
    components: List[str] = field(default_factory=list)
    metrics: List[str] = field(default_factory=list)
    time_range: Optional[Tuple[datetime, datetime]] = None
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_recommendations: bool = True
    template_path: Optional[Path] = None
    output_path: Optional[Path] = None
    formats: List[ReportFormat] = field(default_factory=lambda: [ReportFormat.HTML])
    schedules: List[str] = field(default_factory=list)
    recipients: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "name": self.name,
            "type": self.type.value,
            "title": self.title,
            "description": self.description,
            "components": self.components,
            "metrics": self.metrics,
            "time_range": {
                "start": self.time_range[0].isoformat() if self.time_range else None,
                "end": self.time_range[1].isoformat() if self.time_range else None,
            },
            "include_charts": self.include_charts,
            "include_tables": self.include_tables,
            "include_summary": self.include_summary,
            "include_recommendations": self.include_recommendations,
            "template_path": str(self.template_path) if self.template_path else None,
            "output_path": str(self.output_path) if self.output_path else None,
            "formats": [f.value for f in self.formats],
            "schedules": self.schedules,
            "recipients": self.recipients,
            "metadata": self.metadata,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ReportConfig":
        """Create from dictionary."""
        return cls(
            name=data["name"],
            type=ReportType(data["type"]),
            title=data["title"],
            description=data.get("description", ""),
            components=data.get("components", []),
            metrics=data.get("metrics", []),
            time_range=(
                datetime.fromisoformat(data["time_range"]["start"]) if data.get("time_range", {}).get("start") else None,
                datetime.fromisoformat(data["time_range"]["end"]) if data.get("time_range", {}).get("end") else None,
            ),
            include_charts=data.get("include_charts", True),
            include_tables=data.get("include_tables", True),
            include_summary=data.get("include_summary", True),
            include_recommendations=data.get("include_recommendations", True),
            template_path=Path(data["template_path"]) if data.get("template_path") else None,
            output_path=Path(data["output_path"]) if data.get("output_path") else None,
            formats=[ReportFormat(f) for f in data.get("formats", ["html"])],
            schedules=data.get("schedules", []),
            recipients=data.get("recipients", []),
            metadata=data.get("metadata", {}),
        )


@dataclass
class ReportData:
    """Data for report generation."""

    config: ReportConfig
    performance_data: Dict[str, Any]
    metrics: PerformanceMetrics
    analysis: Optional[PerformanceReport] = None
    charts: List[Dict[str, Any]] = field(default_factory=list)
    tables: List[Dict[str, Any]] = field(default_factory=list)
    recommendations: List[Dict[str, Any]] = field(default_factory=list)
    summary: Dict[str, Any] = field(default_factory=dict)
    raw_data: Dict[str, Any] = field(default_factory=dict)
    generated_at: datetime = field(default_factory=datetime.utcnow)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "config": self.config.to_dict(),
            "performance_data": self.performance_data,
            "metrics": self.metrics.to_dict() if self.metrics else {},
            "analysis": self.analysis.to_dict() if self.analysis else None,
            "charts": self.charts,
            "tables": self.tables,
            "recommendations": self.recommendations,
            "summary": self.summary,
            "raw_data": self.raw_data,
            "generated_at": self.generated_at.isoformat(),
        }


class PerformanceReportGenerator:
    """
    Advanced performance report generation system.
    """

    def __init__(
        self,
        config: Optional[Dict[str, Any]] = None,
        performance_analyzer: Optional[PerformanceAnalyzer] = None,
        metrics_collector: Optional[MetricsCollector] = None,
    ):
        """
        Initialize the report generator.

        Args:
            config: Configuration dictionary
            performance_analyzer: Performance analyzer instance
            metrics_collector: Metrics collector instance
        """
        self.config = config or {}
        self.performance_analyzer = performance_analyzer
        self.metrics_collector = metrics_collector or MetricsCollector()
        self._lock = asyncio.Lock()
        self._reports: Dict[str, ReportData] = {}
        self._configs: Dict[str, ReportConfig] = {}
        self._templates: Dict[str, Template] = {}
        self._scheduled_tasks: Dict[str, asyncio.Task] = {}

        # Load configuration
        self.report_config = self.config.get("performance_report", {})
        self.templates_path = Path(self.report_config.get("templates_path", "./configs/reports/templates"))
        self.output_path = Path(self.report_config.get("output_path", "./reports"))
        self.configs_path = Path(self.report_config.get("configs_path", "./configs/reports/configs.yaml"))
        self.max_reports = self.report_config.get("max_reports", 100)
        self.default_report_type = ReportType(self.report_config.get("default_report_type", "detailed"))

        # Create directories
        self.templates_path.mkdir(parents=True, exist_ok=True)
        self.output_path.mkdir(parents=True, exist_ok=True)

        # Load templates
        self._load_templates()

        # Load report configs
        self._load_configs()

        # Register default report configs
        self._register_default_configs()

        logger.info(f"PerformanceReportGenerator initialized with {len(self._configs)} configs")

    def _load_templates(self):
        """Load report templates."""
        try:
            env = Environment(
                loader=FileSystemLoader(str(self.templates_path)),
                autoescape=True,
            )

            # Load default templates if directory is empty
            if not list(self.templates_path.glob("*.html")):
                self._create_default_templates()

            # Register templates
            for template_file in self.templates_path.glob("*.html"):
                template_name = template_file.stem
                self._templates[template_name] = env.get_template(template_file.name)

            logger.info(f"Loaded {len(self._templates)} templates")

        except Exception as e:
            logger.error(f"Error loading templates: {e}")
            self._create_default_templates()

    def _create_default_templates(self):
        """Create default report templates."""
        # Executive Summary Template
        executive_template = """<!DOCTYPE html>
<html>
<head>
    <title>{{ title }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #2c3e50; }
        .summary { background: #ecf0f1; padding: 20px; border-radius: 5px; }
        .metric-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
        .metric-card { background: white; padding: 15px; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }
        .metric-label { font-size: 14px; color: #7f8c8d; }
        .chart-container { margin: 20px 0; padding: 20px; background: white; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .table-container { margin: 20px 0; overflow-x: auto; }
        table { width: 100%; border-collapse: collapse; }
        th, td { padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #34495e; color: white; }
        .recommendation { background: #f8f9fa; padding: 10px; margin: 10px 0; border-left: 4px solid #3498db; }
        .footer { margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #7f8c8d; }
    </style>
</head>
<body>
    <h1>{{ title }}</h1>
    <p><strong>Generated:</strong> {{ generated_at }}</p>
    <p><strong>Period:</strong> {{ time_range.start }} to {{ time_range.end }}</p>

    <div class="summary">
        <h2>Executive Summary</h2>
        <p>{{ summary_text }}</p>
    </div>

    <h2>Key Metrics</h2>
    <div class="metric-grid">
        {% for metric in key_metrics %}
        <div class="metric-card">
            <div class="metric-value">{{ metric.value }}</div>
            <div class="metric-label">{{ metric.label }}</div>
        </div>
        {% endfor %}
    </div>

    <h2>Performance Analysis</h2>
    <div class="metric-grid">
        {% for metric in performance_metrics %}
        <div class="metric-card">
            <div class="metric-value">{{ metric.value }}</div>
            <div class="metric-label">{{ metric.label }}</div>
        </div>
        {% endfor %}
    </div>

    {% if charts %}
    <h2>Charts</h2>
    {% for chart in charts %}
    <div class="chart-container">
        <h3>{{ chart.title }}</h3>
        <img src="{{ chart.url }}" alt="{{ chart.title }}" />
    </div>
    {% endfor %}
    {% endif %}

    {% if tables %}
    <h2>Detailed Data</h2>
    {% for table in tables %}
    <div class="table-container">
        <h3>{{ table.title }}</h3>
        {{ table.html }}
    </div>
    {% endfor %}
    {% endif %}

    {% if recommendations %}
    <h2>Recommendations</h2>
    {% for rec in recommendations %}
    <div class="recommendation">
        <strong>{{ rec.title }}</strong><br />
        {{ rec.description }}
        <br /><small>Priority: {{ rec.priority }}</small>
    </div>
    {% endfor %}
    {% endif %}

    <div class="footer">
        <p>Generated by Nexus AI Trading System - Performance Report</p>
        <p>Report ID: {{ report_id }}</p>
    </div>
</body>
</html>"""

        # Save templates
        (self.templates_path / "executive.html").write_text(executive_template)
        (self.templates_path / "detailed.html").write_text(executive_template)
        (self.templates_path / "daily.html").write_text(executive_template)
        (self.templates_path / "weekly.html").write_text(executive_template)
        (self.templates_path / "monthly.html").write_text(executive_template)
        (self.templates_path / "quarterly.html").write_text(executive_template)
        (self.templates_path / "annual.html").write_text(executive_template)

        logger.info("Created default templates")

    def _load_configs(self):
        """Load report configurations."""
        try:
            if self.configs_path.exists():
                with open(self.configs_path, "r") as f:
                    data = yaml.safe_load(f)
                    for config_data in data.get("reports", []):
                        config = ReportConfig.from_dict(config_data)
                        self._configs[config.name] = config
                logger.info(f"Loaded {len(self._configs)} report configs")
        except Exception as e:
            logger.error(f"Error loading report configs: {e}")

    def _register_default_configs(self):
        """Register default report configurations."""
        default_configs = [
            ReportConfig(
                name="daily_report",
                type=ReportType.DAILY,
                title="Daily Performance Report",
                description="Daily performance overview for trading operations",
                components=["trading", "system"],
                metrics=["sharpe_ratio", "win_rate", "total_return", "max_drawdown"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.JSON],
                schedules=["daily"],
                recipients=["team@nexustradingia.com"],
            ),
            ReportConfig(
                name="weekly_report",
                type=ReportType.WEEKLY,
                title="Weekly Performance Report",
                description="Weekly performance analysis for trading operations",
                components=["trading", "system", "model"],
                metrics=["sharpe_ratio", "sortino_ratio", "win_rate", "profit_factor", "max_drawdown"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.EXCEL],
                schedules=["weekly"],
                recipients=["team@nexustradingia.com", "management@nexustradingia.com"],
            ),
            ReportConfig(
                name="monthly_report",
                type=ReportType.MONTHLY,
                title="Monthly Performance Report",
                description="Comprehensive monthly performance report",
                components=["trading", "system", "model", "risk"],
                metrics=["sharpe_ratio", "sortino_ratio", "calmar_ratio", "win_rate", "profit_factor",
                        "max_drawdown", "var_95", "cvar_95"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.EXCEL],
                schedules=["monthly"],
                recipients=["team@nexustradingia.com", "management@nexustradingia.com", "investors@nexustradingia.com"],
            ),
            ReportConfig(
                name="quarterly_report",
                type=ReportType.QUARTERLY,
                title="Quarterly Performance Report",
                description="Quarterly performance review with strategic insights",
                components=["trading", "system", "model", "risk", "performance"],
                metrics=["sharpe_ratio", "sortino_ratio", "calmar_ratio", "omega_ratio",
                        "win_rate", "profit_factor", "max_drawdown", "recovery_factor",
                        "var_95", "cvar_95", "beta", "alpha"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.EXCEL, ReportFormat.JSON],
                schedules=["quarterly"],
                recipients=["team@nexustradingia.com", "management@nexustradingia.com",
                           "investors@nexustradingia.com", "board@nexustradingia.com"],
            ),
            ReportConfig(
                name="annual_report",
                type=ReportType.ANNUAL,
                title="Annual Performance Report",
                description="Comprehensive annual performance and strategy review",
                components=["trading", "system", "model", "risk", "performance", "strategy"],
                metrics=["sharpe_ratio", "sortino_ratio", "calmar_ratio", "omega_ratio", "information_ratio",
                        "win_rate", "profit_factor", "max_drawdown", "recovery_factor",
                        "var_95", "cvar_95", "beta", "alpha", "total_return", "annual_return"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.EXCEL, ReportFormat.JSON],
                schedules=["annual"],
                recipients=["team@nexustradingia.com", "management@nexustradingia.com",
                           "investors@nexustradingia.com", "board@nexustradingia.com"],
            ),
            ReportConfig(
                name="risk_report",
                type=ReportType.RISK,
                title="Risk Management Report",
                description="Detailed risk analysis and management report",
                components=["risk", "system"],
                metrics=["max_drawdown", "var_95", "cvar_95", "beta", "alpha",
                        "max_risk_exposure", "avg_risk_exposure", "sharpe_ratio", "sortino_ratio"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.PDF],
                schedules=["weekly"],
                recipients=["risk@nexustradingia.com", "management@nexustradingia.com"],
            ),
            ReportConfig(
                name="model_report",
                type=ReportType.MODEL,
                title="Model Performance Report",
                description="Detailed model performance and accuracy report",
                components=["model"],
                metrics=["accuracy", "precision", "recall", "f1_score", "auc_roc",
                        "mse", "rmse", "mae", "r2_score"],
                include_charts=True,
                include_tables=True,
                include_summary=True,
                include_recommendations=True,
                formats=[ReportFormat.HTML, ReportFormat.JSON],
                schedules=["weekly"],
                recipients=["team@nexustradingia.com", "dev@nexustradingia.com"],
            ),
        ]

        for config in default_configs:
            if config.name not in self._configs:
                self._configs[config.name] = config

        logger.info(f"Registered {len(default_configs)} default report configs")

    async def generate_report(
        self,
        config: Union[ReportConfig, str, Dict[str, Any]],
        data: Optional[Dict[str, Any]] = None,
        format: Optional[Union[ReportFormat, str]] = None,
    ) -> Dict[str, Any]:
        """
        Generate a performance report.

        Args:
            config: Report configuration or name
            data: Additional data for the report
            format: Output format

        Returns:
            Generated report data
        """
        start_time = time.time()

        # Parse config
        if isinstance(config, str):
            config = self._configs.get(config)
            if not config:
                raise ValueError(f"Report config not found: {config}")
        elif isinstance(config, dict):
            config = ReportConfig.from_dict(config)

        if not isinstance(config, ReportConfig):
            raise ValueError("Invalid report configuration")

        # Parse format
        if format:
            if isinstance(format, str):
                format = ReportFormat(format)
        else:
            format = config.formats[0] if config.formats else ReportFormat.HTML

        # Create report data
        report_data = ReportData(
            config=config,
            performance_data=data or {},
            metrics=PerformanceMetrics(),
            generated_at=datetime.utcnow(),
        )

        # Collect performance data
        await self._collect_performance_data(report_data)

        # Calculate metrics
        await self._calculate_metrics(report_data)

        # Generate analysis
        await self._generate_analysis(report_data)

        # Generate charts
        if config.include_charts:
            await self._generate_charts(report_data)

        # Generate tables
        if config.include_tables:
            await self._generate_tables(report_data)

        # Generate summary
        if config.include_summary:
            await self._generate_summary(report_data)

        # Generate recommendations
        if config.include_recommendations:
            await self._generate_recommendations(report_data)

        # Generate reports in requested formats
        results = {}
        for fmt in [format] if format else config.formats:
            result = await self._export_report(report_data, fmt)
            results[fmt.value] = result

        # Store report
        async with self._lock:
            report_id = f"report_{int(time.time())}"
            self._reports[report_id] = report_data

            # Limit stored reports
            if len(self._reports) > self.max_reports:
                oldest = sorted(self._reports.keys())[0]
                del self._reports[oldest]

        # Record metrics
        REPORT_COUNTER.labels(
            report_type=config.type.value,
            format=format.value,
            status="success",
        ).inc()
        REPORT_GENERATION_DURATION.labels(
            report_type=config.type.value,
        ).observe(time.time() - start_time)

        logger.info(f"Report generated: {config.name} ({format.value})")

        return results

    async def _collect_performance_data(self, report_data: ReportData):
        """Collect performance data for the report."""
        try:
            # If performance analyzer is available, use it
            if self.performance_analyzer:
                for component in report_data.config.components:
                    latest_report = await self.performance_analyzer.get_latest_report(component)
                    if latest_report:
                        report_data.performance_data[component] = latest_report.to_dict()

            # Add system metrics
            import psutil
            report_data.performance_data["system"] = {
                "cpu_usage": psutil.cpu_percent(),
                "memory_usage": psutil.virtual_memory().percent,
                "disk_usage": psutil.disk_usage("/").percent,
                "timestamp": datetime.utcnow().isoformat(),
            }

        except Exception as e:
            logger.error(f"Error collecting performance data: {e}")

    async def _calculate_metrics(self, report_data: ReportData):
        """Calculate performance metrics."""
        try:
            # Extract returns from performance data
            returns = []
            trading_data = report_data.performance_data.get("trading", {})

            if "returns" in trading_data:
                returns = trading_data["returns"]
            elif "pnl" in trading_data:
                # Convert PnL to returns
                pnl = trading_data["pnl"]
                if pnl and len(pnl) > 1:
                    returns = np.diff(pnl) / pnl[:-1]

            if returns:
                calculator = PerformanceMetricsCalculator()
                report_data.metrics = calculator.calculate_metrics(
                    returns=returns,
                    trades=trading_data.get("trades"),
                    predictions=trading_data.get("predictions"),
                    actuals=trading_data.get("actuals"),
                )

        except Exception as e:
            logger.error(f"Error calculating metrics: {e}")

    async def _generate_analysis(self, report_data: ReportData):
        """Generate performance analysis."""
        if self.performance_analyzer:
            try:
                component = report_data.config.components[0] if report_data.config.components else "system"
                report_data.analysis = await self.performance_analyzer.analyze_performance(
                    component=component,
                    metrics=report_data.config.metrics,
                    time_range=report_data.config.time_range,
                )
            except Exception as e:
                logger.error(f"Error generating analysis: {e}")

    async def _generate_charts(self, report_data: ReportData):
        """Generate charts for the report."""
        # TODO: Implement chart generation using matplotlib/plotly
        # For now, just create placeholder data
        report_data.charts = [
            {
                "title": "Equity Curve",
                "url": "/charts/equity_curve.png",
                "type": "line",
            },
            {
                "title": "Monthly Returns",
                "url": "/charts/monthly_returns.png",
                "type": "heatmap",
            },
            {
                "title": "Drawdown",
                "url": "/charts/drawdown.png",
                "type": "area",
            },
        ]

    async def _generate_tables(self, report_data: ReportData):
        """Generate tables for the report."""
        # Create table from metrics
        metrics_dict = report_data.metrics.to_dict()
        tables = []

        # Performance table
        perf_table = {
            "title": "Performance Metrics",
            "data": [],
        }

        for category, metrics in metrics_dict.items():
            for metric_name, value in metrics.items():
                if isinstance(value, (int, float)):
                    perf_table["data"].append({
                        "Category": category.capitalize(),
                        "Metric": metric_name.replace("_", " ").title(),
                        "Value": f"{value:.4f}",
                    })

        # Convert to HTML table
        if perf_table["data"]:
            df = pd.DataFrame(perf_table["data"])
            perf_table["html"] = df.to_html(index=False, classes="table table-striped")
            tables.append(perf_table)

        report_data.tables = tables

    async def _generate_summary(self, report_data: ReportData):
        """Generate summary for the report."""
        metrics = report_data.metrics

        summary = {
            "overall_performance": "Good",
            "risk_level": "Moderate",
            "key_insights": [],
            "statistics": {},
        }

        # Add key insights based on metrics
        if metrics.sharpe_ratio > 2:
            summary["key_insights"].append("Excellent risk-adjusted returns")
        elif metrics.sharpe_ratio > 1:
            summary["key_insights"].append("Good risk-adjusted returns")
        elif metrics.sharpe_ratio > 0:
            summary["key_insights"].append("Positive risk-adjusted returns")

        if metrics.max_drawdown < 0.05:
            summary["key_insights"].append("Low drawdown profile")
        elif metrics.max_drawdown < 0.15:
            summary["key_insights"].append("Moderate drawdown profile")

        if metrics.win_rate > 0.6:
            summary["key_insights"].append("High win rate")
        elif metrics.win_rate > 0.4:
            summary["key_insights"].append("Moderate win rate")

        # Add statistics
        summary["statistics"] = {
            "sharpe_ratio": metrics.sharpe_ratio,
            "max_drawdown": metrics.max_drawdown,
            "win_rate": metrics.win_rate,
            "total_return": metrics.total_return,
            "annual_return": metrics.annual_return,
        }

        report_data.summary = summary

    async def _generate_recommendations(self, report_data: ReportData):
        """Generate recommendations for the report."""
        recommendations = []

        metrics = report_data.metrics

        # Sharpe Ratio recommendations
        if metrics.sharpe_ratio < 1:
            recommendations.append({
                "title": "Improve Risk-Adjusted Returns",
                "description": "Sharpe ratio is below 1. Consider reducing risk exposure or improving strategy performance.",
                "priority": "High",
                "category": "Risk Management",
            })
        elif metrics.sharpe_ratio < 2:
            recommendations.append({
                "title": "Optimize Risk-Reward Balance",
                "description": "Sharpe ratio is moderate. Fine-tune position sizing and risk parameters.",
                "priority": "Medium",
                "category": "Optimization",
            })

        # Drawdown recommendations
        if metrics.max_drawdown > 0.2:
            recommendations.append({
                "title": "Reduce Drawdown Risk",
                "description": f"Maximum drawdown is {metrics.max_drawdown:.2%}. Implement tighter stop-losses and position limits.",
                "priority": "High",
                "category": "Risk Management",
            })

        # Win Rate recommendations
        if metrics.win_rate < 0.4:
            recommendations.append({
                "title": "Improve Win Rate",
                "description": f"Win rate is {metrics.win_rate:.2%}. Review entry conditions and signal quality.",
                "priority": "Medium",
                "category": "Strategy Improvement",
            })

        # Profit Factor recommendations
        if metrics.profit_factor < 1.5:
            recommendations.append({
                "title": "Increase Profit Factor",
                "description": f"Profit factor is {metrics.profit_factor:.2f}. Focus on cutting losses and letting winners run.",
                "priority": "Medium",
                "category": "Strategy Improvement",
            })

        report_data.recommendations = recommendations

    async def _export_report(
        self,
        report_data: ReportData,
        format: ReportFormat,
    ) -> Dict[str, Any]:
        """
        Export report in specified format.

        Args:
            report_data: Report data
            format: Export format

        Returns:
            Export result
        """
        result = {
            "format": format.value,
            "path": None,
            "data": None,
        }

        try:
            if format == ReportFormat.HTML:
                result["data"] = await self._export_html(report_data)
            elif format == ReportFormat.JSON:
                result["data"] = json.dumps(report_data.to_dict(), indent=2)
                result["path"] = await self._write_file(report_data, "json", result["data"])
            elif format == ReportFormat.YAML:
                result["data"] = yaml.dump(report_data.to_dict(), default_flow_style=False)
                result["path"] = await self._write_file(report_data, "yaml", result["data"])
            elif format == ReportFormat.CSV:
                result["data"] = await self._export_csv(report_data)
                result["path"] = await self._write_file(report_data, "csv", result["data"])
            elif format == ReportFormat.MARKDOWN:
                result["data"] = await self._export_markdown(report_data)
                result["path"] = await self._write_file(report_data, "md", result["data"])
            elif format == ReportFormat.TEXT:
                result["data"] = await self._export_text(report_data)
                result["path"] = await self._write_file(report_data, "txt", result["data"])
            elif format == ReportFormat.EXCEL:
                result["path"] = await self._export_excel(report_data)
            elif format == ReportFormat.PDF:
                # PDF generation would require additional libraries
                # For now, fallback to HTML
                result["data"] = await self._export_html(report_data)
                result["path"] = await self._write_file(report_data, "html", result["data"])
            else:
                raise ValueError(f"Unsupported format: {format}")

            return result

        except Exception as e:
            logger.error(f"Error exporting report: {e}")
            return result

    async def _export_html(self, report_data: ReportData) -> str:
        """Export report as HTML."""
        # Use template
        template_name = report_data.config.type.value
        if template_name not in self._templates:
            template_name = "detailed"

        template = self._templates.get(template_name)

        if not template:
            return "<html><body>Report data unavailable</body></html>"

        # Render template
        context = {
            "title": report_data.config.title,
            "generated_at": report_data.generated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "time_range": {
                "start": report_data.config.time_range[0].strftime("%Y-%m-%d") if report_data.config.time_range else "N/A",
                "end": report_data.config.time_range[1].strftime("%Y-%m-%d") if report_data.config.time_range else "N/A",
            },
            "summary_text": report_data.summary.get("overall_performance", "Performance summary not available"),
            "key_metrics": [
                {"label": "Sharpe Ratio", "value": f"{report_data.metrics.sharpe_ratio:.3f}"},
                {"label": "Win Rate", "value": f"{report_data.metrics.win_rate:.2%}"},
                {"label": "Max Drawdown", "value": f"{report_data.metrics.max_drawdown:.2%}"},
                {"label": "Total Return", "value": f"{report_data.metrics.total_return:.2%}"},
            ],
            "performance_metrics": [
                {"label": "Sortino Ratio", "value": f"{report_data.metrics.sortino_ratio:.3f}"},
                {"label": "Calmar Ratio", "value": f"{report_data.metrics.calmar_ratio:.3f}"},
                {"label": "Profit Factor", "value": f"{report_data.metrics.profit_factor:.3f}"},
                {"label": "Annual Return", "value": f"{report_data.metrics.annual_return:.2%}"},
            ],
            "charts": report_data.charts,
            "tables": report_data.tables,
            "recommendations": report_data.recommendations,
            "report_id": f"report_{int(time.time())}",
        }

        return template.render(**context)

    async def _export_csv(self, report_data: ReportData) -> str:
        """Export report as CSV."""
        # Convert metrics to CSV
        metrics_dict = report_data.metrics.to_dict()
        rows = []

        for category, metrics in metrics_dict.items():
            for metric_name, value in metrics.items():
                if isinstance(value, (int, float)):
                    rows.append({
                        "Category": category.capitalize(),
                        "Metric": metric_name.replace("_", " ").title(),
                        "Value": value,
                    })

        df = pd.DataFrame(rows)
        return df.to_csv(index=False)

    async def _export_markdown(self, report_data: ReportData) -> str:
        """Export report as Markdown."""
        lines = [
            f"# {report_data.config.title}",
            "",
            f"**Generated:** {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## Executive Summary",
            "",
            report_data.summary.get("overall_performance", "Performance summary not available"),
            "",
            "## Key Metrics",
            "",
            "| Metric | Value |",
            "|--------|-------|",
        ]

        key_metrics = [
            ("Sharpe Ratio", f"{report_data.metrics.sharpe_ratio:.3f}"),
            ("Win Rate", f"{report_data.metrics.win_rate:.2%}"),
            ("Max Drawdown", f"{report_data.metrics.max_drawdown:.2%}"),
            ("Total Return", f"{report_data.metrics.total_return:.2%}"),
            ("Sortino Ratio", f"{report_data.metrics.sortino_ratio:.3f}"),
            ("Profit Factor", f"{report_data.metrics.profit_factor:.3f}"),
        ]

        for metric, value in key_metrics:
            lines.append(f"| {metric} | {value} |")

        lines.append("")

        if report_data.recommendations:
            lines.append("## Recommendations")
            lines.append("")
            for rec in report_data.recommendations:
                lines.append(f"### {rec['title']}")
                lines.append("")
                lines.append(rec['description'])
                lines.append("")
                lines.append(f"**Priority:** {rec['priority']}")
                lines.append("")

        return "\n".join(lines)

    async def _export_text(self, report_data: ReportData) -> str:
        """Export report as plain text."""
        lines = [
            f"{'=' * 80}",
            f"{report_data.config.title}",
            f"{'=' * 80}",
            "",
            f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "EXECUTIVE SUMMARY",
            "-" * 40,
            report_data.summary.get("overall_performance", "Performance summary not available"),
            "",
            "KEY METRICS",
            "-" * 40,
        ]

        key_metrics = [
            ("Sharpe Ratio", f"{report_data.metrics.sharpe_ratio:.3f}"),
            ("Win Rate", f"{report_data.metrics.win_rate:.2%}"),
            ("Max Drawdown", f"{report_data.metrics.max_drawdown:.2%}"),
            ("Total Return", f"{report_data.metrics.total_return:.2%}"),
            ("Sortino Ratio", f"{report_data.metrics.sortino_ratio:.3f}"),
            ("Profit Factor", f"{report_data.metrics.profit_factor:.3f}"),
        ]

        for metric, value in key_metrics:
            lines.append(f"{metric:>20}: {value}")

        if report_data.recommendations:
            lines.append("")
            lines.append("RECOMMENDATIONS")
            lines.append("-" * 40)

            for i, rec in enumerate(report_data.recommendations, 1):
                lines.append(f"{i}. {rec['title']}")
                lines.append(f"   {rec['description']}")
                lines.append(f"   Priority: {rec['priority']}")
                lines.append("")

        lines.append("")
        lines.append(f"{'=' * 80}")
        lines.append("End of Report")

        return "\n".join(lines)

    async def _export_excel(self, report_data: ReportData) -> Path:
        """Export report as Excel."""
        output_path = self.output_path / report_data.config.name
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_path = output_path / f"{report_data.config.name}_{timestamp}.xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = "Summary"

            # Add title
            ws["A1"] = report_data.config.title
            ws["A1"].font = Font(size=16, bold=True)
            ws.merge_cells("A1:C1")

            # Add metadata
            ws["A3"] = "Generated:"
            ws["B3"] = report_data.generated_at.strftime("%Y-%m-%d %H:%M:%S")

            # Add metrics
            row = 5
            ws[f"A{row}"] = "Metric"
            ws[f"B{row}"] = "Value"
            ws[f"{row}:{row}"].font = Font(bold=True)

            metrics_dict = report_data.metrics.to_dict()
            for category, metrics in metrics_dict.items():
                row += 1
                ws[f"A{row}"] = category.capitalize()
                ws[f"A{row}"].font = Font(bold=True)
                ws.merge_cells(f"A{row}:B{row}")

                for metric_name, value in metrics.items():
                    if isinstance(value, (int, float)):
                        row += 1
                        ws[f"A{row}"] = metric_name.replace("_", " ").title()
                        ws[f"B{row}"] = f"{value:.4f}" if isinstance(value, float) else value

            # Recommendations sheet
            if report_data.recommendations:
                ws2 = wb.create_sheet("Recommendations")
                ws2["A1"] = "Recommendations"
                ws2["A1"].font = Font(size=14, bold=True)

                ws2["A3"] = "Title"
                ws2["B3"] = "Description"
                ws2["C3"] = "Priority"
                ws2["D3"] = "Category"

                for row, rec in enumerate(report_data.recommendations, 4):
                    ws2[f"A{row}"] = rec["title"]
                    ws2[f"B{row}"] = rec["description"]
                    ws2[f"C{row}"] = rec["priority"]
                    ws2[f"D{row}"] = rec["category"]

            wb.save(file_path)
            return file_path

        except ImportError:
            logger.warning("openpyxl not installed, using CSV format")
            csv_data = await self._export_csv(report_data)
            csv_path = file_path.with_suffix(".csv")
            csv_path.write_text(csv_data)
            return csv_path

    async def _write_file(
        self,
        report_data: ReportData,
        extension: str,
        content: str,
    ) -> Path:
        """Write report content to file."""
        output_path = self.output_path / report_data.config.name
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_path = output_path / f"{report_data.config.name}_{timestamp}.{extension}"

        async with aiofiles.open(file_path, "w") as f:
            await f.write(content)

        logger.info(f"Report written to {file_path}")
        return file_path

    async def get_report_configs(self) -> List[Dict[str, Any]]:
        """
        Get all report configurations.

        Returns:
            List of report configs
        """
        return [c.to_dict() for c in self._configs.values()]

    async def get_report_config(self, name: str) -> Optional[Dict[str, Any]]:
        """
        Get a specific report configuration.

        Args:
            name: Report config name

        Returns:
            Report config or None
        """
        config = self._configs.get(name)
        return config.to_dict() if config else None

    async def add_report_config(self, config: ReportConfig) -> bool:
        """
        Add a report configuration.

        Args:
            config: Report configuration

        Returns:
            True if added
        """
        async with self._lock:
            if config.name in self._configs:
                return False

            self._configs[config.name] = config
            await self._save_configs()
            logger.info(f"Added report config: {config.name}")
            return True

    async def update_report_config(self, name: str, updates: Dict[str, Any]) -> bool:
        """
        Update a report configuration.

        Args:
            name: Report config name
            updates: Updates to apply

        Returns:
            True if updated
        """
        async with self._lock:
            if name not in self._configs:
                return False

            config = self._configs[name]

            for key, value in updates.items():
                if key == "type":
                    value = ReportType(value)
                elif key == "formats":
                    value = [ReportFormat(f) for f in value]
                elif key == "time_range":
                    value = (
                        datetime.fromisoformat(value["start"]) if value.get("start") else None,
                        datetime.fromisoformat(value["end"]) if value.get("end") else None,
                    )
                setattr(config, key, value)

            await self._save_configs()
            logger.info(f"Updated report config: {name}")
            return True

    async def _save_configs(self):
        """Save report configurations."""
        try:
            data = {
                "reports": [c.to_dict() for c in self._configs.values()]
            }

            with open(self.configs_path, "w") as f:
                yaml.dump(data, f, default_flow_style=False)

        except Exception as e:
            logger.error(f"Error saving report configs: {e}")

    async def shutdown(self):
        """Shutdown the report generator."""
        # Cancel scheduled tasks
        for task in self._scheduled_tasks.values():
            task.cancel()

        logger.info("PerformanceReportGenerator shut down")


# Export singleton
performance_report_generator = PerformanceReportGenerator()
