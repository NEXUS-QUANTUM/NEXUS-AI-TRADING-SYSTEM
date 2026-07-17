"""
NEXUS AI TRADING SYSTEM - Execution Report for AI Trading Bot
Copyright © 2026 NEXUS QUANTUM LTD - All Rights Reserved
CEO: Dr X... - Majority Shareholder

Module: trading/bots/ai_bot/execution/execution_report.py
Description: Générateur de rapports d'exécution pour le bot AI.
             Supporte les rapports détaillés, les résumés, les analyses
             de performance d'exécution, les statistiques de latence,
             les taux de remplissage et les métriques de qualité d'exécution.
             Exporte en HTML, PDF, JSON, Excel et Markdown.
"""

import asyncio
import logging
import time
import json
import os
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

import numpy as np
import pandas as pd

from trading.bots.ai_bot.execution.execution_monitor import ExecutionMetrics, ExecutionAlert
from trading.bots.ai_bot.execution.order_executor import OrderStatus
from shared.exceptions import ReportError

# Configuration du logging
logger = logging.getLogger(__name__)


class ReportFormat(Enum):
    """Formats de rapport."""
    HTML = "html"
    PDF = "pdf"
    JSON = "json"
    EXCEL = "excel"
    MARKDOWN = "markdown"
    CSV = "csv"


class ReportType(Enum):
    """Types de rapport."""
    SUMMARY = "summary"
    DETAILED = "detailed"
    PERFORMANCE = "performance"
    LATENCY = "latency"
    FILL_RATE = "fill_rate"
    ERRORS = "errors"
    ALERTS = "alerts"
    CUSTOM = "custom"


@dataclass
class ExecutionReportConfig:
    """
    Configuration du rapport d'exécution.
    """
    # Format et type
    format: ReportFormat = ReportFormat.HTML
    type: ReportType = ReportType.SUMMARY
    
    # Sections
    include_summary: bool = True
    include_metrics: bool = True
    include_orders: bool = True
    include_latency: bool = True
    include_fill_rates: bool = True
    include_errors: bool = True
    include_alerts: bool = True
    include_charts: bool = True
    
    # Paramètres de sortie
    output_dir: str = "data/execution_reports/"
    filename_prefix: str = "execution_report"
    overwrite: bool = True
    timestamp: bool = True
    
    # Paramètres de visualisation
    chart_style: str = "dark"  # 'dark' ou 'light'
    chart_height: int = 400
    chart_width: int = 800
    
    # Paramètres de performance
    parallel: bool = True
    n_workers: int = 4
    
    def __post_init__(self):
        """Validation des paramètres."""
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)


@dataclass
class ExecutionReportData:
    """
    Données du rapport d'exécution.
    """
    # Métadonnées
    report_id: str
    generated_at: datetime
    period_start: Optional[datetime] = None
    period_end: Optional[datetime] = None
    
    # Métriques
    metrics: Optional[ExecutionMetrics] = None
    metrics_history: Optional[List[ExecutionMetrics]] = None
    
    # Ordres
    orders: List[Dict[str, Any]] = field(default_factory=list)
    order_summary: Dict[str, Any] = field(default_factory=dict)
    
    # Latence
    latency_stats: Dict[str, float] = field(default_factory=dict)
    latency_history: List[float] = field(default_factory=list)
    
    # Taux de remplissage
    fill_rate_stats: Dict[str, float] = field(default_factory=dict)
    fill_rate_history: List[float] = field(default_factory=list)
    
    # Erreurs
    errors: List[Dict[str, Any]] = field(default_factory=list)
    error_summary: Dict[str, int] = field(default_factory=dict)
    
    # Alertes
    alerts: List[ExecutionAlert] = field(default_factory=list)
    alert_summary: Dict[str, int] = field(default_factory=dict)
    
    # Performances
    performance_metrics: Dict[str, float] = field(default_factory=dict)
    comparison_data: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convertit en dictionnaire."""
        return {
            'report_id': self.report_id,
            'generated_at': self.generated_at.isoformat(),
            'period_start': self.period_start.isoformat() if self.period_start else None,
            'period_end': self.period_end.isoformat() if self.period_end else None,
            'metrics': self.metrics.to_dict() if self.metrics else None,
            'order_summary': self.order_summary,
            'latency_stats': self.latency_stats,
            'fill_rate_stats': self.fill_rate_stats,
            'error_summary': self.error_summary,
            'alert_summary': self.alert_summary,
            'performance_metrics': self.performance_metrics
        }


class ExecutionReportGenerator:
    """
    Générateur de rapports d'exécution.
    """
    
    def __init__(self, config: Optional[ExecutionReportConfig] = None):
        """
        Initialise le générateur de rapports.
        
        Args:
            config: Configuration du rapport.
        """
        self.config = config or ExecutionReportConfig()
        
        # Templates
        self._template_env = None
        self._setup_template_engine()
        
        # Cache
        self._report_cache: Dict[str, ExecutionReportData] = {}
        
        logger.info("ExecutionReportGenerator initialisé")
        logger.info(f"Format: {self.config.format.value}")
        logger.info(f"Type: {self.config.type.value}")
    
    def _setup_template_engine(self) -> None:
        """
        Configure le moteur de templates.
        """
        try:
            from jinja2 import Environment, FileSystemLoader, select_autoescape
            
            template_dirs = [
                'templates/execution/',
                'configs/templates/execution/',
                os.path.join(os.path.dirname(__file__), 'templates')
            ]
            
            template_path = None
            for path in template_dirs:
                if os.path.exists(path):
                    template_path = path
                    break
            
            if template_path is None:
                template_path = os.path.dirname(__file__)
            
            self._template_env = Environment(
                loader=FileSystemLoader(template_path),
                autoescape=select_autoescape(['html', 'xml']),
                trim_blocks=True,
                lstrip_blocks=True
            )
            
            # Filtres personnalisés
            self._template_env.filters['format_currency'] = lambda x: f"${x:,.2f}" if x else "$0.00"
            self._template_env.filters['format_percent'] = lambda x: f"{x:.2%}" if x else "0%"
            self._template_env.filters['format_number'] = lambda x: f"{x:,.0f}" if x else "0"
            self._template_env.filters['format_duration'] = self._format_duration
            self._template_env.filters['format_datetime'] = lambda x: x.isoformat() if x else ""
            
        except ImportError:
            logger.warning("Jinja2 non disponible, utiliser HTML basique")
            self._template_env = None
    
    def _format_duration(self, seconds: float) -> str:
        """
        Formate une durée.
        
        Args:
            seconds: Durée en secondes.
            
        Returns:
            Durée formatée.
        """
        if seconds < 60:
            return f"{seconds:.1f}s"
        elif seconds < 3600:
            return f"{seconds/60:.1f}m"
        elif seconds < 86400:
            return f"{seconds/3600:.1f}h"
        else:
            return f"{seconds/86400:.1f}d"
    
    # ============================================================
    # GÉNÉRATION DE RAPPORTS
    # ============================================================
    
    def generate_report(
        self,
        data: ExecutionReportData,
        filename: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Génère un rapport d'exécution.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier (optionnel).
            
        Returns:
            Dictionnaire des fichiers générés.
        """
        logger.info(f"Génération du rapport: {data.report_id}")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.config.filename_prefix}_{timestamp}"
        
        outputs = {}
        
        # Génération selon le format
        if self.config.format == ReportFormat.HTML:
            outputs['html'] = self._generate_html(data, filename)
        elif self.config.format == ReportFormat.PDF:
            outputs['pdf'] = self._generate_pdf(data, filename)
        elif self.config.format == ReportFormat.JSON:
            outputs['json'] = self._generate_json(data, filename)
        elif self.config.format == ReportFormat.EXCEL:
            outputs['excel'] = self._generate_excel(data, filename)
        elif self.config.format == ReportFormat.MARKDOWN:
            outputs['markdown'] = self._generate_markdown(data, filename)
        elif self.config.format == ReportFormat.CSV:
            outputs['csv'] = self._generate_csv(data, filename)
        else:
            raise ReportError(f"Format non supporté: {self.config.format}")
        
        logger.info(f"Rapport généré: {filename}")
        return outputs
    
    # ============================================================
    # HTML
    # ============================================================
    
    def _generate_html(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport HTML.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        filepath = Path(self.config.output_dir) / f"{filename}.html"
        
        try:
            if self._template_env:
                template = self._template_env.get_template('execution_report.html')
                html_content = template.render(
                    report=data.to_dict(),
                    config=self.config,
                    charts=self._generate_charts(data)
                )
            else:
                html_content = self._generate_simple_html(data)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur de génération HTML: {e}")
            raise ReportError(f"Erreur HTML: {e}")
    
    def _generate_simple_html(self, data: ExecutionReportData) -> str:
        """
        Génère un HTML simple sans template.
        
        Args:
            data: Données du rapport.
            
        Returns:
            HTML string.
        """
        html = []
        html.append("<!DOCTYPE html>")
        html.append("<html>")
        html.append("<head>")
        html.append("<meta charset='UTF-8'>")
        html.append(f"<title>Execution Report - {data.report_id}</title>")
        html.append("<style>")
        html.append("body { font-family: Arial, sans-serif; margin: 20px; background: #1a1a2e; color: #e0e0e0; }")
        html.append("h1, h2, h3 { color: #00ff9d; }")
        html.append("table { border-collapse: collapse; width: 100%; margin: 10px 0; }")
        html.append("th, td { border: 1px solid #333; padding: 8px; text-align: left; }")
        html.append("th { background: #2a2a4e; color: #00ff9d; }")
        html.append(".metric-card { display: inline-block; background: #2a2a4e; padding: 15px; margin: 5px; border-radius: 8px; min-width: 150px; }")
        html.append(".metric-value { font-size: 24px; font-weight: bold; color: #00ff9d; }")
        html.append(".metric-label { font-size: 12px; color: #888; }")
        html.append(".positive { color: #00ff9d; }")
        html.append(".negative { color: #ff6b6b; }")
        html.append(".warning { color: #ffd93d; }")
        html.append("</style>")
        html.append("</head>")
        html.append("<body>")
        
        # Header
        html.append(f"<h1>📊 Execution Report</h1>")
        html.append(f"<p><strong>Report ID:</strong> {data.report_id}</p>")
        html.append(f"<p><strong>Generated:</strong> {data.generated_at.isoformat()}</p>")
        
        # Metrics
        if data.metrics:
            html.append("<h2>📈 Performance Metrics</h2>")
            html.append("<div>")
            
            metrics = data.metrics.to_dict()
            metric_items = [
                ('Total Orders', metrics.get('total_orders', 0)),
                ('Fill Rate', f"{metrics.get('fill_rate', 0):.2%}"),
                ('Avg Latency', f"{metrics.get('avg_latency_ms', 0):.2f}ms"),
                ('Throughput', f"{metrics.get('throughput', 0):.2f}/s"),
                ('Slippage', f"{metrics.get('slippage', 0):.4f}"),
                ('Error Rate', f"{metrics.get('error_rate', 0):.2%}")
            ]
            
            for label, value in metric_items:
                html.append(f"""
                <div class="metric-card">
                    <div class="metric-label">{label}</div>
                    <div class="metric-value">{value}</div>
                </div>
                """)
            
            html.append("</div>")
        
        # Order Summary
        if data.order_summary:
            html.append("<h2>📋 Order Summary</h2>")
            html.append("<table>")
            html.append("<tr><th>Status</th><th>Count</th></tr>")
            
            for status, count in data.order_summary.items():
                html.append(f"<tr><td>{status}</td><td>{count}</td></tr>")
            
            html.append("</table>")
        
        # Latency Stats
        if data.latency_stats:
            html.append("<h2>⏱️ Latency Statistics</h2>")
            html.append("<table>")
            html.append("<tr><th>Metric</th><th>Value</th></tr>")
            
            for key, value in data.latency_stats.items():
                html.append(f"<tr><td>{key}</td><td>{value:.2f}ms</td></tr>")
            
            html.append("</table>")
        
        # Errors
        if data.errors:
            html.append("<h2>❌ Errors</h2>")
            html.append("<table>")
            html.append("<tr><th>Timestamp</th><th>Error</th></tr>")
            
            for error in data.errors[:20]:
                html.append(f"<tr><td>{error.get('timestamp', '')}</td><td>{error.get('error', '')}</td></tr>")
            
            html.append("</table>")
        
        # Alerts
        if data.alerts:
            html.append("<h2>⚠️ Alerts</h2>")
            html.append("<table>")
            html.append("<tr><th>Severity</th><th>Message</th><th>Timestamp</th></tr>")
            
            for alert in data.alerts[:20]:
                html.append(f"""
                <tr>
                    <td>{alert.severity.value}</td>
                    <td>{alert.message}</td>
                    <td>{alert.timestamp.isoformat()}</td>
                </tr>
                """)
            
            html.append("</table>")
        
        # Footer
        html.append("<hr>")
        html.append(f"<p style='color: #888; font-size: 12px;'>Generated by NEXUS AI Trading System © 2026</p>")
        html.append("</body>")
        html.append("</html>")
        
        return "\n".join(html)
    
    # ============================================================
    # PDF
    # ============================================================
    
    def _generate_pdf(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport PDF.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        try:
            import weasyprint
            
            # D'abord générer le HTML
            html_file = self._generate_html(data, filename)
            pdf_file = Path(self.config.output_dir) / f"{filename}.pdf"
            
            weasyprint.HTML(filename=html_file).write_pdf(
                pdf_file,
                stylesheets=[self._get_pdf_styles()]
            )
            
            return str(pdf_file)
            
        except ImportError:
            logger.warning("WeasyPrint non disponible, utilisation du HTML")
            return self._generate_html(data, filename)
        except Exception as e:
            logger.error(f"Erreur de génération PDF: {e}")
            raise ReportError(f"Erreur PDF: {e}")
    
    def _get_pdf_styles(self) -> str:
        """
        Styles CSS pour le PDF.
        
        Returns:
            CSS string.
        """
        return """
        @page {
            size: A4;
            margin: 2cm;
        }
        body {
            font-family: Arial, sans-serif;
            font-size: 10pt;
            color: #333;
            background: #fff;
        }
        h1 { color: #2c3e50; font-size: 18pt; }
        h2 { color: #2c3e50; font-size: 14pt; margin-top: 20px; }
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 10px 0;
            font-size: 8pt;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 4px 8px;
            text-align: left;
        }
        th {
            background-color: #2c3e50;
            color: white;
        }
        .metric-card {
            display: inline-block;
            background: #f8f9fa;
            padding: 10px;
            margin: 5px;
            border-radius: 5px;
            min-width: 120px;
        }
        .metric-value {
            font-size: 16px;
            font-weight: bold;
        }
        .positive { color: #27ae60; }
        .negative { color: #e74c3c; }
        .warning { color: #f39c12; }
        """
    
    # ============================================================
    # JSON
    # ============================================================
    
    def _generate_json(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport JSON.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        filepath = Path(self.config.output_dir) / f"{filename}.json"
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data.to_dict(), f, indent=2, default=str)
        
        return str(filepath)
    
    # ============================================================
    # EXCEL
    # ============================================================
    
    def _generate_excel(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport Excel.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.warning("OpenPyXL non disponible, utilisation du CSV")
            return self._generate_csv(data, filename)
        
        filepath = Path(self.config.output_dir) / f"{filename}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 1. Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        
        summary_data = [
            ['Metric', 'Value'],
            ['Report ID', data.report_id],
            ['Generated', data.generated_at.isoformat()],
            ['Period Start', data.period_start.isoformat() if data.period_start else ''],
            ['Period End', data.period_end.isoformat() if data.period_end else ''],
            ['Total Orders', data.metrics.total_orders if data.metrics else 0],
            ['Fill Rate', f"{data.metrics.fill_rate:.2%}" if data.metrics else '0%'],
            ['Avg Latency', f"{data.metrics.avg_latency_ms:.2f}ms" if data.metrics else '0ms']
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        # 2. Metrics Sheet
        if data.metrics:
            ws_metrics = wb.create_sheet("Metrics")
            metrics_data = [['Metric', 'Value']]
            
            for key, value in data.metrics.to_dict().items():
                if isinstance(value, (int, float)):
                    if 'rate' in key or 'latency' in key:
                        metrics_data.append([key, f"{value:.4f}"])
                    else:
                        metrics_data.append([key, value])
                else:
                    metrics_data.append([key, value])
            
            for row_idx, row in enumerate(metrics_data, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_metrics.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            ws_metrics.column_dimensions['A'].width = 25
            ws_metrics.column_dimensions['B'].width = 20
        
        # 3. Orders Sheet
        if data.orders:
            ws_orders = wb.create_sheet("Orders")
            
            order_data = []
            for order in data.orders[:1000]:
                order_data.append({
                    'ID': order.get('id', ''),
                    'Symbol': order.get('symbol', ''),
                    'Side': order.get('side', ''),
                    'Quantity': order.get('quantity', 0),
                    'Price': order.get('price', 0),
                    'Status': order.get('status', ''),
                    'Timestamp': order.get('timestamp', '')
                })
            
            if order_data:
                df = pd.DataFrame(order_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws_orders.append(r)
                
                for col in range(1, len(df.columns) + 1):
                    cell = ws_orders.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
        
        # 4. Alerts Sheet
        if data.alerts:
            ws_alerts = wb.create_sheet("Alerts")
            
            alert_data = []
            for alert in data.alerts[:1000]:
                alert_data.append({
                    'ID': alert.id,
                    'Severity': alert.severity.value,
                    'Message': alert.message,
                    'Timestamp': alert.timestamp.isoformat(),
                    'Order ID': alert.order_id or '',
                    'Acknowledged': alert.acknowledged
                })
            
            if alert_data:
                df = pd.DataFrame(alert_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws_alerts.append(r)
                
                for col in range(1, len(df.columns) + 1):
                    cell = ws_alerts.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Ajustement des colonnes
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_length = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_length
        
        wb.save(filepath)
        
        return str(filepath)
    
    # ============================================================
    # MARKDOWN
    # ============================================================
    
    def _generate_markdown(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport Markdown.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        filepath = Path(self.config.output_dir) / f"{filename}.md"
        
        lines = []
        lines.append(f"# 📊 Execution Report")
        lines.append("")
        lines.append(f"**Report ID:** {data.report_id}")
        lines.append(f"**Generated:** {data.generated_at.isoformat()}")
        lines.append("")
        
        if data.metrics:
            lines.append("## 📈 Performance Metrics")
            lines.append("")
            lines.append("| Metric | Value |")
            lines.append("|--------|-------|")
            
            metrics = data.metrics.to_dict()
            for key, value in metrics.items():
                if isinstance(value, float):
                    if 'rate' in key or 'latency' in key:
                        lines.append(f"| {key} | {value:.4f} |")
                    else:
                        lines.append(f"| {key} | {value:.2f} |")
                else:
                    lines.append(f"| {key} | {value} |")
            lines.append("")
        
        if data.order_summary:
            lines.append("## 📋 Order Summary")
            lines.append("")
            lines.append("| Status | Count |")
            lines.append("|--------|-------|")
            for status, count in data.order_summary.items():
                lines.append(f"| {status} | {count} |")
            lines.append("")
        
        if data.latency_stats:
            lines.append("## ⏱️ Latency Statistics")
            lines.append("")
            lines.append("| Metric | Value |")
            lines.append("|--------|-------|")
            for key, value in data.latency_stats.items():
                lines.append(f"| {key} | {value:.2f}ms |")
            lines.append("")
        
        if data.errors:
            lines.append("## ❌ Errors")
            lines.append("")
            for error in data.errors[:10]:
                lines.append(f"- {error.get('timestamp', '')}: {error.get('error', '')}")
            if len(data.errors) > 10:
                lines.append(f"- ... and {len(data.errors) - 10} more")
            lines.append("")
        
        if data.alerts:
            lines.append("## ⚠️ Alerts")
            lines.append("")
            for alert in data.alerts[:10]:
                lines.append(f"- [{alert.severity.value.upper()}] {alert.message} ({alert.timestamp.isoformat()})")
            if len(data.alerts) > 10:
                lines.append(f"- ... and {len(data.alerts) - 10} more")
            lines.append("")
        
        lines.append("---")
        lines.append("*Generated by NEXUS AI Trading System © 2026*")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return str(filepath)
    
    # ============================================================
    # CSV
    # ============================================================
    
    def _generate_csv(self, data: ExecutionReportData, filename: str) -> str:
        """
        Génère un rapport CSV.
        
        Args:
            data: Données du rapport.
            filename: Nom du fichier.
            
        Returns:
            Chemin du fichier.
        """
        filepath = Path(self.config.output_dir) / f"{filename}.csv"
        
        # Conversion en DataFrame
        rows = []
        
        if data.metrics:
            for key, value in data.metrics.to_dict().items():
                rows.append({
                    'category': 'metrics',
                    'key': key,
                    'value': value
                })
        
        if data.order_summary:
            for status, count in data.order_summary.items():
                rows.append({
                    'category': 'order_summary',
                    'key': status,
                    'value': count
                })
        
        if data.latency_stats:
            for key, value in data.latency_stats.items():
                rows.append({
                    'category': 'latency',
                    'key': key,
                    'value': value
                })
        
        df = pd.DataFrame(rows)
        df.to_csv(filepath, index=False)
        
        return str(filepath)
    
    # ============================================================
    # GRAPHIQUES
    # ============================================================
    
    def _generate_charts(self, data: ExecutionReportData) -> Dict[str, str]:
        """
        Génère des graphiques pour le rapport.
        
        Args:
            data: Données du rapport.
            
        Returns:
            Dictionnaire des graphiques.
        """
        charts = {}
        
        try:
            import matplotlib.pyplot as plt
            import seaborn as sns
            
            # Style
            if self.config.chart_style == 'dark':
                plt.style.use('dark_background')
                colors = ['#00ff9d', '#ff6b6b', '#ffd93d', '#6bcbff']
            else:
                plt.style.use('seaborn-v0_8-whitegrid')
                colors = ['#2ecc71', '#e74c3c', '#f1c40f', '#3498db']
            
            sns.set_palette(colors)
            
            chart_dir = Path(self.config.output_dir) / 'charts'
            chart_dir.mkdir(exist_ok=True)
            
            # 1. Latency Chart
            if data.latency_history:
                fig, ax = plt.subplots(figsize=(12, 6))
                ax.plot(data.latency_history, color='#6bcbff', linewidth=2)
                ax.fill_between(range(len(data.latency_history)), data.latency_history, alpha=0.3)
                ax.set_xlabel('Sample')
                ax.set_ylabel('Latency (ms)')
                ax.set_title('Latency Over Time')
                ax.grid(True, alpha=0.3)
                
                filepath = chart_dir / 'latency.png'
                plt.savefig(filepath, dpi=150, bbox_inches='tight')
                plt.close()
                charts['latency'] = str(filepath)
            
            # 2. Fill Rate Chart
            if data.fill_rate_history:
                fig, ax = plt.subplots(figsize=(12, 6))
                ax.plot(data.fill_rate_history, color='#00ff9d', linewidth=2)
                ax.axhline(y=0.8, color='#ff6b6b', linestyle='--', alpha=0.5, label='Target')
                ax.set_xlabel('Sample')
                ax.set_ylabel('Fill Rate')
                ax.set_title('Fill Rate Over Time')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
                filepath = chart_dir / 'fill_rate.png'
                plt.savefig(filepath, dpi=150, bbox_inches='tight')
                plt.close()
                charts['fill_rate'] = str(filepath)
            
            # 3. Order Distribution
            if data.order_summary:
                fig, ax = plt.subplots(figsize=(10, 8))
                statuses = list(data.order_summary.keys())
                counts = list(data.order_summary.values())
                colors_status = ['#00ff9d' if s == 'filled' else '#ff6b6b' if s == 'rejected' else '#ffd93d' for s in statuses]
                ax.pie(counts, labels=statuses, autopct='%1.1f%%', colors=colors_status, startangle=90)
                ax.set_title('Order Distribution')
                
                filepath = chart_dir / 'order_distribution.png'
                plt.savefig(filepath, dpi=150, bbox_inches='tight')
                plt.close()
                charts['order_distribution'] = str(filepath)
            
            # 4. Metrics Radar Chart
            if data.metrics:
                metrics = data.metrics.to_dict()
                fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(projection='polar'))
                
                categories = ['fill_rate', 'fill_rate', 'throughput', 'avg_latency_ms', 'slippage']
                values = [metrics.get(c, 0) for c in categories]
                values = np.clip(values, 0, 1)
                
                angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
                values += values[:1]
                angles += angles[:1]
                
                ax.plot(angles, values, 'o-', linewidth=2)
                ax.fill(angles, values, alpha=0.25)
                ax.set_xticks(angles[:-1])
                ax.set_xticklabels(categories)
                ax.set_title('Performance Radar')
                
                filepath = chart_dir / 'radar.png'
                plt.savefig(filepath, dpi=150, bbox_inches='tight')
                plt.close()
                charts['radar'] = str(filepath)
            
        except ImportError as e:
            logger.warning(f"Bibliothèque de graphiques non disponible: {e}")
        
        return charts


# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def create_execution_report(
    data: ExecutionReportData,
    format: str = "html",
    output_dir: str = "data/execution_reports/",
    **kwargs
) -> Dict[str, str]:
    """
    Crée un rapport d'exécution.
    
    Args:
        data: Données du rapport.
        format: Format du rapport.
        output_dir: Répertoire de sortie.
        **kwargs: Paramètres supplémentaires.
        
    Returns:
        Dictionnaire des fichiers générés.
    """
    format_map = {
        'html': ReportFormat.HTML,
        'pdf': ReportFormat.PDF,
        'json': ReportFormat.JSON,
        'excel': ReportFormat.EXCEL,
        'markdown': ReportFormat.MARKDOWN,
        'csv': ReportFormat.CSV
    }
    
    config = ExecutionReportConfig(
        format=format_map.get(format, ReportFormat.HTML),
        output_dir=output_dir,
        **kwargs
    )
    
    generator = ExecutionReportGenerator(config)
    return generator.generate_report(data)


# ============================================================
# EXPORTATION
# ============================================================

__all__ = [
    'ExecutionReportGenerator',
    'ExecutionReportConfig',
    'ExecutionReportData',
    'ReportFormat',
    'ReportType',
    'create_execution_report'
]

# ============================================================
# FIN DU FICHIER
# ============================================================
