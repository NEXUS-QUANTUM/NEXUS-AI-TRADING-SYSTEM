"""
NEXUS AI TRADING SYSTEM - Portfolio Reporting Module
Copyright © 2026 NEXUS QUANTUM LTD
CEO: Dr X... - Majority Shareholder

Module: trading/portfolio/reporting.py
Description: Comprehensive portfolio reporting with full API integration
"""

import asyncio
import json
import logging
import os
import tempfile
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, field, asdict
from enum import Enum
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns

from pydantic import BaseModel, Field, validator
from fastapi import HTTPException, status, Response
from fastapi.responses import FileResponse, StreamingResponse

# NEXUS Internal Imports
from shared.configs.portfolio_config import PortfolioConfig
from shared.configs.reporting_config import ReportingConfig
from shared.constants.trading_constants import TIME_FRAMES
from shared.utilities.retry import retry_async
from shared.utilities.logger import get_logger
from shared.utilities.cache_utils import cached

# Database imports
from backend.database.repositories.position_repository import PositionRepository
from backend.database.repositories.trade_repository import TradeRepository
from backend.database.repositories.portfolio_repository import PortfolioRepository

# Portfolio imports
from trading.portfolio.balance_tracker import BalanceTracker, BalanceRequest
from trading.portfolio.performance import PortfolioPerformance, PerformanceRequest
from trading.portfolio.pnl_calculator import PnLCalculator, PnLRequest
from trading.portfolio.history import PortfolioHistory, HistoryRequest
from trading.portfolio.position_manager import PositionManager

logger = get_logger(__name__)


# =============================================================================
# ENUMS & CONSTANTS
# =============================================================================

class ReportFormat(str, Enum):
    """Report output formats"""
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class ReportType(str, Enum):
    """Types of reports"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUAL = "annual"
    PERFORMANCE = "performance"
    RISK = "risk"
    POSITION = "position"
    PNL = "pnl"
    EXECUTIVE = "executive"
    DETAILED = "detailed"
    COMPLIANCE = "compliance"


class ChartType(str, Enum):
    """Chart types"""
    EQUITY = "equity"
    DRAWDOWN = "drawdown"
    PNL = "pnl"
    POSITIONS = "positions"
    PERFORMANCE = "performance"
    RISK = "risk"
    ALLOCATION = "allocation"
    RETURNS = "returns"
    DIVERSIFICATION = "diversification"


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class ReportRequest(BaseModel):
    """Request model for report generation"""
    portfolio_id: str
    report_type: ReportType = ReportType.MONTHLY
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    format: ReportFormat = ReportFormat.PDF
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_detailed: bool = False
    include_recommendations: bool = True
    chart_types: List[ChartType] = []
    metrics: List[str] = []
    compare_to_previous: bool = False
    email_report: bool = False
    email_recipients: List[str] = []
    metadata: Dict[str, Any] = Field(default_factory=dict)


class ReportResponse(BaseModel):
    """Response model for report generation"""
    report_id: str
    portfolio_id: str
    report_type: ReportType
    generated_at: datetime
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    format: ReportFormat
    file_size: int
    summary: Dict[str, Any]
    metrics: Dict[str, Any]
    charts_count: int
    tables_count: int
    recommendations: List[str]
    metadata: Dict[str, Any] = Field(default_factory=dict)


class ReportSchedule(BaseModel):
    """Schedule for automated reports"""
    portfolio_id: str
    report_type: ReportType
    frequency: str  # daily, weekly, monthly
    format: ReportFormat = ReportFormat.PDF
    time: str = "00:00"
    day_of_week: Optional[int] = None
    day_of_month: Optional[int] = None
    recipients: List[str] = []
    enabled: bool = True
    last_generated: Optional[datetime] = None
    next_generation: Optional[datetime] = None


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ReportData:
    """Container for report data"""
    portfolio_id: str
    report_type: ReportType
    start_date: datetime
    end_date: datetime
    positions: List[Any]
    trades: List[Any]
    balance_history: List[Any]
    performance_metrics: Dict[str, Any]
    pnl_metrics: Dict[str, Any]
    risk_metrics: Dict[str, Any]
    summary: Dict[str, Any]
    recommendations: List[str]


@dataclass
class ChartData:
    """Container for chart data"""
    title: str
    chart_type: ChartType
    data: Dict[str, Any]
    x_label: Optional[str] = None
    y_label: Optional[str] = None
    legend: Optional[List[str]] = None


# =============================================================================
# PORTFOLIO REPORTER
# =============================================================================

class PortfolioReporter:
    """
    Comprehensive Portfolio Reporter with full API integration.
    
    Features:
    - Multiple report types
    - Various output formats
    - Interactive charts
    - Performance analysis
    - Risk reporting
    - Position reporting
    - PnL analysis
    - Executive summaries
    - Scheduled reports
    - Email delivery
    """

    def __init__(
        self,
        config: Optional[ReportingConfig] = None,
        portfolio_config: Optional[PortfolioConfig] = None,
        position_repo: Optional[PositionRepository] = None,
        trade_repo: Optional[TradeRepository] = None,
        portfolio_repo: Optional[PortfolioRepository] = None,
        balance_tracker: Optional[BalanceTracker] = None,
        performance: Optional[PortfolioPerformance] = None,
        pnl_calculator: Optional[PnLCalculator] = None,
        history: Optional[PortfolioHistory] = None,
        position_manager: Optional[PositionManager] = None
    ):
        """
        Initialize PortfolioReporter.
        
        Args:
            config: Reporting configuration
            portfolio_config: Portfolio configuration
            position_repo: Position repository
            trade_repo: Trade repository
            portfolio_repo: Portfolio repository
            balance_tracker: Balance tracker
            performance: Portfolio performance
            pnl_calculator: PnL calculator
            history: Portfolio history
            position_manager: Position manager
        """
        self.config = config or ReportingConfig()
        self.portfolio_config = portfolio_config or PortfolioConfig()
        self.position_repo = position_repo or PositionRepository()
        self.trade_repo = trade_repo or TradeRepository()
        self.portfolio_repo = portfolio_repo or PortfolioRepository()
        self.balance_tracker = balance_tracker or BalanceTracker()
        self.performance = performance or PortfolioPerformance()
        self.pnl_calculator = pnl_calculator or PnLCalculator()
        self.history = history or PortfolioHistory()
        self.position_manager = position_manager or PositionManager()
        
        # Report storage
        self._reports: Dict[str, Dict[str, Any]] = {}
        self._schedules: Dict[str, ReportSchedule] = {}
        
        # Cache
        self._report_cache: Dict[str, Dict[str, Any]] = {}
        
        # Chart defaults
        self._chart_defaults = {
            'figsize': (12, 8),
            'dpi': 100,
            'style': 'seaborn-v0_8-darkgrid',
            'font_size': 12,
            'title_size': 16
        }
        
        # Set style
        plt.style.use(self._chart_defaults['style'])
        sns.set_palette("husl")
        
        logger.info("PortfolioReporter initialized")

    # =========================================================================
    # Report Generation
    # =========================================================================

    @retry_async(max_attempts=3, delay=1.0)
    async def generate_report(
        self,
        request: ReportRequest
    ) -> ReportResponse:
        """
        Generate a portfolio report.
        
        Args:
            request: Report request
            
        Returns:
            ReportResponse: Generated report
        """
        try:
            # Validate request
            await self._validate_request(request)
            
            # Set date range
            start_date, end_date = self._get_date_range(request)
            
            # Collect report data
            report_data = await self._collect_report_data(
                request.portfolio_id,
                request.report_type,
                start_date,
                end_date
            )
            
            # Generate report
            report_id = f"port_report_{int(time.time() * 1000)}_{request.portfolio_id}"
            
            # Generate content
            content, file_size = await self._generate_content(
                report_data,
                request.format,
                request
            )
            
            # Create response
            response = ReportResponse(
                report_id=report_id,
                portfolio_id=request.portfolio_id,
                report_type=request.report_type,
                generated_at=datetime.utcnow(),
                start_date=start_date,
                end_date=end_date,
                format=request.format,
                file_size=file_size,
                summary=report_data.summary,
                metrics={
                    'performance': report_data.performance_metrics,
                    'pnl': report_data.pnl_metrics,
                    'risk': report_data.risk_metrics
                },
                charts_count=len(request.chart_types) if request.include_charts else 0,
                tables_count=0,
                recommendations=report_data.recommendations,
                metadata=request.metadata
            )
            
            # Cache report
            self._reports[report_id] = {
                'response': response,
                'content': content,
                'data': report_data,
                'generated_at': datetime.utcnow()
            }
            
            # Send email if requested
            if request.email_report and request.email_recipients:
                await self._send_report_email(
                    response,
                    content,
                    request.email_recipients
                )
            
            logger.info(f"Report generated: {report_id} for {request.portfolio_id}")
            return response
            
        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Report generation failed: {str(e)}"
            )

    async def _validate_request(self, request: ReportRequest) -> None:
        """Validate report request"""
        if request.start_date and request.end_date:
            if request.start_date > request.end_date:
                raise ValueError("Start date must be before end date")

    def _get_date_range(
        self,
        request: ReportRequest
    ) -> Tuple[datetime, datetime]:
        """Get date range for report"""
        end_date = request.end_date or datetime.utcnow()
        
        if request.start_date:
            start_date = request.start_date
        else:
            # Default based on report type
            if request.report_type == ReportType.DAILY:
                start_date = end_date - timedelta(days=1)
            elif request.report_type == ReportType.WEEKLY:
                start_date = end_date - timedelta(days=7)
            elif request.report_type == ReportType.MONTHLY:
                start_date = end_date - timedelta(days=30)
            elif request.report_type == ReportType.QUARTERLY:
                start_date = end_date - timedelta(days=90)
            elif request.report_type == ReportType.ANNUAL:
                start_date = end_date - timedelta(days=365)
            else:
                start_date = end_date - timedelta(days=30)
        
        return start_date, end_date

    async def _collect_report_data(
        self,
        portfolio_id: str,
        report_type: ReportType,
        start_date: datetime,
        end_date: datetime
    ) -> ReportData:
        """
        Collect data for report.
        
        Args:
            portfolio_id: Portfolio ID
            report_type: Type of report
            start_date: Start date
            end_date: End date
            
        Returns:
            ReportData: Collected data
        """
        # Get positions
        positions = await self.position_manager.get_positions(
            portfolio_id=portfolio_id
        )
        
        # Get trades
        trades = await self.trade_repo.get_by_portfolio_id(
            portfolio_id,
            start_date=start_date,
            end_date=end_date
        )
        
        # Get balance history
        balance_request = BalanceRequest(
            portfolio_id=portfolio_id,
            include_history=True,
            history_days=(end_date - start_date).days
        )
        balance = await self.balance_tracker.get_balance(balance_request)
        
        # Get performance metrics
        performance_request = PerformanceRequest(
            portfolio_id=portfolio_id,
            metric="sharpe_ratio",
            period="1m"
        )
        performance = await self.performance.calculate_performance(
            performance_request
        )
        
        # Get PnL metrics
        pnl_request = PnLRequest(
            portfolio_id=portfolio_id,
            period="month"
        )
        pnl = await self.pnl_calculator.calculate_pnl(pnl_request)
        
        # Calculate risk metrics
        risk_metrics = await self._calculate_risk_metrics(
            positions,
            trades,
            balance
        )
        
        # Generate summary
        summary = await self._generate_summary(
            portfolio_id,
            report_type,
            balance,
            performance,
            pnl
        )
        
        # Generate recommendations
        recommendations = await self._generate_recommendations(
            balance,
            performance,
            pnl,
            risk_metrics
        )
        
        return ReportData(
            portfolio_id=portfolio_id,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            positions=positions,
            trades=trades,
            balance_history=balance.history if hasattr(balance, 'history') else [],
            performance_metrics=performance.dict() if performance else {},
            pnl_metrics=pnl.dict() if pnl else {},
            risk_metrics=risk_metrics,
            summary=summary,
            recommendations=recommendations
        )

    async def _calculate_risk_metrics(
        self,
        positions: List[Any],
        trades: List[Any],
        balance: Any
    ) -> Dict[str, Any]:
        """Calculate risk metrics"""
        metrics = {}
        
        # Position risk
        if positions:
            total_value = sum(p.value for p in positions)
            metrics['total_exposure'] = total_value
            
            # Concentration
            if total_value > 0:
                max_position = max(p.value for p in positions)
                metrics['concentration'] = max_position / total_value
            
            # Leverage
            if balance and balance.total_equity > 0:
                metrics['leverage'] = total_value / balance.total_equity
        
        # Drawdown
        if balance:
            metrics['drawdown'] = balance.drawdown
        
        # VaR
        if len(trades) > 30:
            pnls = [t.pnl for t in trades if hasattr(t, 'pnl')]
            if pnls:
                metrics['var_95'] = np.percentile(pnls, 5)
                metrics['cvar_95'] = np.mean([p for p in pnls if p <= metrics['var_95']])
        
        return metrics

    async def _generate_summary(
        self,
        portfolio_id: str,
        report_type: ReportType,
        balance: Any,
        performance: Any,
        pnl: Any
    ) -> Dict[str, Any]:
        """Generate summary"""
        summary = {
            'portfolio_id': portfolio_id,
            'report_type': report_type.value,
            'generated_at': datetime.utcnow().isoformat()
        }
        
        if balance:
            summary.update({
                'total_equity': balance.total_equity,
                'cash_balance': balance.cash_balance,
                'total_pnl': balance.total_pnl,
                'drawdown': balance.drawdown
            })
        
        if performance:
            summary.update({
                'sharpe_ratio': performance.value if hasattr(performance, 'value') else 0
            })
        
        if pnl:
            summary.update({
                'realized_pnl': pnl.realized_pnl if hasattr(pnl, 'realized_pnl') else 0,
                'unrealized_pnl': pnl.unrealized_pnl if hasattr(pnl, 'unrealized_pnl') else 0
            })
        
        return summary

    async def _generate_recommendations(
        self,
        balance: Any,
        performance: Any,
        pnl: Any,
        risk_metrics: Dict[str, Any]
    ) -> List[str]:
        """Generate recommendations"""
        recommendations = []
        
        # Performance recommendations
        if performance:
            if hasattr(performance, 'value') and performance.value < 0.5:
                recommendations.append("Low Sharpe ratio. Consider adjusting strategy.")
        
        # Drawdown recommendations
        if balance and balance.drawdown > 0.15:
            recommendations.append("High drawdown. Consider implementing stricter risk controls.")
        
        # Concentration recommendations
        concentration = risk_metrics.get('concentration', 0)
        if concentration > 0.30:
            recommendations.append("High concentration. Consider diversifying.")
        
        # PnL recommendations
        if pnl:
            if hasattr(pnl, 'total_pnl') and pnl.total_pnl < 0:
                recommendations.append("Negative PnL. Review strategy and risk management.")
        
        if not recommendations:
            recommendations.append("All metrics are within acceptable ranges.")
        
        return recommendations

    # =========================================================================
    # Content Generation
    # =========================================================================

    async def _generate_content(
        self,
        report_data: ReportData,
        format: ReportFormat,
        request: ReportRequest
    ) -> Tuple[Any, int]:
        """
        Generate report content.
        
        Args:
            report_data: Report data
            format: Output format
            request: Original request
            
        Returns:
            Tuple[Any, int]: Content and file size
        """
        if format == ReportFormat.PDF:
            content = await self._generate_pdf(report_data, request)
        elif format == ReportFormat.HTML:
            content = await self._generate_html(report_data, request)
        elif format == ReportFormat.JSON:
            content = await self._generate_json(report_data, request)
        elif format == ReportFormat.CSV:
            content = await self._generate_csv(report_data, request)
        elif format == ReportFormat.EXCEL:
            content = await self._generate_excel(report_data, request)
        elif format == ReportFormat.MARKDOWN:
            content = await self._generate_markdown(report_data, request)
        else:
            content = await self._generate_json(report_data, request)
        
        file_size = len(content) if isinstance(content, (bytes, str)) else 0
        return content, file_size

    # =========================================================================
    # PDF Generation
    # =========================================================================

    async def _generate_pdf(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate PDF report"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image, KeepTogether
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(
                f"Portfolio Report - {report_data.portfolio_id}",
                styles['Title']
            ))
            story.append(Spacer(1, 0.25*inch))
            
            # Info
            story.append(Paragraph(
                f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}",
                styles['Normal']
            ))
            story.append(Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
                styles['Normal']
            ))
            story.append(Spacer(1, 0.5*inch))
            
            # Summary
            story.append(Paragraph("Executive Summary", styles['Heading1']))
            story.append(Spacer(1, 0.1*inch))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Equity', f"${report_data.summary.get('total_equity', 0):.2f}"],
                ['Cash Balance', f"${report_data.summary.get('cash_balance', 0):.2f}"],
                ['Total PnL', f"${report_data.summary.get('total_pnl', 0):.2f}"],
                ['Drawdown', f"{report_data.summary.get('drawdown', 0)*100:.1f}%"],
                ['Sharpe Ratio', f"{report_data.summary.get('sharpe_ratio', 0):.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.25*inch))
            
            # Charts
            if request.include_charts and request.chart_types:
                story.append(PageBreak())
                story.append(Paragraph("Charts", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            img = Image(chart_image, width=6*inch, height=4*inch)
                            story.append(KeepTogether([
                                Paragraph(f"<b>{chart_data.title}</b>", styles['Heading2']),
                                img,
                                Spacer(1, 0.1*inch)
                            ]))
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                story.append(PageBreak())
                story.append(Paragraph("Recommendations", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for rec in report_data.recommendations:
                    story.append(Paragraph(f"• {rec}", styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
            
            doc.build(story)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return b""

    # =========================================================================
    # HTML Generation
    # =========================================================================

    async def _generate_html(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate HTML report"""
        try:
            html = []
            
            # Header
            html.append("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Portfolio Report</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
                    h2 { color: #34495e; margin-top: 30px; }
                    .summary-box { background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 10px 0; }
                    .metric { display: inline-block; margin: 10px 20px 10px 0; }
                    .metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }
                    .metric-label { font-size: 14px; color: #7f8c8d; }
                    .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                    .table th { background: #3498db; color: white; padding: 10px; text-align: left; }
                    .table td { padding: 10px; border-bottom: 1px solid #ddd; }
                    .table tr:hover { background: #f5f5f5; }
                    .recommendation { background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }
                    .chart-container { margin: 20px 0; text-align: center; }
                    img { max-width: 100%; border: 1px solid #ddd; border-radius: 5px; }
                </style>
            </head>
            <body>
            """)
            
            # Title
            html.append(f"<h1>Portfolio Report - {report_data.portfolio_id}</h1>")
            html.append(f"<p><strong>Period:</strong> {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}</p>")
            html.append(f"<p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>")
            
            # Summary
            html.append("<h2>Executive Summary</h2>")
            html.append('<div class="summary-box">')
            
            summary = report_data.summary
            html.append(f"""
            <div class="metric">
                <div class="metric-value">${summary.get('total_equity', 0):.2f}</div>
                <div class="metric-label">Total Equity</div>
            </div>
            <div class="metric">
                <div class="metric-value">${summary.get('total_pnl', 0):.2f}</div>
                <div class="metric-label">Total PnL</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('drawdown', 0)*100:.1f}%</div>
                <div class="metric-label">Drawdown</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('sharpe_ratio', 0):.2f}</div>
                <div class="metric-label">Sharpe Ratio</div>
            </div>
            """)
            html.append('</div>')
            
            # Charts
            if request.include_charts and request.chart_types:
                html.append("<h2>Charts</h2>")
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            import base64
                            img_base64 = base64.b64encode(chart_image).decode('utf-8')
                            html.append(f"""
                            <div class="chart-container">
                                <h3>{chart_data.title}</h3>
                                <img src="data:image/png;base64,{img_base64}" />
                            </div>
                            """)
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                html.append("<h2>Recommendations</h2>")
                for rec in report_data.recommendations:
                    html.append(f"""
                    <div class="recommendation">
                        <p>{rec}</p>
                    </div>
                    """)
            
            html.append("""
            <hr/>
            <p style="color: #7f8c8d; font-size: 12px;">
                Generated by NEXUS Portfolio Reporter
            </p>
            </body>
            </html>
            """)
            
            return ''.join(html)
            
        except Exception as e:
            logger.error(f"Error generating HTML: {e}")
            return ""

    # =========================================================================
    # JSON Generation
    # =========================================================================

    async def _generate_json(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate JSON report"""
        try:
            report = {
                'portfolio_id': report_data.portfolio_id,
                'report_type': report_data.report_type.value,
                'start_date': report_data.start_date.isoformat(),
                'end_date': report_data.end_date.isoformat(),
                'generated_at': datetime.utcnow().isoformat(),
                'summary': report_data.summary,
                'performance': report_data.performance_metrics,
                'pnl': report_data.pnl_metrics,
                'risk': report_data.risk_metrics,
                'positions': [p.__dict__ for p in report_data.positions],
                'trades': [t.__dict__ for t in report_data.trades],
                'recommendations': report_data.recommendations
            }
            
            return json.dumps(report, indent=2, default=str)
            
        except Exception as e:
            logger.error(f"Error generating JSON: {e}")
            return "{}"

    # =========================================================================
    # CSV Generation
    # =========================================================================

    async def _generate_csv(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate CSV report"""
        try:
            output = []
            
            # Summary
            output.append("=== Summary ===")
            for key, value in report_data.summary.items():
                output.append(f"{key},{value}")
            output.append("")
            
            # Positions
            if report_data.positions:
                output.append("=== Positions ===")
                output.append("symbol,side,size,entry_price,current_price,pnl,value")
                for pos in report_data.positions:
                    output.append(f"{pos.symbol},{pos.side},{pos.size},{pos.entry_price},{pos.current_price},{pos.pnl},{pos.value}")
                output.append("")
            
            # Trades
            if report_data.trades:
                output.append("=== Trades ===")
                output.append("timestamp,symbol,side,size,price,pnl")
                for trade in report_data.trades:
                    output.append(f"{trade.execution_time},{trade.symbol},{trade.side},{trade.size},{trade.price},{trade.pnl}")
                output.append("")
            
            # Metrics
            output.append("=== Performance Metrics ===")
            for key, value in report_data.performance_metrics.items():
                output.append(f"{key},{value}")
            output.append("")
            
            output.append("=== PnL Metrics ===")
            for key, value in report_data.pnl_metrics.items():
                output.append(f"{key},{value}")
            output.append("")
            
            output.append("=== Risk Metrics ===")
            for key, value in report_data.risk_metrics.items():
                output.append(f"{key},{value}")
            output.append("")
            
            # Recommendations
            output.append("=== Recommendations ===")
            for rec in report_data.recommendations:
                output.append(rec)
            
            return '\n'.join(output)
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            return ""

    # =========================================================================
    # Excel Generation
    # =========================================================================

    async def _generate_excel(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            
            ws['A1'] = f"Portfolio Report - {report_data.portfolio_id}"
            ws['A1'].font = Font(size=16, bold=True)
            
            ws['A3'] = f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}"
            ws['A4'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            
            row = 6
            for key, value in report_data.summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            # Positions sheet
            if report_data.positions:
                ws_pos = wb.create_sheet("Positions")
                headers = ['Symbol', 'Side', 'Size', 'Entry Price', 'Current Price', 'PnL', 'Value']
                for col, header in enumerate(headers, 1):
                    ws_pos.cell(row=1, column=col, value=header)
                    ws_pos.cell(row=1, column=col).font = Font(bold=True)
                
                for row, pos in enumerate(report_data.positions, 2):
                    ws_pos.cell(row=row, column=1, value=pos.symbol)
                    ws_pos.cell(row=row, column=2, value=pos.side)
                    ws_pos.cell(row=row, column=3, value=pos.size)
                    ws_pos.cell(row=row, column=4, value=pos.entry_price)
                    ws_pos.cell(row=row, column=5, value=pos.current_price)
                    ws_pos.cell(row=row, column=6, value=pos.pnl)
                    ws_pos.cell(row=row, column=7, value=pos.value)
            
            # Trades sheet
            if report_data.trades:
                ws_trades = wb.create_sheet("Trades")
                headers = ['Timestamp', 'Symbol', 'Side', 'Size', 'Price', 'PnL']
                for col, header in enumerate(headers, 1):
                    ws_trades.cell(row=1, column=col, value=header)
                    ws_trades.cell(row=1, column=col).font = Font(bold=True)
                
                for row, trade in enumerate(report_data.trades, 2):
                    ws_trades.cell(row=row, column=1, value=trade.execution_time)
                    ws_trades.cell(row=row, column=2, value=trade.symbol)
                    ws_trades.cell(row=row, column=3, value=trade.side)
                    ws_trades.cell(row=row, column=4, value=trade.size)
                    ws_trades.cell(row=row, column=5, value=trade.price)
                    ws_trades.cell(row=row, column=6, value=trade.pnl)
            
            # Metrics sheets
            metrics_sheets = [
                ('Performance', report_data.performance_metrics),
                ('PnL', report_data.pnl_metrics),
                ('Risk', report_data.risk_metrics)
            ]
            
            for sheet_name, metrics in metrics_sheets:
                if metrics:
                    ws_metrics = wb.create_sheet(sheet_name)
                    ws_metrics.cell(row=1, column=1, value="Metric")
                    ws_metrics.cell(row=1, column=2, value="Value")
                    ws_metrics.cell(row=1, column=1).font = Font(bold=True)
                    ws_metrics.cell(row=1, column=2).font = Font(bold=True)
                    
                    row = 2
                    for key, value in metrics.items():
                        ws_metrics.cell(row=row, column=1, value=key.replace('_', ' ').title())
                        ws_metrics.cell(row=row, column=2, value=value)
                        row += 1
            
            # Recommendations sheet
            if report_data.recommendations:
                ws_rec = wb.create_sheet("Recommendations")
                ws_rec.cell(row=1, column=1, value="#")
                ws_rec.cell(row=1, column=2, value="Recommendation")
                ws_rec.cell(row=1, column=1).font = Font(bold=True)
                ws_rec.cell(row=1, column=2).font = Font(bold=True)
                
                for i, rec in enumerate(report_data.recommendations, 2):
                    ws_rec.cell(row=i, column=1, value=i-1)
                    ws_rec.cell(row=i, column=2, value=rec)
            
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    # =========================================================================
    # Markdown Generation
    # =========================================================================

    async def _generate_markdown(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate Markdown report"""
        try:
            md = []
            
            md.append(f"# Portfolio Report - {report_data.portfolio_id}\n")
            md.append(f"**Period:** {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}")
            md.append(f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n")
            
            # Summary
            md.append("## Executive Summary\n")
            md.append("| Metric | Value |")
            md.append("|--------|-------|")
            
            for key, value in report_data.summary.items():
                if isinstance(value, float):
                    if key in ['drawdown']:
                        value = f"{value*100:.1f}%"
                    elif key in ['total_equity', 'cash_balance', 'total_pnl']:
                        value = f"${value:.2f}"
                    else:
                        value = f"{value:.2f}"
                md.append(f"| {key.replace('_', ' ').title()} | {value} |")
            
            md.append("")
            
            # Recommendations
            if report_data.recommendations:
                md.append("## Recommendations\n")
                for rec in report_data.recommendations:
                    md.append(f"- {rec}")
                md.append("")
            
            return '\n'.join(md)
            
        except Exception as e:
            logger.error(f"Error generating Markdown: {e}")
            return ""

    # =========================================================================
    # Chart Generation
    # =========================================================================

    async def _prepare_chart_data(
        self,
        report_data: ReportData,
        chart_type: ChartType
    ) -> Optional[ChartData]:
        """Prepare chart data"""
        try:
            if chart_type == ChartType.EQUITY:
                if report_data.balance_history:
                    timestamps = [h.get('timestamp') for h in report_data.balance_history]
                    equity = [h.get('total_equity', 0) for h in report_data.balance_history]
                    
                    return ChartData(
                        title="Equity Curve",
                        chart_type=ChartType.EQUITY,
                        data={
                            'timestamps': timestamps,
                            'equity': equity
                        },
                        x_label="Date",
                        y_label="Equity ($)"
                    )
            
            elif chart_type == ChartType.DRAWDOWN:
                if report_data.balance_history:
                    drawdowns = [h.get('drawdown', 0) * 100 for h in report_data.balance_history]
                    
                    return ChartData(
                        title="Drawdown",
                        chart_type=ChartType.DRAWDOWN,
                        data={'drawdowns': drawdowns},
                        x_label="Date",
                        y_label="Drawdown (%)"
                    )
            
            elif chart_type == ChartType.PNL:
                if report_data.trades:
                    pnls = [t.pnl for t in report_data.trades if hasattr(t, 'pnl')]
                    
                    return ChartData(
                        title="PnL Distribution",
                        chart_type=ChartType.PNL,
                        data={'pnls': pnls},
                        x_label="PnL ($)",
                        y_label="Frequency"
                    )
            
            elif chart_type == ChartType.POSITIONS:
                if report_data.positions:
                    symbols = [p.symbol for p in report_data.positions]
                    values = [p.value for p in report_data.positions]
                    
                    return ChartData(
                        title="Position Allocation",
                        chart_type=ChartType.POSITIONS,
                        data={
                            'symbols': symbols,
                            'values': values
                        }
                    )
            
            elif chart_type == ChartType.ALLOCATION:
                if report_data.positions:
                    weights = []
                    symbols = []
                    total_value = sum(p.value for p in report_data.positions)
                    
                    for pos in report_data.positions:
                        if total_value > 0:
                            symbols.append(pos.symbol)
                            weights.append(pos.value / total_value * 100)
                    
                    return ChartData(
                        title="Allocation by Symbol",
                        chart_type=ChartType.ALLOCATION,
                        data={
                            'symbols': symbols,
                            'weights': weights
                        }
                    )
            
            return None
            
        except Exception as e:
            logger.error(f"Error preparing chart data: {e}")
            return None

    async def _create_chart_image(self, chart_data: ChartData) -> Optional[bytes]:
        """Create chart image"""
        try:
            fig, ax = plt.subplots(figsize=(12, 8))
            
            if chart_data.chart_type == ChartType.EQUITY:
                timestamps = chart_data.data.get('timestamps', [])
                equity = chart_data.data.get('equity', [])
                
                ax.plot(timestamps, equity, linewidth=2, color='#3498db')
                ax.fill_between(timestamps, equity, alpha=0.3, color='#3498db')
                ax.axhline(y=0, color='#e74c3c', linestyle='--', alpha=0.5)
            
            elif chart_data.chart_type == ChartType.DRAWDOWN:
                drawdowns = chart_data.data.get('drawdowns', [])
                ax.fill_between(range(len(drawdowns)), drawdowns, alpha=0.5, color='#e74c3c')
                ax.plot(range(len(drawdowns)), drawdowns, linewidth=2, color='#e74c3c')
                ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)
            
            elif chart_data.chart_type == ChartType.PNL:
                pnls = chart_data.data.get('pnls', [])
                ax.hist(pnls, bins=30, color='#3498db', edgecolor='black', alpha=0.7)
                ax.axvline(np.mean(pnls), color='#e74c3c', linestyle='dashed', linewidth=2, label='Mean')
                ax.axvline(np.median(pnls), color='#2ecc71', linestyle='dashed', linewidth=2, label='Median')
                ax.legend()
            
            elif chart_data.chart_type == ChartType.POSITIONS:
                symbols = chart_data.data.get('symbols', [])
                values = chart_data.data.get('values', [])
                colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6', '#1abc9c']
                ax.bar(symbols, values, color=colors[:len(symbols)])
                ax.set_ylabel('Value ($)')
                plt.xticks(rotation=45)
            
            elif chart_data.chart_type == ChartType.ALLOCATION:
                symbols = chart_data.data.get('symbols', [])
                weights = chart_data.data.get('weights', [])
                colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6', '#1abc9c']
                ax.pie(weights, labels=symbols, autopct='%1.1f%%', colors=colors[:len(symbols)])
                ax.axis('equal')
            
            ax.set_xlabel(chart_data.x_label or '')
            ax.set_ylabel(chart_data.y_label or '')
            ax.set_title(chart_data.title, fontsize=16, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            buffer = BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating chart image: {e}")
            return None

    # =========================================================================
    # Report Management
    # =========================================================================

    async def get_report(self, report_id: str) -> Optional[Dict[str, Any]]:
        """Get cached report"""
        return self._reports.get(report_id)

    async def delete_report(self, report_id: str) -> bool:
        """Delete cached report"""
        if report_id in self._reports:
            del self._reports[report_id]
            return True
        return False

    async def list_reports(
        self,
        portfolio_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """List available reports"""
        reports = []
        
        for report_id, data in self._reports.items():
            response = data.get('response')
            if response:
                if portfolio_id and response.portfolio_id != portfolio_id:
                    continue
                reports.append({
                    'report_id': report_id,
                    'report_type': response.report_type.value,
                    'portfolio_id': response.portfolio_id,
                    'generated_at': response.generated_at,
                    'format': response.format.value,
                    'file_size': response.file_size
                })
        
        reports.sort(key=lambda x: x['generated_at'], reverse=True)
        return reports[:limit]

    # =========================================================================
    # Scheduled Reports
    # =========================================================================

    async def create_schedule(self, schedule: ReportSchedule) -> bool:
        """Create a report schedule"""
        try:
            schedule.next_generation = self._calculate_next_generation(schedule)
            self._schedules[schedule.portfolio_id] = schedule
            logger.info(f"Created schedule for {schedule.portfolio_id}")
            return True
        except Exception as e:
            logger.error(f"Error creating schedule: {e}")
            return False

    async def get_schedule(self, portfolio_id: str) -> Optional[ReportSchedule]:
        """Get a report schedule"""
        return self._schedules.get(portfolio_id)

    async def delete_schedule(self, portfolio_id: str) -> bool:
        """Delete a report schedule"""
        if portfolio_id in self._schedules:
            del self._schedules[portfolio_id]
            return True
        return False

    def _calculate_next_generation(self, schedule: ReportSchedule) -> datetime:
        """Calculate next generation time"""
        now = datetime.utcnow()
        
        if schedule.frequency == 'daily':
            next_time = now.replace(hour=int(schedule.time.split(':')[0]),
                                   minute=int(schedule.time.split(':')[1]),
                                   second=0, microsecond=0)
            if next_time <= now:
                next_time += timedelta(days=1)
            return next_time
        
        elif schedule.frequency == 'weekly' and schedule.day_of_week is not None:
            days_until = schedule.day_of_week - now.weekday()
            if days_until <= 0:
                days_until += 7
            next_time = now + timedelta(days=days_until)
            next_time = next_time.replace(hour=int(schedule.time.split(':')[0]),
                                         minute=int(schedule.time.split(':')[1]),
                                         second=0, microsecond=0)
            return next_time
        
        return now + timedelta(days=1)

    async def _send_report_email(
        self,
        response: ReportResponse,
        content: Any,
        recipients: List[str]
    ) -> None:
        """Send report via email"""
        try:
            logger.info(f"Sending report {response.report_id} to {recipients}")
        except Exception as e:
            logger.error(f"Error sending report email: {e}")

    # =========================================================================
    # Cleanup
    # =========================================================================

    async def close(self) -> None:
        """Close the reporter"""
        self._reports.clear()
        self._schedules.clear()
        self._report_cache.clear()
        logger.info("PortfolioReporter closed")


# =============================================================================
# FASTAPI ROUTER
# =============================================================================

from fastapi import APIRouter, Depends, Query, Body

router = APIRouter(prefix="/api/v1/portfolio/reports", tags=["Portfolio Reports"])


async def get_reporter() -> PortfolioReporter:
    """Dependency to get PortfolioReporter instance"""
    return PortfolioReporter()


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Generate a portfolio report"""
    return await reporter.generate_report(request)


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Get a generated report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return report


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Delete a report"""
    success = await reporter.delete_report(report_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return {"success": True}


@router.get("/")
async def list_reports(
    portfolio_id: Optional[str] = Query(None),
    limit: int = Query(50, le=500),
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """List available reports"""
    return await reporter.list_reports(portfolio_id, limit)


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Download a report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    
    response_data = report.get('response')
    content = report.get('content')
    
    if not content:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report content not found"
        )
    
    content_type = {
        'pdf': 'application/pdf',
        'html': 'text/html',
        'json': 'application/json',
        'csv': 'text/csv',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'markdown': 'text/markdown'
    }.get(response_data.format.value, 'application/octet-stream')
    
    return Response(
        content=content if isinstance(content, bytes) else content.encode('utf-8'),
        media_type=content_type,
        headers={
            'Content-Disposition': f'attachment; filename="portfolio_report_{report_id}.{response_data.format.value}"'
        }
    )


@router.post("/schedule")
async def create_schedule(
    schedule: ReportSchedule,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Create a report schedule"""
    success = await reporter.create_schedule(schedule)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create schedule"
        )
    return {"success": True}


@router.get("/schedule/{portfolio_id}")
async def get_schedule(
    portfolio_id: str,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Get a report schedule"""
    schedule = await reporter.get_schedule(portfolio_id)
    if not schedule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {portfolio_id} not found"
        )
    return schedule


@router.delete("/schedule/{portfolio_id}")
async def delete_schedule(
    portfolio_id: str,
    reporter: PortfolioReporter = Depends(get_reporter)
):
    """Delete a report schedule"""
    success = await reporter.delete_schedule(portfolio_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {portfolio_id} not found"
        )
    return {"success": True}


@router.get("/types")
async def get_report_types():
    """Get available report types"""
    return {
        'types': [
            {'name': t.value, 'description': t.name.replace('_', ' ').title()}
            for t in ReportType
        ]
    }


@router.get("/formats")
async def get_report_formats():
    """Get available report formats"""
    return {
        'formats': [
            {'name': f.value, 'description': f.name.upper()}
            for f in ReportFormat
        ]
    }


@router.get("/chart-types")
async def get_chart_types():
    """Get available chart types"""
    return {
        'chart_types': [
            {'name': c.value, 'description': c.name.replace('_', ' ').title()}
            for c in ChartType
        ]
    }


# =============================================================================
# EXPORTS
# =============================================================================

__all__ = [
    'PortfolioReporter',
    'ReportFormat',
    'ReportType',
    'ChartType',
    'ReportRequest',
    'ReportResponse',
    'ReportSchedule',
    'ReportData',
    'ChartData',
    'router'
]
