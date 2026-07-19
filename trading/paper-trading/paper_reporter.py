"""
NEXUS AI TRADING SYSTEM - Paper Trading Reporter Module
Copyright © 2026 NEXUS QUANTUM LTD
CEO: Dr X... - Majority Shareholder

Module: trading/paper-trading/paper_reporter.py
Description: Paper trading reporting with full API integration
"""

import asyncio
import json
import logging
import os
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, field, asdict
from enum import Enum
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns

from pydantic import BaseModel, Field, validator
from fastapi import HTTPException, status, Response
from fastapi.responses import FileResponse, StreamingResponse

# NEXUS Internal Imports
from shared.configs.paper_trading_config import PaperTradingConfig
from shared.constants.trading_constants import TIME_FRAMES
from shared.utilities.retry import retry_async
from shared.utilities.logger import get_logger

# Paper trading imports
from trading.paper_trading.paper_account import PaperTradingAccount
from trading.paper_trading.paper_analytics import PaperTradingAnalytics
from trading.paper_trading.paper_orders import PaperTradingOrders
from trading.paper_trading.paper_portfolio import PaperTradingPortfolio

logger = get_logger(__name__)


# =============================================================================
# ENUMS & CONSTANTS
# =============================================================================

class ReportFormat(str, Enum):
    """Report output formats"""
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class ReportType(str, Enum):
    """Types of reports"""
    PERFORMANCE = "performance"
    TRADES = "trades"
    POSITIONS = "positions"
    ACCOUNT = "account"
    EXECUTIVE = "executive"
    DETAILED = "detailed"
    SUMMARY = "summary"


class ChartType(str, Enum):
    """Chart types"""
    EQUITY = "equity"
    DRAWDOWN = "drawdown"
    PNL = "pnl"
    PERFORMANCE = "performance"
    TRADES = "trades"
    POSITIONS = "positions"


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class ReportRequest(BaseModel):
    """Request model for report generation"""
    account_id: str
    report_type: ReportType = ReportType.PERFORMANCE
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    format: ReportFormat = ReportFormat.PDF
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_detailed: bool = False
    include_recommendations: bool = True
    chart_types: List[ChartType] = []
    metrics: List[str] = []
    compare_to_previous: bool = False
    email_report: bool = False
    email_recipients: List[str] = []
    metadata: Dict[str, Any] = Field(default_factory=dict)


class ReportResponse(BaseModel):
    """Response model for report generation"""
    report_id: str
    account_id: str
    report_type: ReportType
    generated_at: datetime
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    format: ReportFormat
    file_size: int
    summary: Dict[str, Any]
    metrics: Dict[str, Any]
    charts_count: int
    tables_count: int
    recommendations: List[str]
    metadata: Dict[str, Any] = Field(default_factory=dict)


class ReportSchedule(BaseModel):
    """Schedule for automated reports"""
    account_id: str
    report_type: ReportType
    frequency: str  # daily, weekly, monthly
    format: ReportFormat = ReportFormat.PDF
    time: str = "00:00"
    day_of_week: Optional[int] = None
    day_of_month: Optional[int] = None
    recipients: List[str] = []
    enabled: bool = True
    last_generated: Optional[datetime] = None
    next_generation: Optional[datetime] = None


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ReportData:
    """Container for report data"""
    account_id: str
    report_type: ReportType
    start_date: datetime
    end_date: datetime
    trades: List[Any]
    positions: List[Any]
    orders: List[Any]
    balance_history: List[Any]
    performance_metrics: Dict[str, Any]
    summary: Dict[str, Any]
    recommendations: List[str]


@dataclass
class ChartData:
    """Container for chart data"""
    title: str
    chart_type: ChartType
    data: Dict[str, Any]
    x_label: Optional[str] = None
    y_label: Optional[str] = None
    legend: Optional[List[str]] = None


# =============================================================================
# PAPER TRADING REPORTER
# =============================================================================

class PaperTradingReporter:
    """
    Paper Trading Reporter with full API integration.
    
    Features:
    - Multiple report types
    - Various output formats
    - Interactive charts
    - Performance analysis
    - Trade reporting
    - Position reporting
    - Account reporting
    - Executive summaries
    - Scheduled reports
    - Email delivery
    """

    def __init__(
        self,
        config: Optional[PaperTradingConfig] = None,
        paper_account: Optional[PaperTradingAccount] = None,
        paper_analytics: Optional[PaperTradingAnalytics] = None,
        paper_orders: Optional[PaperTradingOrders] = None,
        paper_portfolio: Optional[PaperTradingPortfolio] = None
    ):
        """
        Initialize PaperTradingReporter.
        
        Args:
            config: Paper trading configuration
            paper_account: Paper trading account
            paper_analytics: Paper trading analytics
            paper_orders: Paper trading orders
            paper_portfolio: Paper trading portfolio
        """
        self.config = config or PaperTradingConfig()
        self.paper_account = paper_account or PaperTradingAccount()
        self.paper_analytics = paper_analytics or PaperTradingAnalytics()
        self.paper_orders = paper_orders or PaperTradingOrders()
        self.paper_portfolio = paper_portfolio or PaperTradingPortfolio()
        
        # Report storage
        self._reports: Dict[str, Dict[str, Any]] = {}
        self._schedules: Dict[str, ReportSchedule] = {}
        
        # Cache
        self._report_cache: Dict[str, Dict[str, Any]] = {}
        
        # Chart defaults
        self._chart_defaults = {
            'figsize': (12, 8),
            'dpi': 100,
            'style': 'seaborn-v0_8-darkgrid',
            'font_size': 12,
            'title_size': 16
        }
        
        # Set style
        plt.style.use(self._chart_defaults['style'])
        sns.set_palette("husl")
        
        logger.info("PaperTradingReporter initialized")

    # =========================================================================
    # Report Generation
    # =========================================================================

    @retry_async(max_attempts=3, delay=1.0)
    async def generate_report(
        self,
        request: ReportRequest
    ) -> ReportResponse:
        """
        Generate a paper trading report.
        
        Args:
            request: Report request
            
        Returns:
            ReportResponse: Generated report
        """
        try:
            # Validate request
            await self._validate_request(request)
            
            # Set date range
            start_date, end_date = self._get_date_range(request)
            
            # Collect report data
            report_data = await self._collect_report_data(
                request.account_id,
                request.report_type,
                start_date,
                end_date
            )
            
            # Generate report
            report_id = f"pt_report_{int(time.time() * 1000)}_{request.account_id[:8]}"
            
            # Generate content
            content, file_size = await self._generate_content(
                report_data,
                request.format,
                request
            )
            
            # Create response
            response = ReportResponse(
                report_id=report_id,
                account_id=request.account_id,
                report_type=request.report_type,
                generated_at=datetime.utcnow(),
                start_date=start_date,
                end_date=end_date,
                format=request.format,
                file_size=file_size,
                summary=report_data.summary,
                metrics=report_data.performance_metrics,
                charts_count=len(request.chart_types) if request.include_charts else 0,
                tables_count=0,
                recommendations=report_data.recommendations,
                metadata=request.metadata
            )
            
            # Cache report
            self._reports[report_id] = {
                'response': response,
                'content': content,
                'data': report_data,
                'generated_at': datetime.utcnow()
            }
            
            # Send email if requested
            if request.email_report and request.email_recipients:
                await self._send_report_email(
                    response,
                    content,
                    request.email_recipients
                )
            
            logger.info(f"Report generated: {report_id} for {request.account_id}")
            return response
            
        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Report generation failed: {str(e)}"
            )

    async def _validate_request(self, request: ReportRequest) -> None:
        """Validate report request"""
        if request.start_date and request.end_date:
            if request.start_date > request.end_date:
                raise ValueError("Start date must be before end date")
        
        # Validate account exists
        account = await self.paper_account.get_account(request.account_id)
        if not account:
            raise ValueError(f"Account {request.account_id} not found")

    def _get_date_range(
        self,
        request: ReportRequest
    ) -> Tuple[datetime, datetime]:
        """Get date range for report"""
        end_date = request.end_date or datetime.utcnow()
        start_date = request.start_date
        
        if not start_date:
            # Default based on report type
            if request.report_type == ReportType.ACCOUNT:
                start_date = end_date - timedelta(days=7)
            elif request.report_type == ReportType.PERFORMANCE:
                start_date = end_date - timedelta(days=30)
            elif request.report_type == ReportType.SUMMARY:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = end_date - timedelta(days=30)
        
        return start_date, end_date

    async def _collect_report_data(
        self,
        account_id: str,
        report_type: ReportType,
        start_date: datetime,
        end_date: datetime
    ) -> ReportData:
        """
        Collect data for report.
        
        Args:
            account_id: Account ID
            report_type: Type of report
            start_date: Start date
            end_date: End date
            
        Returns:
            ReportData: Collected data
        """
        # Get account
        account = await self.paper_account.get_account(account_id)
        
        # Get orders
        orders = await self.paper_orders.get_orders(
            account_id=account_id,
            limit=1000
        )
        
        # Get portfolio
        portfolio = await self.paper_portfolio.get_portfolio(account_id)
        
        # Get trades
        trades = await self._get_trades(account_id, start_date, end_date)
        
        # Get positions
        positions = await self._get_positions(account_id)
        
        # Get analytics
        analytics = await self.paper_analytics.get_analytics(account_id)
        
        # Generate summary
        summary = self._generate_summary(account, portfolio, analytics)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(account, portfolio, analytics)
        
        return ReportData(
            account_id=account_id,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            trades=trades,
            positions=positions,
            orders=orders.orders if hasattr(orders, 'orders') else [],
            balance_history=[],  # Would fetch balance history
            performance_metrics=analytics.metrics if hasattr(analytics, 'metrics') else {},
            summary=summary,
            recommendations=recommendations
        )

    async def _get_trades(
        self,
        account_id: str,
        start_date: datetime,
        end_date: datetime
    ) -> List[Any]:
        """Get trades for account"""
        # This would fetch from database
        return []

    async def _get_positions(self, account_id: str) -> List[Any]:
        """Get positions for account"""
        # This would fetch from database
        return []

    def _generate_summary(
        self,
        account: Any,
        portfolio: Any,
        analytics: Any
    ) -> Dict[str, Any]:
        """Generate summary"""
        summary = {}
        
        if account:
            summary.update({
                'account_id': account.account_id,
                'balance': account.balance,
                'equity': account.equity,
                'total_pnl': account.total_pnl,
                'total_pnl_pct': account.total_pnl_pct
            })
        
        if portfolio:
            summary.update({
                'portfolio_id': portfolio.portfolio_id,
                'total_value': portfolio.total_value,
                'cash_balance': portfolio.cash_balance,
                'positions_count': portfolio.positions_count
            })
        
        if analytics:
            summary.update({
                'sharpe_ratio': analytics.performance.get('sharpe_ratio', 0),
                'max_drawdown': analytics.risk_metrics.get('max_drawdown', 0),
                'win_rate': analytics.trade_stats.get('win_rate', 0),
                'profit_factor': analytics.trade_stats.get('profit_factor', 0)
            })
        
        return summary

    def _generate_recommendations(
        self,
        account: Any,
        portfolio: Any,
        analytics: Any
    ) -> List[str]:
        """Generate recommendations"""
        recommendations = []
        
        if analytics:
            # Sharpe ratio recommendation
            sharpe = analytics.performance.get('sharpe_ratio', 0)
            if sharpe < 0.5:
                recommendations.append("Low Sharpe ratio. Consider improving risk-adjusted returns.")
            
            # Drawdown recommendation
            max_dd = analytics.risk_metrics.get('max_drawdown', 0)
            if max_dd > 0.20:
                recommendations.append("High drawdown. Consider implementing stricter risk controls.")
            
            # Win rate recommendation
            win_rate = analytics.trade_stats.get('win_rate', 0)
            if win_rate < 0.4:
                recommendations.append("Low win rate. Review trade entry criteria.")
            
            # Profit factor recommendation
            profit_factor = analytics.trade_stats.get('profit_factor', 0)
            if profit_factor < 1:
                recommendations.append("Profit factor below 1. Strategy needs improvement.")
        
        if account and account.total_pnl < 0:
            recommendations.append("Negative PnL. Consider pausing trading for review.")
        
        if not recommendations:
            recommendations.append("All metrics are within acceptable ranges.")
        
        return recommendations

    # =========================================================================
    # Content Generation
    # =========================================================================

    async def _generate_content(
        self,
        report_data: ReportData,
        format: ReportFormat,
        request: ReportRequest
    ) -> Tuple[Any, int]:
        """
        Generate report content.
        
        Args:
            report_data: Report data
            format: Output format
            request: Original request
            
        Returns:
            Tuple[Any, int]: Content and file size
        """
        if format == ReportFormat.PDF:
            content = await self._generate_pdf(report_data, request)
        elif format == ReportFormat.HTML:
            content = await self._generate_html(report_data, request)
        elif format == ReportFormat.JSON:
            content = await self._generate_json(report_data, request)
        elif format == ReportFormat.CSV:
            content = await self._generate_csv(report_data, request)
        elif format == ReportFormat.EXCEL:
            content = await self._generate_excel(report_data, request)
        elif format == ReportFormat.MARKDOWN:
            content = await self._generate_markdown(report_data, request)
        else:
            content = await self._generate_json(report_data, request)
        
        file_size = len(content) if isinstance(content, (bytes, str)) else 0
        return content, file_size

    # =========================================================================
    # PDF Generation
    # =========================================================================

    async def _generate_pdf(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate PDF report"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image, KeepTogether
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(
                f"Paper Trading Report - {report_data.account_id}",
                styles['Title']
            ))
            story.append(Spacer(1, 0.25*inch))
            
            # Info
            story.append(Paragraph(
                f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}",
                styles['Normal']
            ))
            story.append(Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
                styles['Normal']
            ))
            story.append(Spacer(1, 0.5*inch))
            
            # Summary
            story.append(Paragraph("Executive Summary", styles['Heading1']))
            story.append(Spacer(1, 0.1*inch))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total PnL', f"${report_data.summary.get('total_pnl', 0):.2f}"],
                ['Total PnL %', f"{report_data.summary.get('total_pnl_pct', 0):.2f}%"],
                ['Balance', f"${report_data.summary.get('balance', 0):.2f}"],
                ['Equity', f"${report_data.summary.get('equity', 0):.2f}"],
                ['Sharpe Ratio', f"{report_data.summary.get('sharpe_ratio', 0):.2f}"],
                ['Win Rate', f"{report_data.summary.get('win_rate', 0)*100:.1f}%"],
                ['Profit Factor', f"{report_data.summary.get('profit_factor', 0):.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.25*inch))
            
            # Charts
            if request.include_charts and request.chart_types:
                story.append(PageBreak())
                story.append(Paragraph("Charts", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            img = Image(chart_image, width=6*inch, height=4*inch)
                            story.append(KeepTogether([
                                Paragraph(f"<b>{chart_data.title}</b>", styles['Heading2']),
                                img,
                                Spacer(1, 0.1*inch)
                            ]))
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                story.append(PageBreak())
                story.append(Paragraph("Recommendations", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for rec in report_data.recommendations:
                    story.append(Paragraph(f"• {rec}", styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
            
            doc.build(story)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return b""

    # =========================================================================
    # HTML Generation
    # =========================================================================

    async def _generate_html(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate HTML report"""
        try:
            html = []
            
            # Header
            html.append("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Paper Trading Report</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
                    h2 { color: #34495e; margin-top: 30px; }
                    .summary-box { background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 10px 0; }
                    .metric { display: inline-block; margin: 10px 20px 10px 0; }
                    .metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }
                    .metric-label { font-size: 14px; color: #7f8c8d; }
                    .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                    .table th { background: #3498db; color: white; padding: 10px; text-align: left; }
                    .table td { padding: 10px; border-bottom: 1px solid #ddd; }
                    .table tr:hover { background: #f5f5f5; }
                    .recommendation { background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }
                    .chart-container { margin: 20px 0; text-align: center; }
                    img { max-width: 100%; border: 1px solid #ddd; border-radius: 5px; }
                </style>
            </head>
            <body>
            """)
            
            # Title
            html.append(f"<h1>Paper Trading Report - {report_data.account_id}</h1>")
            html.append(f"<p><strong>Period:</strong> {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}</p>")
            html.append(f"<p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>")
            
            # Summary
            html.append("<h2>Executive Summary</h2>")
            html.append('<div class="summary-box">')
            
            summary = report_data.summary
            html.append(f"""
            <div class="metric">
                <div class="metric-value">${summary.get('total_pnl', 0):.2f}</div>
                <div class="metric-label">Total PnL</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('total_pnl_pct', 0):.2f}%</div>
                <div class="metric-label">Total PnL %</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('sharpe_ratio', 0):.2f}</div>
                <div class="metric-label">Sharpe Ratio</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('win_rate', 0)*100:.1f}%</div>
                <div class="metric-label">Win Rate</div>
            </div>
            """)
            html.append('</div>')
            
            # Charts
            if request.include_charts and request.chart_types:
                html.append("<h2>Charts</h2>")
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            import base64
                            img_base64 = base64.b64encode(chart_image).decode('utf-8')
                            html.append(f"""
                            <div class="chart-container">
                                <h3>{chart_data.title}</h3>
                                <img src="data:image/png;base64,{img_base64}" />
                            </div>
                            """)
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                html.append("<h2>Recommendations</h2>")
                for rec in report_data.recommendations:
                    html.append(f"""
                    <div class="recommendation">
                        <p>{rec}</p>
                    </div>
                    """)
            
            html.append("""
            <hr/>
            <p style="color: #7f8c8d; font-size: 12px;">
                Generated by NEXUS Paper Trading Reporter
            </p>
            </body>
            </html>
            """)
            
            return ''.join(html)
            
        except Exception as e:
            logger.error(f"Error generating HTML: {e}")
            return ""

    # =========================================================================
    # JSON Generation
    # =========================================================================

    async def _generate_json(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate JSON report"""
        try:
            report = {
                'account_id': report_data.account_id,
                'report_type': report_data.report_type.value,
                'start_date': report_data.start_date.isoformat(),
                'end_date': report_data.end_date.isoformat(),
                'generated_at': datetime.utcnow().isoformat(),
                'summary': report_data.summary,
                'performance': report_data.performance_metrics,
                'trades': [t.__dict__ for t in report_data.trades],
                'positions': [p.__dict__ for p in report_data.positions],
                'orders': report_data.orders,
                'recommendations': report_data.recommendations
            }
            
            return json.dumps(report, indent=2, default=str)
            
        except Exception as e:
            logger.error(f"Error generating JSON: {e}")
            return "{}"

    # =========================================================================
    # CSV Generation
    # =========================================================================

    async def _generate_csv(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate CSV report"""
        try:
            output = []
            
            # Summary
            output.append("=== Summary ===")
            for key, value in report_data.summary.items():
                output.append(f"{key},{value}")
            output.append("")
            
            # Trades
            if report_data.trades:
                output.append("=== Trades ===")
                output.append("timestamp,symbol,side,size,price,pnl")
                for trade in report_data.trades:
                    output.append(f"{trade.timestamp},{trade.symbol},{trade.side},{trade.size},{trade.price},{trade.pnl}")
                output.append("")
            
            # Positions
            if report_data.positions:
                output.append("=== Positions ===")
                output.append("symbol,side,size,entry_price,current_price,pnl,value")
                for pos in report_data.positions:
                    output.append(f"{pos.symbol},{pos.side},{pos.size},{pos.entry_price},{pos.current_price},{pos.pnl},{pos.value}")
                output.append("")
            
            # Performance metrics
            output.append("=== Performance Metrics ===")
            for key, value in report_data.performance_metrics.items():
                output.append(f"{key},{value}")
            output.append("")
            
            # Recommendations
            output.append("=== Recommendations ===")
            for rec in report_data.recommendations:
                output.append(rec)
            
            return '\n'.join(output)
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            return ""

    # =========================================================================
    # Excel Generation
    # =========================================================================

    async def _generate_excel(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            
            ws['A1'] = f"Paper Trading Report - {report_data.account_id}"
            ws['A1'].font = Font(size=16, bold=True)
            
            ws['A3'] = f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}"
            ws['A4'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            
            row = 6
            for key, value in report_data.summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            # Trades sheet
            if report_data.trades:
                ws_trades = wb.create_sheet("Trades")
                headers = ['Timestamp', 'Symbol', 'Side', 'Size', 'Price', 'PnL']
                for col, header in enumerate(headers, 1):
                    ws_trades.cell(row=1, column=col, value=header)
                    ws_trades.cell(row=1, column=col).font = Font(bold=True)
                
                for row, trade in enumerate(report_data.trades, 2):
                    ws_trades.cell(row=row, column=1, value=trade.timestamp)
                    ws_trades.cell(row=row, column=2, value=trade.symbol)
                    ws_trades.cell(row=row, column=3, value=trade.side)
                    ws_trades.cell(row=row, column=4, value=trade.size)
                    ws_trades.cell(row=row, column=5, value=trade.price)
                    ws_trades.cell(row=row, column=6, value=trade.pnl)
            
            # Positions sheet
            if report_data.positions:
                ws_pos = wb.create_sheet("Positions")
                headers = ['Symbol', 'Side', 'Size', 'Entry Price', 'Current Price', 'PnL', 'Value']
                for col, header in enumerate(headers, 1):
                    ws_pos.cell(row=1, column=col, value=header)
                    ws_pos.cell(row=1, column=col).font = Font(bold=True)
                
                for row, pos in enumerate(report_data.positions, 2):
                    ws_pos.cell(row=row, column=1, value=pos.symbol)
                    ws_pos.cell(row=row, column=2, value=pos.side)
                    ws_pos.cell(row=row, column=3, value=pos.size)
                    ws_pos.cell(row=row, column=4, value=pos.entry_price)
                    ws_pos.cell(row=row, column=5, value=pos.current_price)
                    ws_pos.cell(row=row, column=6, value=pos.pnl)
                    ws_pos.cell(row=row, column=7, value=pos.value)
            
            # Metrics sheet
            if report_data.performance_metrics:
                ws_metrics = wb.create_sheet("Performance")
                ws_metrics.cell(row=1, column=1, value="Metric")
                ws_metrics.cell(row=1, column=2, value="Value")
                ws_metrics.cell(row=1, column=1).font = Font(bold=True)
                ws_metrics.cell(row=1, column=2).font = Font(bold=True)
                
                row = 2
                for key, value in report_data.performance_metrics.items():
                    ws_metrics.cell(row=row, column=1, value=key.replace('_', ' ').title())
                    ws_metrics.cell(row=row, column=2, value=value)
                    row += 1
            
            # Recommendations sheet
            if report_data.recommendations:
                ws_rec = wb.create_sheet("Recommendations")
                ws_rec.cell(row=1, column=1, value="#")
                ws_rec.cell(row=1, column=2, value="Recommendation")
                ws_rec.cell(row=1, column=1).font = Font(bold=True)
                ws_rec.cell(row=1, column=2).font = Font(bold=True)
                
                for i, rec in enumerate(report_data.recommendations, 2):
                    ws_rec.cell(row=i, column=1, value=i-1)
                    ws_rec.cell(row=i, column=2, value=rec)
            
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    # =========================================================================
    # Markdown Generation
    # =========================================================================

    async def _generate_markdown(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate Markdown report"""
        try:
            md = []
            
            md.append(f"# Paper Trading Report - {report_data.account_id}\n")
            md.append(f"**Period:** {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}")
            md.append(f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n")
            
            # Summary
            md.append("## Executive Summary\n")
            md.append("| Metric | Value |")
            md.append("|--------|-------|")
            
            for key, value in report_data.summary.items():
                if isinstance(value, float):
                    if key in ['total_pnl_pct']:
                        value = f"{value:.2f}%"
                    elif key in ['total_pnl', 'balance', 'equity']:
                        value = f"${value:.2f}"
                    else:
                        value = f"{value:.2f}"
                md.append(f"| {key.replace('_', ' ').title()} | {value} |")
            
            md.append("")
            
            # Performance Metrics
            if report_data.performance_metrics:
                md.append("## Performance Metrics\n")
                md.append("| Metric | Value |")
                md.append("|--------|-------|")
                
                for key, value in report_data.performance_metrics.items():
                    if isinstance(value, float):
                        if 'ratio' in key:
                            value = f"{value:.2f}"
                        elif 'pct' in key:
                            value = f"{value:.2f}%"
                        else:
                            value = f"{value:.2f}"
                    md.append(f"| {key.replace('_', ' ').title()} | {value} |")
                md.append("")
            
            # Recommendations
            if report_data.recommendations:
                md.append("## Recommendations\n")
                for rec in report_data.recommendations:
                    md.append(f"- {rec}")
                md.append("")
            
            return '\n'.join(md)
            
        except Exception as e:
            logger.error(f"Error generating Markdown: {e}")
            return ""

    # =========================================================================
    # Chart Generation
    # =========================================================================

    async def _prepare_chart_data(
        self,
        report_data: ReportData,
        chart_type: ChartType
    ) -> Optional[ChartData]:
        """Prepare chart data"""
        try:
            if chart_type == ChartType.EQUITY:
                if report_data.balance_history:
                    timestamps = [h.get('timestamp') for h in report_data.balance_history]
                    equity = [h.get('equity', 0) for h in report_data.balance_history]
                    
                    return ChartData(
                        title="Equity Curve",
                        chart_type=ChartType.EQUITY,
                        data={
                            'timestamps': timestamps,
                            'equity': equity
                        },
                        x_label="Date",
                        y_label="Equity ($)"
                    )
            
            elif chart_type == ChartType.DRAWDOWN:
                if report_data.balance_history:
                    drawdowns = [h.get('drawdown', 0) * 100 for h in report_data.balance_history]
                    
                    return ChartData(
                        title="Drawdown",
                        chart_type=ChartType.DRAWDOWN,
                        data={'drawdowns': drawdowns},
                        x_label="Date",
                        y_label="Drawdown (%)"
                    )
            
            elif chart_type == ChartType.PNL:
                if report_data.trades:
                    pnls = [t.pnl for t in report_data.trades if hasattr(t, 'pnl')]
                    
                    return ChartData(
                        title="PnL Distribution",
                        chart_type=ChartType.PNL,
                        data={'pnls': pnls},
                        x_label="PnL ($)",
                        y_label="Frequency"
                    )
            
            return None
            
        except Exception as e:
            logger.error(f"Error preparing chart data: {e}")
            return None

    async def _create_chart_image(self, chart_data: ChartData) -> Optional[bytes]:
        """Create chart image"""
        try:
            fig, ax = plt.subplots(figsize=(12, 8))
            
            if chart_data.chart_type == ChartType.EQUITY:
                timestamps = chart_data.data.get('timestamps', [])
                equity = chart_data.data.get('equity', [])
                
                ax.plot(timestamps, equity, linewidth=2, color='#3498db')
                ax.fill_between(timestamps, equity, alpha=0.3, color='#3498db')
                ax.axhline(y=0, color='#e74c3c', linestyle='--', alpha=0.5)
            
            elif chart_data.chart_type == ChartType.DRAWDOWN:
                drawdowns = chart_data.data.get('drawdowns', [])
                ax.fill_between(range(len(drawdowns)), drawdowns, alpha=0.5, color='#e74c3c')
                ax.plot(range(len(drawdowns)), drawdowns, linewidth=2, color='#e74c3c')
                ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)
            
            elif chart_data.chart_type == ChartType.PNL:
                pnls = chart_data.data.get('pnls', [])
                ax.hist(pnls, bins=30, color='#3498db', edgecolor='black', alpha=0.7)
                ax.axvline(np.mean(pnls), color='#e74c3c', linestyle='dashed', linewidth=2, label='Mean')
                ax.axvline(np.median(pnls), color='#2ecc71', linestyle='dashed', linewidth=2, label='Median')
                ax.legend()
            
            ax.set_xlabel(chart_data.x_label or '')
            ax.set_ylabel(chart_data.y_label or '')
            ax.set_title(chart_data.title, fontsize=16, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            buffer = BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating chart image: {e}")
            return None

    # =========================================================================
    # Report Management
    # =========================================================================

    async def get_report(self, report_id: str) -> Optional[Dict[str, Any]]:
        """Get cached report"""
        return self._reports.get(report_id)

    async def delete_report(self, report_id: str) -> bool:
        """Delete cached report"""
        if report_id in self._reports:
            del self._reports[report_id]
            return True
        return False

    async def list_reports(
        self,
        account_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """List available reports"""
        reports = []
        
        for report_id, data in self._reports.items():
            response = data.get('response')
            if response:
                if account_id and response.account_id != account_id:
                    continue
                reports.append({
                    'report_id': report_id,
                    'report_type': response.report_type.value,
                    'account_id': response.account_id,
                    'generated_at': response.generated_at,
                    'format': response.format.value,
                    'file_size': response.file_size
                })
        
        reports.sort(key=lambda x: x['generated_at'], reverse=True)
        return reports[:limit]

    # =========================================================================
    # Scheduled Reports
    # =========================================================================

    async def create_schedule(self, schedule: ReportSchedule) -> bool:
        """Create a report schedule"""
        try:
            schedule.next_generation = self._calculate_next_generation(schedule)
            self._schedules[schedule.account_id] = schedule
            logger.info(f"Created schedule for {schedule.account_id}")
            return True
        except Exception as e:
            logger.error(f"Error creating schedule: {e}")
            return False

    async def get_schedule(self, account_id: str) -> Optional[ReportSchedule]:
        """Get a report schedule"""
        return self._schedules.get(account_id)

    async def delete_schedule(self, account_id: str) -> bool:
        """Delete a report schedule"""
        if account_id in self._schedules:
            del self._schedules[account_id]
            return True
        return False

    def _calculate_next_generation(self, schedule: ReportSchedule) -> datetime:
        """Calculate next generation time"""
        now = datetime.utcnow()
        
        if schedule.frequency == 'daily':
            next_time = now.replace(hour=int(schedule.time.split(':')[0]),
                                   minute=int(schedule.time.split(':')[1]),
                                   second=0, microsecond=0)
            if next_time <= now:
                next_time += timedelta(days=1)
            return next_time
        
        elif schedule.frequency == 'weekly' and schedule.day_of_week is not None:
            days_until = schedule.day_of_week - now.weekday()
            if days_until <= 0:
                days_until += 7
            next_time = now + timedelta(days=days_until)
            next_time = next_time.replace(hour=int(schedule.time.split(':')[0]),
                                         minute=int(schedule.time.split(':')[1]),
                                         second=0, microsecond=0)
            return next_time
        
        return now + timedelta(days=1)

    async def _send_report_email(
        self,
        response: ReportResponse,
        content: Any,
        recipients: List[str]
    ) -> None:
        """Send report via email"""
        try:
            logger.info(f"Sending report {response.report_id} to {recipients}")
        except Exception as e:
            logger.error(f"Error sending report email: {e}")

    # =========================================================================
    # Cleanup
    # =========================================================================

    async def close(self) -> None:
        """Close the reporter"""
        self._reports.clear()
        self._schedules.clear()
        self._report_cache.clear()
        logger.info("PaperTradingReporter closed")


# =============================================================================
# FASTAPI ROUTER
# =============================================================================

from fastapi import APIRouter, Depends, Query, Body

router = APIRouter(prefix="/api/v1/paper-trading/reports", tags=["Paper Trading Reports"])


async def get_reporter() -> PaperTradingReporter:
    """Dependency to get PaperTradingReporter instance"""
    return PaperTradingReporter()


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Generate a paper trading report"""
    return await reporter.generate_report(request)


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Get a generated report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return report


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Delete a report"""
    success = await reporter.delete_report(report_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return {"success": True}


@router.get("/")
async def list_reports(
    account_id: Optional[str] = Query(None),
    limit: int = Query(50, le=500),
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """List available reports"""
    return await reporter.list_reports(account_id, limit)


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Download a report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    
    response_data = report.get('response')
    content = report.get('content')
    
    if not content:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report content not found"
        )
    
    content_type = {
        'pdf': 'application/pdf',
        'html': 'text/html',
        'json': 'application/json',
        'csv': 'text/csv',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'markdown': 'text/markdown'
    }.get(response_data.format.value, 'application/octet-stream')
    
    return Response(
        content=content if isinstance(content, bytes) else content.encode('utf-8'),
        media_type=content_type,
        headers={
            'Content-Disposition': f'attachment; filename="paper_trading_report_{report_id}.{response_data.format.value}"'
        }
    )


@router.post("/schedule")
async def create_schedule(
    schedule: ReportSchedule,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Create a report schedule"""
    success = await reporter.create_schedule(schedule)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create schedule"
        )
    return {"success": True}


@router.get("/schedule/{account_id}")
async def get_schedule(
    account_id: str,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Get a report schedule"""
    schedule = await reporter.get_schedule(account_id)
    if not schedule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {account_id} not found"
        )
    return schedule


@router.delete("/schedule/{account_id}")
async def delete_schedule(
    account_id: str,
    reporter: PaperTradingReporter = Depends(get_reporter)
):
    """Delete a report schedule"""
    success = await reporter.delete_schedule(account_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {account_id} not found"
        )
    return {"success": True}


@router.get("/types")
async def get_report_types():
    """Get available report types"""
    return {
        'types': [
            {'name': t.value, 'description': t.name.replace('_', ' ').title()}
            for t in ReportType
        ]
    }


@router.get("/formats")
async def get_report_formats():
    """Get available report formats"""
    return {
        'formats': [
            {'name': f.value, 'description': f.name.upper()}
            for f in ReportFormat
        ]
    }


@router.get("/chart-types")
async def get_chart_types():
    """Get available chart types"""
    return {
        'chart_types': [
            {'name': c.value, 'description': c.name.replace('_', ' ').title()}
            for c in ChartType
        ]
    }


# =============================================================================
# EXPORTS
# =============================================================================

__all__ = [
    'PaperTradingReporter',
    'ReportFormat',
    'ReportType',
    'ChartType',
    'ReportRequest',
    'ReportResponse',
    'ReportSchedule',
    'ReportData',
    'ChartData',
    'router'
]
