"""
NEXUS AI TRADING SYSTEM - Market Making Reporter Module
Copyright © 2026 NEXUS QUANTUM LTD
CEO: Dr X... - Majority Shareholder

Module: trading/market-making/reporter.py
Description: Comprehensive market making reporting with full API integration
"""

import asyncio
import json
import logging
import os
import tempfile
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, field, asdict
from enum import Enum
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns

from pydantic import BaseModel, Field, validator
from fastapi import HTTPException, status, Response
from fastapi.responses import FileResponse, StreamingResponse

# NEXUS Internal Imports
from shared.configs.market_making_config import MarketMakingConfig
from shared.configs.reporting_config import ReportingConfig
from shared.constants.trading_constants import TIME_FRAMES
from shared.utilities.retry import retry_async
from shared.utilities.logger import get_logger
from shared.utilities.cache_utils import cached

# Database imports
from backend.database.repositories.order_repository import OrderRepository
from backend.database.repositories.trade_repository import TradeRepository
from backend.database.repositories.position_repository import PositionRepository

# Market making imports
from trading.market_making.analytics import MarketMakingAnalytics
from trading.market_making.market_maker import MarketMaker

logger = get_logger(__name__)


# =============================================================================
# ENUMS & CONSTANTS
# =============================================================================

class ReportFormat(str, Enum):
    """Report output formats"""
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class ReportType(str, Enum):
    """Types of reports"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    PERFORMANCE = "performance"
    RISK = "risk"
    EXECUTION = "execution"
    LIQUIDITY = "liquidity"
    INVENTORY = "inventory"
    PROFITABILITY = "profitability"
    SUMMARY = "summary"
    DETAILED = "detailed"
    EXECUTIVE = "executive"
    COMPLIANCE = "compliance"


class ChartType(str, Enum):
    """Chart types"""
    EQUITY = "equity"
    DRAWDOWN = "drawdown"
    SPREAD = "spread"
    VOLUME = "volume"
    INVENTORY = "inventory"
    PNL = "pnl"
    TRADES = "trades"
    LIQUIDITY = "liquidity"
    RISK = "risk"
    HEATMAP = "heatmap"


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class ReportRequest(BaseModel):
    """Request model for report generation"""
    report_type: ReportType
    symbol: str
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    format: ReportFormat = ReportFormat.PDF
    include_charts: bool = True
    include_tables: bool = True
    include_summary: bool = True
    include_detailed: bool = False
    include_recommendations: bool = True
    chart_types: List[ChartType] = []
    metrics: List[str] = []
    compare_to_previous: bool = False
    email_report: bool = False
    email_recipients: List[str] = []
    metadata: Dict[str, Any] = Field(default_factory=dict)


class ReportResponse(BaseModel):
    """Response model for report generation"""
    report_id: str
    report_type: ReportType
    symbol: str
    generated_at: datetime
    start_date: Optional[datetime]
    end_date: Optional[datetime]
    format: ReportFormat
    file_size: int
    summary: Dict[str, Any]
    metrics: Dict[str, Any]
    charts_count: int
    tables_count: int
    recommendations: List[str]


class ReportSchedule(BaseModel):
    """Schedule for automated reports"""
    report_type: ReportType
    symbol: str
    frequency: str  # daily, weekly, monthly
    format: ReportFormat = ReportFormat.PDF
    time: str = "00:00"
    day_of_week: Optional[int] = None
    day_of_month: Optional[int] = None
    recipients: List[str] = []
    enabled: bool = True
    last_generated: Optional[datetime] = None
    next_generation: Optional[datetime] = None


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ReportData:
    """Container for report data"""
    symbol: str
    report_type: ReportType
    start_date: datetime
    end_date: datetime
    positions: List[Any]
    trades: List[Any]
    orders: List[Any]
    performance: Dict[str, Any]
    risk_metrics: Dict[str, Any]
    liquidity_metrics: Dict[str, Any]
    inventory_metrics: Dict[str, Any]
    profitability_metrics: Dict[str, Any]
    summary: Dict[str, Any]
    recommendations: List[str]


@dataclass
class ChartData:
    """Container for chart data"""
    title: str
    chart_type: ChartType
    data: Dict[str, Any]
    x_label: Optional[str] = None
    y_label: Optional[str] = None
    legend: Optional[List[str]] = None


# =============================================================================
# MARKET MAKING REPORTER
# =============================================================================

class MarketMakingReporter:
    """
    Comprehensive Market Making Reporter with full API integration.
    
    Features:
    - Multiple report types
    - Various output formats
    - Interactive charts
    - Performance analysis
    - Risk reporting
    - Liquidity analysis
    - Inventory reporting
    - Profitability analysis
    - Executive summaries
    - Scheduled reports
    - Email delivery
    """

    def __init__(
        self,
        config: Optional[ReportingConfig] = None,
        market_making_config: Optional[MarketMakingConfig] = None,
        order_repo: Optional[OrderRepository] = None,
        trade_repo: Optional[TradeRepository] = None,
        position_repo: Optional[PositionRepository] = None,
        analytics: Optional[MarketMakingAnalytics] = None,
        market_maker: Optional[MarketMaker] = None
    ):
        """
        Initialize MarketMakingReporter.
        
        Args:
            config: Reporting configuration
            market_making_config: Market making configuration
            order_repo: Order repository
            trade_repo: Trade repository
            position_repo: Position repository
            analytics: Market making analytics
            market_maker: Market maker instance
        """
        self.config = config or ReportingConfig()
        self.mm_config = market_making_config or MarketMakingConfig()
        self.order_repo = order_repo or OrderRepository()
        self.trade_repo = trade_repo or TradeRepository()
        self.position_repo = position_repo or PositionRepository()
        self.analytics = analytics or MarketMakingAnalytics()
        self.market_maker = market_maker or MarketMaker()
        
        # Report storage
        self._reports: Dict[str, Dict[str, Any]] = {}
        self._schedules: Dict[str, ReportSchedule] = {}
        
        # Cache
        self._report_cache: Dict[str, Dict[str, Any]] = {}
        
        # Chart defaults
        self._chart_defaults = {
            'figsize': (12, 8),
            'dpi': 100,
            'style': 'seaborn-v0_8-darkgrid',
            'font_size': 12,
            'title_size': 16
        }
        
        # Set style
        plt.style.use(self._chart_defaults['style'])
        sns.set_palette("husl")
        
        logger.info("MarketMakingReporter initialized")

    # =========================================================================
    # Report Generation
    # =========================================================================

    @retry_async(max_attempts=3, delay=1.0)
    async def generate_report(
        self,
        request: ReportRequest
    ) -> ReportResponse:
        """
        Generate a report.
        
        Args:
            request: Report request
            
        Returns:
            ReportResponse: Generated report
        """
        try:
            # Validate request
            await self._validate_request(request)
            
            # Set date range
            start_date, end_date = self._get_date_range(request)
            
            # Collect report data
            report_data = await self._collect_report_data(
                request.symbol,
                request.report_type,
                start_date,
                end_date
            )
            
            # Generate report
            report_id = f"mm_report_{int(time.time() * 1000)}_{request.symbol}"
            
            # Generate content
            content, file_size = await self._generate_content(
                report_data,
                request.format,
                request
            )
            
            # Create response
            response = ReportResponse(
                report_id=report_id,
                report_type=request.report_type,
                symbol=request.symbol,
                generated_at=datetime.utcnow(),
                start_date=start_date,
                end_date=end_date,
                format=request.format,
                file_size=file_size,
                summary=report_data.summary,
                metrics={
                    'performance': report_data.performance,
                    'risk': report_data.risk_metrics,
                    'liquidity': report_data.liquidity_metrics,
                    'inventory': report_data.inventory_metrics,
                    'profitability': report_data.profitability_metrics
                },
                charts_count=len(request.chart_types) if request.include_charts else 0,
                tables_count=0,
                recommendations=report_data.recommendations
            )
            
            # Cache report
            self._reports[report_id] = {
                'response': response,
                'content': content,
                'data': report_data,
                'generated_at': datetime.utcnow()
            }
            
            # Send email if requested
            if request.email_report and request.email_recipients:
                await self._send_report_email(
                    response,
                    content,
                    request.email_recipients
                )
            
            logger.info(f"Report generated: {report_id} for {request.symbol}")
            return response
            
        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Report generation failed: {str(e)}"
            )

    async def _validate_request(self, request: ReportRequest) -> None:
        """Validate report request"""
        if request.start_date and request.end_date:
            if request.start_date > request.end_date:
                raise ValueError("Start date must be before end date")

    def _get_date_range(
        self,
        request: ReportRequest
    ) -> Tuple[datetime, datetime]:
        """Get date range for report"""
        end_date = request.end_date or datetime.utcnow()
        
        if request.start_date:
            start_date = request.start_date
        else:
            # Default based on report type
            if request.report_type == ReportType.DAILY:
                start_date = end_date - timedelta(days=1)
            elif request.report_type == ReportType.WEEKLY:
                start_date = end_date - timedelta(days=7)
            elif request.report_type == ReportType.MONTHLY:
                start_date = end_date - timedelta(days=30)
            else:
                start_date = end_date - timedelta(days=7)
        
        return start_date, end_date

    async def _collect_report_data(
        self,
        symbol: str,
        report_type: ReportType,
        start_date: datetime,
        end_date: datetime
    ) -> ReportData:
        """
        Collect data for report.
        
        Args:
            symbol: Symbol
            report_type: Type of report
            start_date: Start date
            end_date: End date
            
        Returns:
            ReportData: Collected data
        """
        # Get positions
        positions = await self.position_repo.get_by_symbol(symbol)
        
        # Get trades
        trades = await self.trade_repo.get_by_symbol(
            symbol,
            start_date=start_date,
            end_date=end_date
        )
        
        # Get orders
        orders = await self.order_repo.get_by_symbol(
            symbol,
            start_date=start_date,
            end_date=end_date
        )
        
        # Calculate metrics
        performance_metrics = await self._calculate_performance_metrics(trades)
        risk_metrics = await self._calculate_risk_metrics(positions, trades)
        liquidity_metrics = await self._calculate_liquidity_metrics(orders, trades)
        inventory_metrics = await self._calculate_inventory_metrics(positions)
        profitability_metrics = await self._calculate_profitability_metrics(trades)
        
        # Generate summary
        summary = await self._generate_summary(
            symbol,
            report_type,
            performance_metrics,
            risk_metrics,
            profitability_metrics
        )
        
        # Generate recommendations
        recommendations = await self._generate_recommendations(
            performance_metrics,
            risk_metrics,
            liquidity_metrics,
            inventory_metrics
        )
        
        return ReportData(
            symbol=symbol,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            positions=positions,
            trades=trades,
            orders=orders,
            performance=performance_metrics,
            risk_metrics=risk_metrics,
            liquidity_metrics=liquidity_metrics,
            inventory_metrics=inventory_metrics,
            profitability_metrics=profitability_metrics,
            summary=summary,
            recommendations=recommendations
        )

    # =========================================================================
    # Metrics Calculations
    # =========================================================================

    async def _calculate_performance_metrics(
        self,
        trades: List[Any]
    ) -> Dict[str, Any]:
        """Calculate performance metrics"""
        metrics = {}
        
        if not trades:
            return metrics
        
        total_trades = len(trades)
        winning_trades = [t for t in trades if t.pnl > 0]
        losing_trades = [t for t in trades if t.pnl < 0]
        
        metrics['total_trades'] = total_trades
        metrics['winning_trades'] = len(winning_trades)
        metrics['losing_trades'] = len(losing_trades)
        metrics['win_rate'] = len(winning_trades) / total_trades if total_trades > 0 else 0
        
        total_pnl = sum(t.pnl for t in trades)
        metrics['total_pnl'] = total_pnl
        metrics['avg_pnl'] = total_pnl / total_trades if total_trades > 0 else 0
        
        gross_profit = sum(t.pnl for t in winning_trades)
        gross_loss = abs(sum(t.pnl for t in losing_trades))
        metrics['gross_profit'] = gross_profit
        metrics['gross_loss'] = gross_loss
        metrics['profit_factor'] = gross_profit / gross_loss if gross_loss > 0 else float('inf')
        
        # Average trade duration
        durations = []
        for trade in trades:
            if hasattr(trade, 'created_at') and hasattr(trade, 'execution_time'):
                duration = (trade.execution_time - trade.created_at).total_seconds()
                durations.append(duration)
        
        metrics['avg_trade_duration'] = np.mean(durations) if durations else 0
        
        # Trade frequency
        if trades:
            first_trade = min(t.created_at for t in trades)
            last_trade = max(t.execution_time for t in trades)
            duration = (last_trade - first_trade).total_seconds()
            metrics['trade_frequency'] = total_trades / duration if duration > 0 else 0
        
        return metrics

    async def _calculate_risk_metrics(
        self,
        positions: List[Any],
        trades: List[Any]
    ) -> Dict[str, Any]:
        """Calculate risk metrics"""
        metrics = {}
        
        # Position risk
        if positions:
            total_position = sum(abs(p.size) for p in positions)
            metrics['total_position'] = total_position
            metrics['avg_position'] = total_position / len(positions) if positions else 0
            metrics['max_position'] = max(abs(p.size) for p in positions) if positions else 0
        
        # Drawdown
        if trades:
            equity = [0]
            for trade in trades:
                equity.append(equity[-1] + trade.pnl)
            
            peak = max(equity)
            current = equity[-1]
            metrics['current_drawdown'] = (peak - current) / peak if peak > 0 else 0
            
            # Max drawdown
            max_dd = 0
            peak = equity[0]
            for value in equity:
                if value > peak:
                    peak = value
                dd = (peak - value) / peak if peak > 0 else 0
                max_dd = max(max_dd, dd)
            metrics['max_drawdown'] = max_dd
        
        # Sharpe ratio
        if len(trades) > 1:
            returns = [t.pnl for t in trades]
            avg_return = np.mean(returns)
            std_return = np.std(returns)
            if std_return > 0:
                metrics['sharpe_ratio'] = avg_return / std_return * np.sqrt(252)
            else:
                metrics['sharpe_ratio'] = 0
        
        # Var
        if len(trades) > 30:
            returns = [t.pnl for t in trades]
            metrics['var_95'] = np.percentile(returns, 5)
            metrics['cvar_95'] = np.mean([r for r in returns if r <= metrics['var_95']])
        
        return metrics

    async def _calculate_liquidity_metrics(
        self,
        orders: List[Any],
        trades: List[Any]
    ) -> Dict[str, Any]:
        """Calculate liquidity metrics"""
        metrics = {}
        
        if orders:
            total_orders = len(orders)
            filled_orders = [o for o in orders if o.status == 'filled']
            
            metrics['total_orders'] = total_orders
            metrics['filled_orders'] = len(filled_orders)
            metrics['fill_rate'] = len(filled_orders) / total_orders if total_orders > 0 else 0
            
            # Order size
            sizes = [o.size for o in orders if hasattr(o, 'size')]
            metrics['avg_order_size'] = np.mean(sizes) if sizes else 0
            metrics['max_order_size'] = max(sizes) if sizes else 0
        
        if trades:
            # Trade size
            sizes = [t.size for t in trades if hasattr(t, 'size')]
            metrics['avg_trade_size'] = np.mean(sizes) if sizes else 0
            
            # Spread
            spreads = []
            for trade in trades:
                if hasattr(trade, 'spread'):
                    spreads.append(trade.spread)
            metrics['avg_spread'] = np.mean(spreads) if spreads else 0
            
            # Volume
            metrics['total_volume'] = sum(sizes) if sizes else 0
        
        return metrics

    async def _calculate_inventory_metrics(
        self,
        positions: List[Any]
    ) -> Dict[str, Any]:
        """Calculate inventory metrics"""
        metrics = {}
        
        if positions:
            # Current inventory
            total_inventory = sum(p.size for p in positions)
            metrics['current_inventory'] = total_inventory
            
            # Inventory value
            inventory_value = sum(p.size * p.entry_price for p in positions)
            metrics['inventory_value'] = inventory_value
            
            # Average entry price
            if total_inventory != 0:
                metrics['avg_entry_price'] = inventory_value / total_inventory
            else:
                metrics['avg_entry_price'] = 0
            
            # Position distribution
            long_positions = [p for p in positions if p.size > 0]
            short_positions = [p for p in positions if p.size < 0]
            metrics['long_positions'] = len(long_positions)
            metrics['short_positions'] = len(short_positions)
            
            # Unrealized PnL
            unrealized_pnl = 0
            for position in positions:
                if hasattr(position, 'current_price'):
                    pnl = (position.current_price - position.entry_price) * position.size
                    unrealized_pnl += pnl
            metrics['unrealized_pnl'] = unrealized_pnl
        
        return metrics

    async def _calculate_profitability_metrics(
        self,
        trades: List[Any]
    ) -> Dict[str, Any]:
        """Calculate profitability metrics"""
        metrics = {}
        
        if trades:
            # Total PnL
            total_pnl = sum(t.pnl for t in trades)
            metrics['total_pnl'] = total_pnl
            
            # Per trade
            metrics['avg_pnl_per_trade'] = total_pnl / len(trades) if trades else 0
            
            # Win/loss
            winning_trades = [t for t in trades if t.pnl > 0]
            losing_trades = [t for t in trades if t.pnl < 0]
            
            metrics['win_rate'] = len(winning_trades) / len(trades) if trades else 0
            metrics['avg_win'] = np.mean([t.pnl for t in winning_trades]) if winning_trades else 0
            metrics['avg_loss'] = abs(np.mean([t.pnl for t in losing_trades])) if losing_trades else 0
            
            if metrics['avg_loss'] > 0:
                metrics['risk_reward_ratio'] = metrics['avg_win'] / metrics['avg_loss']
            else:
                metrics['risk_reward_ratio'] = 0
            
            # Profit factor
            gross_profit = sum(t.pnl for t in winning_trades)
            gross_loss = abs(sum(t.pnl for t in losing_trades))
            metrics['profit_factor'] = gross_profit / gross_loss if gross_loss > 0 else float('inf')
            
            # Daily PnL
            if len(trades) > 0:
                # Group by date
                daily_pnl = {}
                for trade in trades:
                    date = trade.execution_time.date()
                    daily_pnl[date] = daily_pnl.get(date, 0) + trade.pnl
                
                metrics['avg_daily_pnl'] = np.mean(list(daily_pnl.values()))
                metrics['max_daily_pnl'] = max(daily_pnl.values())
                metrics['min_daily_pnl'] = min(daily_pnl.values())
        
        return metrics

    async def _generate_summary(
        self,
        symbol: str,
        report_type: ReportType,
        performance: Dict[str, Any],
        risk: Dict[str, Any],
        profitability: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Generate summary"""
        summary = {
            'symbol': symbol,
            'report_type': report_type.value,
            'period': f"{self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}",
            'generated_at': datetime.utcnow().isoformat()
        }
        
        # Add key metrics
        if performance:
            summary.update({
                'total_trades': performance.get('total_trades', 0),
                'win_rate': performance.get('win_rate', 0),
                'total_pnl': performance.get('total_pnl', 0),
                'profit_factor': performance.get('profit_factor', 0)
            })
        
        if risk:
            summary.update({
                'max_drawdown': risk.get('max_drawdown', 0),
                'sharpe_ratio': risk.get('sharpe_ratio', 0),
                'var_95': risk.get('var_95', 0)
            })
        
        if profitability:
            summary.update({
                'avg_pnl_per_trade': profitability.get('avg_pnl_per_trade', 0),
                'avg_daily_pnl': profitability.get('avg_daily_pnl', 0)
            })
        
        return summary

    async def _generate_recommendations(
        self,
        performance: Dict[str, Any],
        risk: Dict[str, Any],
        liquidity: Dict[str, Any],
        inventory: Dict[str, Any]
    ) -> List[str]:
        """Generate recommendations"""
        recommendations = []
        
        # Performance recommendations
        if performance.get('win_rate', 0) < 0.4:
            recommendations.append("Win rate is below 40%. Consider reviewing entry/exit criteria.")
        
        if performance.get('profit_factor', 0) < 1.0:
            recommendations.append("Profit factor is below 1.0. Strategy is not profitable.")
        
        # Risk recommendations
        if risk.get('max_drawdown', 0) > 0.2:
            recommendations.append("Max drawdown exceeds 20%. Consider adding risk controls.")
        
        if risk.get('sharpe_ratio', 0) < 0.5:
            recommendations.append("Sharpe ratio is below 0.5. Risk-adjusted returns need improvement.")
        
        # Liquidity recommendations
        if liquidity.get('fill_rate', 0) < 0.5:
            recommendations.append("Fill rate is below 50%. Consider adjusting order placement.")
        
        if liquidity.get('avg_spread', 0) > 0.01:
            recommendations.append("Average spread is high. Consider optimizing quote placement.")
        
        # Inventory recommendations
        if abs(inventory.get('current_inventory', 0)) > 100:
            recommendations.append("Large inventory position. Consider reducing exposure.")
        
        if len(recommendations) == 0:
            recommendations.append("All metrics are within acceptable ranges. Continue current strategy.")
        
        return recommendations

    # =========================================================================
    # Content Generation
    # =========================================================================

    async def _generate_content(
        self,
        report_data: ReportData,
        format: ReportFormat,
        request: ReportRequest
    ) -> Tuple[Any, int]:
        """
        Generate report content.
        
        Args:
            report_data: Report data
            format: Output format
            request: Original request
            
        Returns:
            Tuple[Any, int]: Content and file size
        """
        if format == ReportFormat.PDF:
            content = await self._generate_pdf(report_data, request)
        elif format == ReportFormat.HTML:
            content = await self._generate_html(report_data, request)
        elif format == ReportFormat.JSON:
            content = await self._generate_json(report_data, request)
        elif format == ReportFormat.CSV:
            content = await self._generate_csv(report_data, request)
        elif format == ReportFormat.EXCEL:
            content = await self._generate_excel(report_data, request)
        elif format == ReportFormat.MARKDOWN:
            content = await self._generate_markdown(report_data, request)
        else:
            content = await self._generate_json(report_data, request)
        
        file_size = len(content) if isinstance(content, (bytes, str)) else 0
        return content, file_size

    # =========================================================================
    # PDF Generation
    # =========================================================================

    async def _generate_pdf(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate PDF report"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image, KeepTogether
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(
                f"Market Making Report - {report_data.symbol}",
                styles['Title']
            ))
            story.append(Spacer(1, 0.25*inch))
            
            # Info
            story.append(Paragraph(
                f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}",
                styles['Normal']
            ))
            story.append(Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
                styles['Normal']
            ))
            story.append(Spacer(1, 0.5*inch))
            
            # Summary
            story.append(Paragraph("Executive Summary", styles['Heading1']))
            story.append(Spacer(1, 0.1*inch))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Trades', str(report_data.summary.get('total_trades', 0))],
                ['Win Rate', f"{report_data.summary.get('win_rate', 0)*100:.1f}%"],
                ['Total PnL', f"${report_data.summary.get('total_pnl', 0):.2f}"],
                ['Profit Factor', f"{report_data.summary.get('profit_factor', 0):.2f}"],
                ['Max Drawdown', f"{report_data.summary.get('max_drawdown', 0)*100:.1f}%"],
                ['Sharpe Ratio', f"{report_data.summary.get('sharpe_ratio', 0):.2f}"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.25*inch))
            
            # Charts
            if request.include_charts and request.chart_types:
                story.append(PageBreak())
                story.append(Paragraph("Charts", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            img = Image(chart_image, width=6*inch, height=4*inch)
                            story.append(KeepTogether([
                                Paragraph(f"<b>{chart_data.title}</b>", styles['Heading2']),
                                img,
                                Spacer(1, 0.1*inch)
                            ]))
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                story.append(PageBreak())
                story.append(Paragraph("Recommendations", styles['Heading1']))
                story.append(Spacer(1, 0.1*inch))
                
                for rec in report_data.recommendations:
                    story.append(Paragraph(f"• {rec}", styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
            
            doc.build(story)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return b""

    # =========================================================================
    # HTML Generation
    # =========================================================================

    async def _generate_html(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate HTML report"""
        try:
            html = []
            
            # Header
            html.append("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Market Making Report</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
                    h2 { color: #34495e; margin-top: 30px; }
                    .summary-box { background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 10px 0; }
                    .metric { display: inline-block; margin: 10px 20px 10px 0; }
                    .metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }
                    .metric-label { font-size: 14px; color: #7f8c8d; }
                    .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                    .table th { background: #3498db; color: white; padding: 10px; text-align: left; }
                    .table td { padding: 10px; border-bottom: 1px solid #ddd; }
                    .table tr:hover { background: #f5f5f5; }
                    .recommendation { background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }
                    .chart-container { margin: 20px 0; text-align: center; }
                    img { max-width: 100%; border: 1px solid #ddd; border-radius: 5px; }
                </style>
            </head>
            <body>
            """)
            
            # Title
            html.append(f"<h1>Market Making Report - {report_data.symbol}</h1>")
            html.append(f"<p><strong>Period:</strong> {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}</p>")
            html.append(f"<p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>")
            
            # Summary
            html.append("<h2>Executive Summary</h2>")
            html.append('<div class="summary-box">')
            
            summary = report_data.summary
            html.append(f"""
            <div class="metric">
                <div class="metric-value">{summary.get('total_trades', 0)}</div>
                <div class="metric-label">Total Trades</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('win_rate', 0)*100:.1f}%</div>
                <div class="metric-label">Win Rate</div>
            </div>
            <div class="metric">
                <div class="metric-value">${summary.get('total_pnl', 0):.2f}</div>
                <div class="metric-label">Total PnL</div>
            </div>
            <div class="metric">
                <div class="metric-value">{summary.get('sharpe_ratio', 0):.2f}</div>
                <div class="metric-label">Sharpe Ratio</div>
            </div>
            """)
            html.append('</div>')
            
            # Charts
            if request.include_charts and request.chart_types:
                html.append("<h2>Charts</h2>")
                for chart_type in request.chart_types:
                    chart_data = await self._prepare_chart_data(report_data, chart_type)
                    if chart_data:
                        chart_image = await self._create_chart_image(chart_data)
                        if chart_image:
                            import base64
                            img_base64 = base64.b64encode(chart_image).decode('utf-8')
                            html.append(f"""
                            <div class="chart-container">
                                <h3>{chart_data.title}</h3>
                                <img src="data:image/png;base64,{img_base64}" />
                            </div>
                            """)
            
            # Recommendations
            if request.include_recommendations and report_data.recommendations:
                html.append("<h2>Recommendations</h2>")
                for rec in report_data.recommendations:
                    html.append(f"""
                    <div class="recommendation">
                        <p>{rec}</p>
                    </div>
                    """)
            
            html.append("""
            <hr/>
            <p style="color: #7f8c8d; font-size: 12px;">
                Generated by NEXUS Market Making Reporter
            </p>
            </body>
            </html>
            """)
            
            return ''.join(html)
            
        except Exception as e:
            logger.error(f"Error generating HTML: {e}")
            return ""

    # =========================================================================
    # JSON Generation
    # =========================================================================

    async def _generate_json(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate JSON report"""
        try:
            report = {
                'symbol': report_data.symbol,
                'report_type': report_data.report_type.value,
                'start_date': report_data.start_date.isoformat(),
                'end_date': report_data.end_date.isoformat(),
                'generated_at': datetime.utcnow().isoformat(),
                'summary': report_data.summary,
                'performance': report_data.performance,
                'risk': report_data.risk_metrics,
                'liquidity': report_data.liquidity_metrics,
                'inventory': report_data.inventory_metrics,
                'profitability': report_data.profitability_metrics,
                'recommendations': report_data.recommendations
            }
            
            return json.dumps(report, indent=2, default=str)
            
        except Exception as e:
            logger.error(f"Error generating JSON: {e}")
            return "{}"

    # =========================================================================
    # CSV Generation
    # =========================================================================

    async def _generate_csv(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate CSV report"""
        try:
            output = []
            
            # Trades
            if report_data.trades:
                output.append("=== Trades ===")
                output.append("timestamp,symbol,side,size,price,pnl")
                for trade in report_data.trades:
                    output.append(f"{trade.execution_time},{trade.symbol},{trade.side},{trade.size},{trade.price},{trade.pnl}")
            
            # Performance metrics
            output.append("\n=== Performance Metrics ===")
            for key, value in report_data.performance.items():
                output.append(f"{key},{value}")
            
            # Risk metrics
            output.append("\n=== Risk Metrics ===")
            for key, value in report_data.risk_metrics.items():
                output.append(f"{key},{value}")
            
            # Liquidity metrics
            output.append("\n=== Liquidity Metrics ===")
            for key, value in report_data.liquidity_metrics.items():
                output.append(f"{key},{value}")
            
            # Inventory metrics
            output.append("\n=== Inventory Metrics ===")
            for key, value in report_data.inventory_metrics.items():
                output.append(f"{key},{value}")
            
            # Profitability metrics
            output.append("\n=== Profitability Metrics ===")
            for key, value in report_data.profitability_metrics.items():
                output.append(f"{key},{value}")
            
            # Recommendations
            output.append("\n=== Recommendations ===")
            for i, rec in enumerate(report_data.recommendations, 1):
                output.append(f"{i},{rec}")
            
            return '\n'.join(output)
            
        except Exception as e:
            logger.error(f"Error generating CSV: {e}")
            return ""

    # =========================================================================
    # Excel Generation
    # =========================================================================

    async def _generate_excel(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> bytes:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            
            ws['A1'] = f"Market Making Report - {report_data.symbol}"
            ws['A1'].font = Font(size=16, bold=True)
            
            ws['A3'] = f"Period: {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}"
            ws['A4'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            
            row = 6
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for key, value in report_data.summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            # Trades sheet
            if report_data.trades:
                ws_trades = wb.create_sheet("Trades")
                headers = ['Timestamp', 'Symbol', 'Side', 'Size', 'Price', 'PnL']
                for col, header in enumerate(headers, 1):
                    ws_trades.cell(row=1, column=col, value=header)
                    ws_trades.cell(row=1, column=col).font = Font(bold=True)
                
                for row, trade in enumerate(report_data.trades, 2):
                    ws_trades.cell(row=row, column=1, value=trade.execution_time)
                    ws_trades.cell(row=row, column=2, value=trade.symbol)
                    ws_trades.cell(row=row, column=3, value=trade.side)
                    ws_trades.cell(row=row, column=4, value=trade.size)
                    ws_trades.cell(row=row, column=5, value=trade.price)
                    ws_trades.cell(row=row, column=6, value=trade.pnl)
            
            # Metrics sheets
            metrics_sheets = [
                ('Performance', report_data.performance),
                ('Risk', report_data.risk_metrics),
                ('Liquidity', report_data.liquidity_metrics),
                ('Inventory', report_data.inventory_metrics),
                ('Profitability', report_data.profitability_metrics)
            ]
            
            for sheet_name, metrics in metrics_sheets:
                if metrics:
                    ws_metrics = wb.create_sheet(sheet_name)
                    ws_metrics.cell(row=1, column=1, value="Metric")
                    ws_metrics.cell(row=1, column=2, value="Value")
                    ws_metrics.cell(row=1, column=1).font = Font(bold=True)
                    ws_metrics.cell(row=1, column=2).font = Font(bold=True)
                    
                    row = 2
                    for key, value in metrics.items():
                        ws_metrics.cell(row=row, column=1, value=key.replace('_', ' ').title())
                        ws_metrics.cell(row=row, column=2, value=value)
                        row += 1
            
            # Recommendations sheet
            if report_data.recommendations:
                ws_rec = wb.create_sheet("Recommendations")
                ws_rec.cell(row=1, column=1, value="#")
                ws_rec.cell(row=1, column=2, value="Recommendation")
                ws_rec.cell(row=1, column=1).font = Font(bold=True)
                ws_rec.cell(row=1, column=2).font = Font(bold=True)
                
                for i, rec in enumerate(report_data.recommendations, 2):
                    ws_rec.cell(row=i, column=1, value=i-1)
                    ws_rec.cell(row=i, column=2, value=rec)
            
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return b""

    # =========================================================================
    # Markdown Generation
    # =========================================================================

    async def _generate_markdown(
        self,
        report_data: ReportData,
        request: ReportRequest
    ) -> str:
        """Generate Markdown report"""
        try:
            md = []
            
            md.append(f"# Market Making Report - {report_data.symbol}\n")
            md.append(f"**Period:** {report_data.start_date.strftime('%Y-%m-%d')} to {report_data.end_date.strftime('%Y-%m-%d')}")
            md.append(f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n")
            
            # Summary
            md.append("## Executive Summary\n")
            md.append("| Metric | Value |")
            md.append("|--------|-------|")
            
            for key, value in report_data.summary.items():
                if isinstance(value, float):
                    if key in ['win_rate', 'max_drawdown']:
                        value = f"{value*100:.1f}%"
                    elif key == 'total_pnl':
                        value = f"${value:.2f}"
                    else:
                        value = f"{value:.2f}"
                md.append(f"| {key.replace('_', ' ').title()} | {value} |")
            
            md.append("")
            
            # Recommendations
            if report_data.recommendations:
                md.append("## Recommendations\n")
                for rec in report_data.recommendations:
                    md.append(f"- {rec}")
                md.append("")
            
            return '\n'.join(md)
            
        except Exception as e:
            logger.error(f"Error generating Markdown: {e}")
            return ""

    # =========================================================================
    # Chart Generation
    # =========================================================================

    async def _prepare_chart_data(
        self,
        report_data: ReportData,
        chart_type: ChartType
    ) -> Optional[ChartData]:
        """Prepare chart data"""
        try:
            if chart_type == ChartType.EQUITY:
                if report_data.trades:
                    equity = [0]
                    timestamps = [report_data.start_date]
                    
                    for trade in report_data.trades:
                        equity.append(equity[-1] + trade.pnl)
                        if hasattr(trade, 'execution_time'):
                            timestamps.append(trade.execution_time)
                    
                    return ChartData(
                        title="Equity Curve",
                        chart_type=ChartType.EQUITY,
                        data={
                            'timestamps': timestamps,
                            'equity': equity
                        },
                        x_label="Date",
                        y_label="Equity ($)"
                    )
            
            elif chart_type == ChartType.DRAWDOWN:
                if report_data.trades:
                    equity = [0]
                    for trade in report_data.trades:
                        equity.append(equity[-1] + trade.pnl)
                    
                    peak = max(equity)
                    drawdowns = [(peak - e) / peak * 100 if peak > 0 else 0 for e in equity]
                    
                    return ChartData(
                        title="Drawdown",
                        chart_type=ChartType.DRAWDOWN,
                        data={'drawdowns': drawdowns},
                        x_label="Trade Number",
                        y_label="Drawdown (%)"
                    )
            
            elif chart_type == ChartType.PNL:
                if report_data.trades:
                    pnls = [t.pnl for t in report_data.trades]
                    
                    return ChartData(
                        title="PnL Distribution",
                        chart_type=ChartType.PNL,
                        data={'pnls': pnls},
                        x_label="PnL ($)",
                        y_label="Frequency"
                    )
            
            return None
            
        except Exception as e:
            logger.error(f"Error preparing chart data: {e}")
            return None

    async def _create_chart_image(self, chart_data: ChartData) -> Optional[bytes]:
        """Create chart image"""
        try:
            fig, ax = plt.subplots(figsize=(12, 8))
            
            if chart_data.chart_type == ChartType.EQUITY:
                timestamps = chart_data.data.get('timestamps', [])
                equity = chart_data.data.get('equity', [])
                
                ax.plot(timestamps, equity, linewidth=2, color='#3498db')
                ax.fill_between(timestamps, equity, alpha=0.3, color='#3498db')
                ax.axhline(y=0, color='#e74c3c', linestyle='--', alpha=0.5)
            
            elif chart_data.chart_type == ChartType.DRAWDOWN:
                drawdowns = chart_data.data.get('drawdowns', [])
                ax.fill_between(range(len(drawdowns)), drawdowns, alpha=0.5, color='#e74c3c')
                ax.plot(range(len(drawdowns)), drawdowns, linewidth=2, color='#e74c3c')
                ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)
            
            elif chart_data.chart_type == ChartType.PNL:
                pnls = chart_data.data.get('pnls', [])
                ax.hist(pnls, bins=30, color='#3498db', edgecolor='black', alpha=0.7)
                ax.axvline(np.mean(pnls), color='#e74c3c', linestyle='dashed', linewidth=2, label='Mean')
                ax.axvline(np.median(pnls), color='#2ecc71', linestyle='dashed', linewidth=2, label='Median')
                ax.legend()
            
            ax.set_xlabel(chart_data.x_label or '')
            ax.set_ylabel(chart_data.y_label or '')
            ax.set_title(chart_data.title, fontsize=16, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            buffer = BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating chart image: {e}")
            return None

    # =========================================================================
    # Report Management
    # =========================================================================

    async def get_report(self, report_id: str) -> Optional[Dict[str, Any]]:
        """Get cached report"""
        return self._reports.get(report_id)

    async def delete_report(self, report_id: str) -> bool:
        """Delete cached report"""
        if report_id in self._reports:
            del self._reports[report_id]
            return True
        return False

    async def list_reports(
        self,
        symbol: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """List available reports"""
        reports = []
        
        for report_id, data in self._reports.items():
            response = data.get('response')
            if response:
                if symbol and response.symbol != symbol:
                    continue
                reports.append({
                    'report_id': report_id,
                    'report_type': response.report_type.value,
                    'symbol': response.symbol,
                    'generated_at': response.generated_at,
                    'format': response.format.value,
                    'file_size': response.file_size
                })
        
        reports.sort(key=lambda x: x['generated_at'], reverse=True)
        return reports[:limit]

    # =========================================================================
    # Scheduled Reports
    # =========================================================================

    async def create_schedule(self, schedule: ReportSchedule) -> bool:
        """Create a report schedule"""
        try:
            schedule.next_generation = self._calculate_next_generation(schedule)
            self._schedules[schedule.symbol] = schedule
            logger.info(f"Created schedule for {schedule.symbol}")
            return True
        except Exception as e:
            logger.error(f"Error creating schedule: {e}")
            return False

    async def get_schedule(self, symbol: str) -> Optional[ReportSchedule]:
        """Get a report schedule"""
        return self._schedules.get(symbol)

    async def delete_schedule(self, symbol: str) -> bool:
        """Delete a report schedule"""
        if symbol in self._schedules:
            del self._schedules[symbol]
            return True
        return False

    def _calculate_next_generation(self, schedule: ReportSchedule) -> datetime:
        """Calculate next generation time"""
        now = datetime.utcnow()
        
        if schedule.frequency == 'daily':
            next_time = now.replace(hour=int(schedule.time.split(':')[0]),
                                   minute=int(schedule.time.split(':')[1]),
                                   second=0, microsecond=0)
            if next_time <= now:
                next_time += timedelta(days=1)
            return next_time
        
        elif schedule.frequency == 'weekly' and schedule.day_of_week is not None:
            days_until = schedule.day_of_week - now.weekday()
            if days_until <= 0:
                days_until += 7
            next_time = now + timedelta(days=days_until)
            next_time = next_time.replace(hour=int(schedule.time.split(':')[0]),
                                         minute=int(schedule.time.split(':')[1]),
                                         second=0, microsecond=0)
            return next_time
        
        return now + timedelta(days=1)

    async def _send_report_email(
        self,
        response: ReportResponse,
        content: Any,
        recipients: List[str]
    ) -> None:
        """Send report via email"""
        try:
            logger.info(f"Sending report {response.report_id} to {recipients}")
        except Exception as e:
            logger.error(f"Error sending report email: {e}")

    # =========================================================================
    # Cleanup
    # =========================================================================

    async def close(self) -> None:
        """Close the reporter"""
        self._reports.clear()
        self._schedules.clear()
        self._report_cache.clear()
        logger.info("MarketMakingReporter closed")


# =============================================================================
# FASTAPI ROUTER
# =============================================================================

from fastapi import APIRouter, Depends, Query, Body

router = APIRouter(prefix="/api/v1/market-making/reports", tags=["Market Making Reports"])


async def get_reporter() -> MarketMakingReporter:
    """Dependency to get MarketMakingReporter instance"""
    return MarketMakingReporter()


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Generate a report"""
    return await reporter.generate_report(request)


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Get a generated report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return report


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Delete a report"""
    success = await reporter.delete_report(report_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    return {"success": True}


@router.get("/")
async def list_reports(
    symbol: Optional[str] = Query(None),
    limit: int = Query(50, le=500),
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """List available reports"""
    return await reporter.list_reports(symbol, limit)


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Download a report"""
    report = await reporter.get_report(report_id)
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found"
        )
    
    response_data = report.get('response')
    content = report.get('content')
    
    if not content:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report content not found"
        )
    
    content_type = {
        'pdf': 'application/pdf',
        'html': 'text/html',
        'json': 'application/json',
        'csv': 'text/csv',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'markdown': 'text/markdown'
    }.get(response_data.format.value, 'application/octet-stream')
    
    return Response(
        content=content if isinstance(content, bytes) else content.encode('utf-8'),
        media_type=content_type,
        headers={
            'Content-Disposition': f'attachment; filename="mm_report_{report_id}.{response_data.format.value}"'
        }
    )


@router.post("/schedule")
async def create_schedule(
    schedule: ReportSchedule,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Create a report schedule"""
    success = await reporter.create_schedule(schedule)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create schedule"
        )
    return {"success": True}


@router.get("/schedule/{symbol}")
async def get_schedule(
    symbol: str,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Get a report schedule"""
    schedule = await reporter.get_schedule(symbol)
    if not schedule:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {symbol} not found"
        )
    return schedule


@router.delete("/schedule/{symbol}")
async def delete_schedule(
    symbol: str,
    reporter: MarketMakingReporter = Depends(get_reporter)
):
    """Delete a report schedule"""
    success = await reporter.delete_schedule(symbol)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Schedule for {symbol} not found"
        )
    return {"success": True}


@router.get("/types")
async def get_report_types():
    """Get available report types"""
    return {
        'types': [
            {'name': t.value, 'description': t.name.replace('_', ' ').title()}
            for t in ReportType
        ]
    }


@router.get("/formats")
async def get_report_formats():
    """Get available report formats"""
    return {
        'formats': [
            {'name': f.value, 'description': f.name.upper()}
            for f in ReportFormat
        ]
    }


@router.get("/chart-types")
async def get_chart_types():
    """Get available chart types"""
    return {
        'chart_types': [
            {'name': c.value, 'description': c.name.replace('_', ' ').title()}
            for c in ChartType
        ]
    }


# =============================================================================
# EXPORTS
# =============================================================================

__all__ = [
    'MarketMakingReporter',
    'ReportFormat',
    'ReportType',
    'ChartType',
    'ReportRequest',
    'ReportResponse',
    'ReportSchedule',
    'ReportData',
    'ChartData',
    'router'
]
